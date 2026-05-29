from io import BytesIO
from datetime import date
import json
import os
import sys
from tempfile import NamedTemporaryFile
import types
import uuid

from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


def _build_roster_file(rows: list[tuple[str, str]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("student_code", "full_name"))
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_exam_file(rows: list[tuple[str, str, float, str, str]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("student_code", "full_name", "score", "note", "teacher_comment"))
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_notescc_exam_file(rows: list[tuple[str, str, float]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "NotesCC"
    for idx, (code, name, score) in enumerate(rows, start=18):
        sheet.cell(idx, 3).value = code
        sheet.cell(idx, 4).value = name
        sheet.cell(idx, 7).value = score
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_tiny_png() -> bytes:
    return (
        b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
        b"\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDAT\x08\xd7c\xf8\xcf"
        b"\xc0\x00\x00\x03\x01\x01\x00\xc9\xfe\x92\xef\x00\x00\x00\x00IEND\xaeB`\x82"
    )


def _build_notescc_list_file(rows: list[tuple[str, str, str, str]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "NotesCC"
    sheet.cell(16, 2).value = "ID"
    sheet.cell(16, 3).value = "student_code"
    sheet.cell(16, 4).value = "name"
    sheet.cell(16, 6).value = "birth_date"
    for idx, (external_id, code, name, birth_date) in enumerate(rows, start=18):
        sheet.cell(idx, 2).value = external_id
        sheet.cell(idx, 3).value = code
        sheet.cell(idx, 4).value = name
        sheet.cell(idx, 6).value = birth_date
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_exam_list_file(rows: list[tuple[str, str, str, float, float, float, float]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("id", "name", "birth_date", "note_1", "note_2", "note_3", "note"))
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_holiday_file(headers: tuple[str, ...], rows: list[tuple]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "holiday_import"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_pdf_file(lines: list[str]) -> bytes:
    output = BytesIO()
    pdf = canvas.Canvas(output, pagesize=A4)
    _, page_height = A4
    y = page_height - 48
    for line in lines:
        pdf.drawString(42, y, line)
        y -= 18
        if y < 48:
            pdf.showPage()
            y = page_height - 48
    pdf.save()
    return output.getvalue()


def _build_timetable_csv(rows: list[tuple[str, str, str, str, str, str, str, str]]) -> bytes:
    lines = ["teacher_key,class_name,subject,weekday,start_time,end_time,room,group"]
    for row in rows:
        lines.append(",".join(row))
    return ("\n".join(lines) + "\n").encode("utf-8")


def _build_timetable_ics(events: list[dict]) -> bytes:
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Teacher Progress//Tests//EN",
    ]
    for event in events:
        lines.extend(
            [
                "BEGIN:VEVENT",
                f"SUMMARY:{event.get('summary', '')}",
                f"DTSTART:{event.get('dtstart', '')}",
                f"DTEND:{event.get('dtend', '')}",
            ]
        )
        if event.get("rrule"):
            lines.append(f"RRULE:{event.get('rrule')}")
        if event.get("location"):
            lines.append(f"LOCATION:{event.get('location')}")
        if event.get("description"):
            lines.append(f"DESCRIPTION:{event.get('description')}")
        lines.append("END:VEVENT")
    lines.append("END:VCALENDAR")
    return ("\r\n".join(lines) + "\r\n").encode("utf-8")


def _flatten_checklist(nodes: list[dict]) -> list[dict]:
    output: list[dict] = []
    for node in nodes:
        output.append(node)
        output.extend(_flatten_checklist(node.get("children", [])))
    return output


def _first_leaf_checklist_item(nodes: list[dict]) -> dict | None:
    flat = _flatten_checklist(nodes)
    if not flat:
        return None
    parent_ids = {
        int(row["parent_item_id"])
        for row in flat
        if isinstance(row, dict) and row.get("parent_item_id") is not None
    }
    return next(
        (
            row
            for row in flat
            if isinstance(row, dict) and int(row.get("id") or 0) > 0 and int(row["id"]) not in parent_ids
        ),
        None,
    )


def _auth_headers(client) -> dict[str, str]:
    owner_payload = {"email": "owner@app.local", "password": "OwnerPass123", "full_name": "Owner"}
    bootstrap_resp = client.post("/auth/bootstrap-owner", json=owner_payload)
    assert bootstrap_resp.status_code in (201, 400)
    login_resp = client.post("/auth/login", json={"email": owner_payload["email"], "password": owner_payload["password"]})
    assert login_resp.status_code == 200
    token = login_resp.json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _login_headers(client, email: str, password: str) -> dict[str, str]:
    login_resp = client.post("/auth/login", json={"email": email, "password": password})
    assert login_resp.status_code == 200
    return {"Authorization": f"Bearer {login_resp.json()['access_token']}"}


def _create_teacher_and_login(client, owner_headers: dict[str, str]) -> tuple[int, dict[str, str]]:
    email = f"teacher_{uuid.uuid4().hex[:8]}@app.local"
    password = "TeacherPass123"
    create_resp = client.post(
        "/auth/users",
        headers=owner_headers,
        json={"email": email, "password": password, "full_name": "Teacher", "role": "teacher"},
    )
    assert create_resp.status_code == 201
    teacher_id = create_resp.json()["id"]
    login_resp = client.post("/auth/login", json={"email": email, "password": password})
    assert login_resp.status_code == 200
    return teacher_id, {"Authorization": f"Bearer {login_resp.json()['access_token']}"}


def _close_any_active_unit(client, headers: dict[str, str]) -> None:
    classes_resp = client.get("/classes", headers=headers)
    assert classes_resp.status_code == 200
    for row in classes_resp.json():
        if not isinstance(row, dict) or not row.get("id"):
            continue
        class_id = int(row["id"])
        workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
        assert workspace_resp.status_code == 200
        active_unit = workspace_resp.json().get("active_unit")
        if not isinstance(active_unit, dict) or not active_unit.get("id"):
            continue
        close_resp = client.post(
            f"/workflow/classes/{class_id}/units/{active_unit['id']}/close",
            headers=headers,
        )
        if close_resp.status_code == 200:
            continue
        active_session = workspace_resp.json().get("active_session")
        if (
            close_resp.status_code == 409
            and isinstance(active_session, dict)
            and active_session.get("id")
        ):
            end_payload = {}
            if active_session.get("session_date"):
                end_payload["session_date"] = active_session.get("session_date")
            if active_session.get("start_time"):
                end_payload["start_time"] = active_session.get("start_time")
                end_payload["end_time"] = active_session.get("start_time")
            end_resp = client.post(
                f"/workflow/classes/{class_id}/sessions/{active_session['id']}/end",
                headers=headers,
                json=end_payload,
            )
            assert end_resp.status_code == 200
            retry_close_resp = client.post(
                f"/workflow/classes/{class_id}/units/{active_unit['id']}/close",
                headers=headers,
            )
            assert retry_close_resp.status_code == 200
            continue
        assert close_resp.status_code == 200


def test_heuristic_extract_supports_compact_numbered_and_french_keywords():
    from app.services.extraction import _heuristic_extract

    raw_text = (
        "Chapitre2: Calcul litteral - Identites remarquables\n"
        "1.Developpement par la distributivite\n"
        "Propriete: Developper un produit.\n"
        "Exemples: Calculs de demonstration.\n"
        "Application: Developper et simplifier.\n"
        "2.Developpement a l'aide des identites remarquables\n"
        "Exercice 1: Resoudre l'equation.\n"
    )
    extracted = _heuristic_extract(raw_text)
    assert any(item.startswith("Chapitre2") for item in extracted["lesson_headings"])
    assert "1.Developpement par la distributivite" in extracted["lesson_headings"]
    assert any("2.Developpement" in item for item in extracted["lesson_headings"])
    assert len(extracted["activities"]) >= 2
    assert len(extracted["exercises"]) >= 1


def test_extract_structured_progress_falls_back_when_openai_returns_empty(monkeypatch):
    from app.services import extraction

    def fake_openai(raw_text, image_path=None):
        return (
            {
                "confidence": 0.92,
                "lesson_headings": [],
                "activities": [],
                "exercises": [],
                "raw_text": raw_text,
                "provider": "openai",
                "model": "gpt-fake",
                "fallback_reason": None,
            },
            None,
        )

    monkeypatch.setattr(extraction, "_openai_extract", fake_openai)
    parsed = extraction.extract_structured_progress(
        "1.Developpement par la distributivite\nExercice 1: factoriser"
    )
    assert parsed["provider"] == "heuristic"
    assert parsed["fallback_reason"] == "openai_empty_result"
    assert parsed["lesson_headings"]
    assert parsed["exercises"]


def test_resolve_raw_text_uses_best_ocr_candidate(monkeypatch):
    from PIL import Image
    from app.services import extraction

    with NamedTemporaryFile(suffix=".png", delete=False) as temp:
        temp_path = temp.name
    try:
        image = Image.new("RGB", (120, 40), color="white")
        image.save(temp_path)

        fake_module = types.SimpleNamespace()
        fake_module.Output = types.SimpleNamespace(DICT="dict")

        def fake_image_to_string(_image, lang=None, config=""):
            if config == "--psm 4":
                return "1.Developpement par la distributivite\nExercice 1: factoriser"
            return "blurry noise"

        def fake_image_to_data(_image, lang=None, config="", output_type=None):
            if config == "--psm 4":
                return {"conf": ["88", "84"], "text": ["1.Developpement", "Exercice"]}
            return {"conf": ["12"], "text": ["noise"]}

        fake_module.image_to_string = fake_image_to_string
        fake_module.image_to_data = fake_image_to_data
        monkeypatch.setitem(sys.modules, "pytesseract", fake_module)

        text = extraction.resolve_raw_text(temp_path, None)
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)

    assert "1.Developpement" in text
    assert "Exercice 1" in text


def test_sanitize_result_reclassifies_and_cleans_lines():
    from app.services.extraction import _sanitize_result

    parsed = _sanitize_result(
        {
            "confidence": 0.8,
            "lesson_headings": ["  1.2 Equations  ", "• Exercice 2 page 44"],
            "activities": ["\u25aa Propriete: distributivite"],
            "exercises": [],
        },
        "raw",
        provider="openai",
        model="gpt-fake",
    )
    assert "1.2 Equations" in parsed["lesson_headings"]
    assert any("Exercice 2 page 44" in row for row in parsed["exercises"])
    assert any("Propriete: distributivite" in row for row in parsed["activities"])


def test_extract_structured_progress_augments_openai_with_heuristic(monkeypatch):
    from app.services import extraction

    def fake_openai(raw_text, image_path=None):
        return (
            {
                "confidence": 0.61,
                "lesson_headings": ["Chapitre 2: Calcul litteral"],
                "activities": [],
                "exercises": [],
                "raw_text": raw_text,
                "provider": "openai",
                "model": "gpt-fake",
                "fallback_reason": None,
            },
            None,
        )

    monkeypatch.setattr(extraction, "_openai_extract", fake_openai)
    raw_text = (
        "1.1 Developpement par la distributivite\n"
        "Propriete: k(a+b)=ka+kb\n"
        "Exercice 1: simplifier\n"
    )
    parsed = extraction.extract_structured_progress(raw_text)
    assert parsed["provider"] == "openai"
    assert any("1.1" in row for row in parsed["lesson_headings"])
    assert any("Propriete" in row for row in parsed["activities"])
    assert any("Exercice 1" in row for row in parsed["exercises"])
    assert parsed["fallback_reason"] == "openai_augmented_with_heuristic"


def test_health(client):
    response = client.get("/health")
    assert response.status_code == 200
    assert response.json()["status"] == "ok"


def test_workflow_unit_session_lifecycle(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Workflow 2APIC", "subject": "Math", "level": "2APIC"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    roster = _build_roster_file([("STD001", "Alice"), ("STD002", "Bob")])
    import_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200
    students = client.get(f"/classes/{class_id}/students", headers=headers).json()
    assert len(students) == 2

    source_text_lines = [
        "Chapitre 1: Calcul litteral",
        "1.1 Developpement",
        "Propriete: distributivite",
        "Exemples: simplifier",
        "2. Identites remarquables",
        "Definition: (a+b)^2",
    ]
    start_unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Chapitre 1",
            "planned_hours": "6",
            "source_text": "\n".join(source_text_lines),
        },
    )
    assert start_unit_resp.status_code == 201
    unit = start_unit_resp.json()
    assert unit["status"] == "active"
    assert unit["unit_type"] == "chapter"
    assert unit["progress_total"] >= 1

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_resp.status_code == 200
    workspace = workspace_resp.json()
    assert workspace["active_unit"]["id"] == unit["id"]
    assert workspace["active_session"] is None

    parent_item_id = workspace["active_unit"]["checklist"][0]["id"]
    add_item_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items",
        headers=headers,
        json={"title": "3.1 Revision guidee", "item_kind": "exercise", "parent_item_id": parent_item_id},
    )
    assert add_item_resp.status_code == 201
    added_item = add_item_resp.json()
    assert added_item["unit_id"] == unit["id"]
    assert added_item["title"] == "3.1 Revision guidee"
    assert added_item["item_kind"] == "exercise"
    assert added_item["parent_item_id"] == parent_item_id

    update_item_resp = client.put(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items/{added_item['id']}",
        headers=headers,
        json={"title": "3.1 Revision guidee (updated)", "item_kind": "example"},
    )
    assert update_item_resp.status_code == 200
    updated_item = update_item_resp.json()
    assert updated_item["title"] == "3.1 Revision guidee (updated)"
    assert updated_item["item_kind"] == "example"

    delete_item_resp = client.delete(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items/{added_item['id']}",
        headers=headers,
    )
    assert delete_item_resp.status_code == 200
    assert delete_item_resp.json()["deleted"] is True

    absent_ids = [students[0]["id"]]
    start_session_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/start",
        headers=headers,
        json={"absent_student_ids": absent_ids},
    )
    assert start_session_resp.status_code == 201
    session = start_session_resp.json()
    session_id = session["id"]
    assert session["absent_count"] == 1
    assert sorted(session["absent_student_ids"]) == sorted(absent_ids)
    assert session["unit_session_number"] == 1

    parent_item_id = workspace["active_unit"]["checklist"][0]["id"]
    toggle_parent_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/{session_id}/items/{parent_item_id}/toggle",
        headers=headers,
        json={"checked": True},
    )
    assert toggle_parent_resp.status_code == 409
    assert toggle_parent_resp.json()["detail"] == "Checklist headings auto-complete after their child rows are completed."

    leaf_item = _first_leaf_checklist_item(workspace["active_unit"]["checklist"])
    assert leaf_item is not None
    item_id = int(leaf_item["id"])
    toggle_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/{session_id}/items/{item_id}/toggle",
        headers=headers,
        json={"checked": True},
    )
    assert toggle_resp.status_code == 200
    assert toggle_resp.json()["is_completed"] is True

    workspace_after_leaf_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_after_leaf_resp.status_code == 200
    workspace_after_leaf = workspace_after_leaf_resp.json()
    assert workspace_after_leaf["active_unit"]["checklist"][0]["is_completed"] is False
    assert int(workspace_after_leaf["active_session"]["checked_items_count"]) == 1
    assert len(workspace_after_leaf["active_session"]["checked_item_paths"]) == 1

    # Forward-only checklist flow: unchecking is blocked even while session is open.
    toggle_uncheck_open_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/{session_id}/items/{item_id}/toggle",
        headers=headers,
        json={"checked": False},
    )
    assert toggle_uncheck_open_resp.status_code == 409

    end_session_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/{session_id}/end",
        headers=headers,
        json={
            "session_date": "2026-03-03",
            "start_time": "08:00:00",
            "end_time": "10:00:00",
            "absent_student_ids": absent_ids,
            "note": "Session complete",
        },
    )
    assert end_session_resp.status_code == 200
    assert end_session_resp.json()["end_time"] == "10:00:00"

    toggle_closed_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/{session_id}/items/{item_id}/toggle",
        headers=headers,
        json={"checked": False},
    )
    assert toggle_closed_resp.status_code == 409

    keep_end_time_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/{session_id}/end",
        headers=headers,
        json={"note": "Session note only update"},
    )
    assert keep_end_time_resp.status_code == 200
    assert keep_end_time_resp.json()["end_time"] == "10:00:00"

    # Past sessions remain editable.
    edit_past_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/{session_id}/end",
        headers=headers,
        json={"note": "Session edited after save", "end_time": "10:30:00"},
    )
    assert edit_past_resp.status_code == 200
    assert edit_past_resp.json()["end_time"] == "10:30:00"

    # Multiple sessions on same day are allowed.
    second_session_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/start",
        headers=headers,
        json={"absent_student_ids": []},
    )
    assert second_session_resp.status_code == 201
    assert second_session_resp.json()["unit_session_number"] == 2

    unit_sessions_resp = client.get(f"/workflow/units/{unit['id']}/sessions", headers=headers)
    assert unit_sessions_resp.status_code == 200
    unit_sessions = unit_sessions_resp.json()
    assert len(unit_sessions) >= 2
    assert unit_sessions[0]["unit_id"] == unit["id"]
    assert unit_sessions[0]["unit_session_number"] == 1
    assert unit_sessions[1]["unit_session_number"] == 2
    assert unit_sessions[1]["end_time"] is None

    calendar_resp = client.get(f"/workflow/classes/{class_id}/calendar", headers=headers)
    assert calendar_resp.status_code == 200
    events = calendar_resp.json()
    assert len(events) >= 1
    assert events[0]["class_id"] == class_id
    session_event = next((row for row in events if row["session_id"] == session_id), None)
    assert session_event is not None
    assert sorted(session_event["absent_student_ids"]) == sorted(absent_ids)
    assert session_event["unit_session_number"] == 1
    calendar_export_resp = client.get(f"/workflow/classes/{class_id}/calendar/export.xlsx", headers=headers)
    assert calendar_export_resp.status_code == 200
    assert (
        calendar_export_resp.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    export_workbook = load_workbook(filename=BytesIO(calendar_export_resp.content), data_only=True)
    export_sheet = export_workbook.active
    assert export_sheet.cell(1, 1).value == "session_id"
    assert export_sheet.cell(1, 2).value == "session_date"
    assert export_sheet.max_row >= 2
    export_session_ids = {
        int(value)
        for value in [export_sheet.cell(row, 1).value for row in range(2, export_sheet.max_row + 1)]
        if value is not None
    }
    assert session_id in export_session_ids
    calendar_pdf_resp = client.get(
        f"/workflow/classes/{class_id}/calendar/export.pdf?date_from=2026-01-01&date_to=2026-12-31",
        headers=headers,
    )
    assert calendar_pdf_resp.status_code == 200
    assert calendar_pdf_resp.headers["content-type"] == "application/pdf"
    assert len(calendar_pdf_resp.content or b"") > 100

    close_unit_resp = client.post(f"/workflow/classes/{class_id}/units/{unit['id']}/close", headers=headers)
    assert close_unit_resp.status_code == 409  # second session still open

    # End second session then close unit.
    second_id = second_session_resp.json()["id"]
    close_open_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/{second_id}/end",
        headers=headers,
        json={"end_time": "23:59:00"},
    )
    assert close_open_resp.status_code == 200
    close_unit_resp = client.post(f"/workflow/classes/{class_id}/units/{unit['id']}/close", headers=headers)
    assert close_unit_resp.status_code == 200
    assert close_unit_resp.json()["status"] == "closed"
    close_unit_again_resp = client.post(f"/workflow/classes/{class_id}/units/{unit['id']}/close", headers=headers)
    assert close_unit_again_resp.status_code == 409

    exam_correction_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "exam_correction", "title": "Correction CC1"},
    )
    assert exam_correction_resp.status_code == 201
    assert exam_correction_resp.json()["unit_type"] == "exam_correction"


def test_workflow_calendar_session_create_with_unit_and_absences(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Workflow Calendar Create", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    roster = _build_roster_file([("STD110", "Mina"), ("STD111", "Hakim")])
    import_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200
    students = client.get(f"/classes/{class_id}/students", headers=headers).json()
    assert len(students) == 2

    start_unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "chapter", "title": "Calendar Unit", "source_text": "Chapter seed"},
    )
    assert start_unit_resp.status_code == 201
    unit_id = int(start_unit_resp.json()["id"])

    absent_id = int(students[0]["id"])
    create_resp = client.post(
        f"/workflow/classes/{class_id}/sessions",
        headers=headers,
        json={
            "session_date": "2026-03-04",
            "start_time": "09:00:00",
            "end_time": "10:30:00",
            "note": "Calendar block",
            "unit_id": unit_id,
            "absent_student_ids": [absent_id],
        },
    )
    assert create_resp.status_code == 201
    session_payload = create_resp.json()
    session_id = int(session_payload["id"])
    assert session_payload["unit_id"] == unit_id
    assert session_payload["absent_count"] == 1
    assert sorted(session_payload["absent_student_ids"]) == [absent_id]
    assert session_payload["unit_session_number"] == 1

    detail_resp = client.get(f"/sessions/{session_id}", headers=headers)
    assert detail_resp.status_code == 200
    attendance = detail_resp.json()["attendance"]
    assert len(attendance) == 2
    absent_rows = [row for row in attendance if row["status"] == "absent"]
    assert len(absent_rows) == 1
    assert int(absent_rows[0]["student_id"]) == absent_id

    calendar_resp = client.get(f"/workflow/classes/{class_id}/calendar", headers=headers)
    assert calendar_resp.status_code == 200
    created_event = next((row for row in calendar_resp.json() if int(row["session_id"]) == session_id), None)
    assert created_event is not None
    assert int(created_event["unit_id"]) == unit_id
    assert created_event["unit_title"] == "Calendar Unit"
    assert sorted(created_event["absent_student_ids"]) == [absent_id]
    assert created_event["unit_session_number"] == 1

    invalid_unit_resp = client.post(
        f"/workflow/classes/{class_id}/sessions",
        headers=headers,
        json={"session_date": "2026-03-05", "unit_id": 999999},
    )
    assert invalid_unit_resp.status_code == 404

    invalid_absent_resp = client.post(
        f"/workflow/classes/{class_id}/sessions",
        headers=headers,
        json={"session_date": "2026-03-05", "absent_student_ids": [999999]},
    )
    assert invalid_absent_resp.status_code == 400
    assert "Unknown student ids" in str(invalid_absent_resp.json()["detail"])


def test_workflow_slot_actions_create_new_unit_and_continue(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Workflow Slot Actions", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    roster = _build_roster_file([("STD210", "Sara"), ("STD211", "Nabil")])
    import_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200

    new_unit_resp = client.post(
        f"/workflow/classes/{class_id}/slot-actions",
        headers=headers,
        json={
            "action": "new_unit_session",
            "session_date": "2026-03-04",
            "start_time": "08:00:00",
            "end_time": "09:00:00",
            "note": "First session from slot",
            "unit_type": "chapter",
            "unit_title": "Slot Unit",
            "planned_hours": 4,
            "source_text": "Chapter slot seed",
        },
    )
    assert new_unit_resp.status_code == 201
    new_unit_payload = new_unit_resp.json()
    assert new_unit_payload["unit"] is not None
    assert new_unit_payload["unit"]["title"] == "Slot Unit"
    assert new_unit_payload["session"]["unit_id"] == new_unit_payload["unit"]["id"]
    assert new_unit_payload["session"]["unit_session_number"] == 1
    first_leaf = _first_leaf_checklist_item(new_unit_payload["unit"].get("checklist") or [])
    first_check_id = int(first_leaf["id"]) if isinstance(first_leaf, dict) and first_leaf.get("id") else None

    continue_resp = client.post(
        f"/workflow/classes/{class_id}/slot-actions",
        headers=headers,
        json={
            "action": "continue_unit_session",
            "session_date": "2026-03-05",
            "start_time": "08:00:00",
            "end_time": "09:00:00",
            "unit_id": new_unit_payload["unit"]["id"],
            "checked_item_ids": [first_check_id] if first_check_id is not None else [],
        },
    )
    assert continue_resp.status_code == 201
    continue_payload = continue_resp.json()
    assert continue_payload["unit"] is None
    assert continue_payload["session"]["unit_id"] == new_unit_payload["unit"]["id"]
    assert continue_payload["session"]["unit_session_number"] == 2
    if first_check_id is not None:
        assert int(continue_payload["session"]["checked_items_count"]) >= 1
        workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
        assert workspace_resp.status_code == 200
        workspace_payload = workspace_resp.json()
        assert workspace_payload.get("active_unit") is not None
        assert int(workspace_payload["active_unit"]["progress_done"]) >= 1

    invalid_new_unit_resp = client.post(
        f"/workflow/classes/{class_id}/slot-actions",
        headers=headers,
        json={
            "action": "new_unit_session",
            "session_date": "2026-03-06",
            "start_time": "08:00:00",
            "end_time": "09:00:00",
            "unit_title": "Missing type",
        },
    )
    assert invalid_new_unit_resp.status_code == 400


def test_workflow_auto_plan_week_and_unit(client):
    headers = _auth_headers(client)
    class_name = f"AUTO-PLAN-{uuid.uuid4().hex[:6]}"
    class_resp = client.post("/classes", json={"name": class_name, "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    roster = _build_roster_file([("STD500", "A"), ("STD501", "B")])
    import_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200

    csv_content = _build_timetable_csv(
        [
            ("owner@school.edu", class_name, "Math", "Monday", "08:00", "09:00", "R12", "G1"),
            ("owner@school.edu", class_name, "Math", "Wednesday", "10:00", "11:00", "R12", "G1"),
        ]
    )
    apply_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "append_new_slots",
            "effective_from": "2026-09-01",
            "create_missing_classes": "false",
        },
        files={"file": ("emploi.csv", csv_content, "text/csv")},
    )
    assert apply_resp.status_code == 200
    assert int(apply_resp.json()["applied_rows"]) >= 2

    existing_resp = client.post(
        f"/workflow/classes/{class_id}/sessions",
        headers=headers,
        json={
            "session_date": "2026-09-07",
            "start_time": "08:00:00",
            "end_time": "09:00:00",
            "note": "Existing manual session",
        },
    )
    assert existing_resp.status_code == 201

    week_plan_resp = client.post(
        f"/workflow/classes/{class_id}/auto-plan",
        headers=headers,
        json={
            "action": "load_week_plan",
            "week_start": "2026-09-07",
        },
    )
    assert week_plan_resp.status_code == 200
    week_payload = week_plan_resp.json()
    assert week_payload["action"] == "load_week_plan"
    assert int(week_payload["created_count"]) >= 1
    assert int(week_payload["skipped_existing_count"]) >= 1
    assert all(row["unit_id"] is None for row in (week_payload.get("created_sessions") or []))

    plan_unit_resp = client.post(
        f"/workflow/classes/{class_id}/auto-plan",
        headers=headers,
        json={
            "action": "plan_unit",
            "plan_mode": "new_unit",
            "start_date": "2026-09-08",
            "session_count": 3,
            "unit_type": "chapter",
            "unit_title": "Auto Planned Unit",
            "source_text": "Auto Planned Unit",
        },
    )
    assert plan_unit_resp.status_code == 200
    unit_payload = plan_unit_resp.json()
    assert unit_payload["action"] == "plan_unit"
    assert int(unit_payload["requested_count"]) == 3
    assert int(unit_payload["created_count"]) >= 1
    assert unit_payload["target_unit_id"] is not None
    created_rows = unit_payload.get("created_sessions") or []
    assert len(created_rows) == int(unit_payload["created_count"])
    assert all(int(row["unit_id"]) == int(unit_payload["target_unit_id"]) for row in created_rows)
    unit_numbers = [int(row["unit_session_number"]) for row in created_rows if row.get("unit_session_number") is not None]
    assert unit_numbers == sorted(unit_numbers)


def test_workflow_auto_plan_unit_skips_holidays_and_jumps_forward(client):
    headers = _auth_headers(client)
    class_name = f"AUTO-PLAN-HOL-{uuid.uuid4().hex[:6]}"
    class_resp = client.post("/classes", json={"name": class_name, "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    roster = _build_roster_file([("STD900", "A"), ("STD901", "B")])
    import_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200

    csv_content = _build_timetable_csv(
        [
            ("owner@school.edu", class_name, "Math", "Wednesday", "08:00", "09:00", "R12", "G1"),
        ]
    )
    apply_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "append_new_slots",
            "effective_from": "2026-09-01",
            "create_missing_classes": "false",
        },
        files={"file": ("emploi.csv", csv_content, "text/csv")},
    )
    assert apply_resp.status_code == 200
    assert int(apply_resp.json()["applied_rows"]) >= 1

    auto_plan_resp = client.post(
        f"/workflow/classes/{class_id}/auto-plan",
        headers=headers,
        json={
            "action": "plan_unit",
            "plan_mode": "new_unit",
            "start_date": "2026-11-16",
            "session_count": 2,
            "unit_type": "exercise_series",
            "unit_title": "Series Auto Holiday",
            "source_text": "Series Auto Holiday",
        },
    )
    assert auto_plan_resp.status_code == 200
    payload = auto_plan_resp.json()
    assert int(payload["created_count"]) == 2
    assert int(payload["failed_count"]) == 0
    assert int(payload["skipped_holiday_count"]) >= 1
    created_dates = [str(row["session_date"]) for row in (payload.get("created_sessions") or [])]
    assert "2026-11-18" not in created_dates
    assert "2026-11-25" in created_dates
    assert "2026-12-02" in created_dates


def test_workflow_auto_plan_unit_dry_run_preview_does_not_create_data(client):
    headers = _auth_headers(client)
    class_name = f"AUTO-PLAN-DRY-{uuid.uuid4().hex[:6]}"
    class_resp = client.post("/classes", json={"name": class_name, "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    roster = _build_roster_file([("STD910", "A"), ("STD911", "B")])
    import_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200

    csv_content = _build_timetable_csv(
        [
            ("owner@school.edu", class_name, "Math", "Monday", "08:00", "09:00", "R12", "G1"),
            ("owner@school.edu", class_name, "Math", "Wednesday", "10:00", "11:00", "R12", "G1"),
        ]
    )
    apply_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "append_new_slots",
            "effective_from": "2026-09-01",
            "create_missing_classes": "false",
        },
        files={"file": ("emploi.csv", csv_content, "text/csv")},
    )
    assert apply_resp.status_code == 200
    assert int(apply_resp.json()["applied_rows"]) >= 2

    preview_resp = client.post(
        f"/workflow/classes/{class_id}/auto-plan",
        headers=headers,
        json={
            "action": "plan_unit",
            "dry_run": True,
            "plan_mode": "new_unit",
            "start_date": "2026-09-07",
            "session_count": 3,
            "unit_type": "exercise_series",
            "unit_title": "Dry Run Series",
            "source_text": "Dry Run Series",
        },
    )
    assert preview_resp.status_code == 200
    preview_payload = preview_resp.json()
    assert preview_payload["action"] == "plan_unit"
    assert preview_payload["created_count"] == 0
    assert int(preview_payload["planned_count"]) >= 1
    assert len(preview_payload.get("planned_slots") or []) == int(preview_payload["planned_count"])
    assert preview_payload["target_unit_id"] is None
    assert preview_payload["target_unit_title"] == "Dry Run Series"

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_resp.status_code == 200
    assert workspace_resp.json().get("active_unit") is None

    calendar_resp = client.get(f"/workflow/classes/{class_id}/calendar", headers=headers)
    assert calendar_resp.status_code == 200
    assert len(calendar_resp.json()) == 0


def test_workspace_load_survives_corrupted_checklist_cycle(client):
    from app.database import SessionLocal
    from app.models import WorkflowChecklistItem

    owner_headers = _auth_headers(client)
    _, teacher_headers = _create_teacher_and_login(client, owner_headers)

    class_resp = client.post(
        "/classes",
        headers=teacher_headers,
        json={"name": "APIC 2", "subject": "Math", "level": "2"},
    )
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    source_pdf = _build_pdf_file(
        [
            "Chapitre 1 : Nombres rationnels",
            "1. Definition d'un nombre rationnel",
            "2. Exemples et applications",
        ]
    )
    start_unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=teacher_headers,
        data={"unit_type": "chapter", "title": "Chapitre 1"},
        files={"file": ("chapter.pdf", source_pdf, "application/pdf")},
    )
    assert start_unit_resp.status_code == 201
    unit_id = int(start_unit_resp.json()["id"])

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=teacher_headers)
    assert workspace_resp.status_code == 200
    root_id = int(workspace_resp.json()["active_unit"]["checklist"][0]["id"])

    add_item_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/items",
        headers=teacher_headers,
        json={"title": "1.1 Exemple guide", "item_kind": "example", "parent_item_id": root_id},
    )
    assert add_item_resp.status_code == 201
    child_id = int(add_item_resp.json()["id"])

    db = SessionLocal()
    try:
        row = db.get(WorkflowChecklistItem, child_id)
        assert row is not None
        row.parent_item_id = child_id
        db.commit()
    finally:
        db.close()

    repaired_workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=teacher_headers)
    assert repaired_workspace_resp.status_code == 200
    payload = repaired_workspace_resp.json()
    assert payload["active_unit"] is not None
    assert payload["active_unit"]["id"] == unit_id
    assert payload["active_unit"]["checklist"]
    assert int(payload["active_unit"]["checklist"][0]["id"]) == root_id


def test_workflow_auto_plan_new_unit_dry_run_respects_active_unit_conflict(client):
    headers = _auth_headers(client)
    class_name = f"AUTO-PLAN-CONFLICT-{uuid.uuid4().hex[:6]}"
    class_resp = client.post("/classes", json={"name": class_name, "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    roster = _build_roster_file([("STD920", "A"), ("STD921", "B")])
    import_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200

    first_unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "exam",
            "title": "Existing Active Unit",
        },
    )
    assert first_unit_resp.status_code == 201

    dry_run_resp = client.post(
        f"/workflow/classes/{class_id}/auto-plan",
        headers=headers,
        json={
            "action": "plan_unit",
            "dry_run": True,
            "plan_mode": "new_unit",
            "start_date": "2026-09-07",
            "session_count": 2,
            "unit_type": "exercise_series",
            "unit_title": "Should Conflict",
            "source_text": "Should Conflict",
        },
    )
    assert dry_run_resp.status_code == 409
    assert "active unit already exists" in str(dry_run_resp.json()["detail"]).lower()


def test_workflow_start_session_uses_next_timetable_slot_after_last_unit_session(client):
    headers = _auth_headers(client)
    class_name = f"NEXT-SESSION-{uuid.uuid4().hex[:6]}"
    class_resp = client.post("/classes", json={"name": class_name, "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    roster = _build_roster_file([("STD930", "A"), ("STD931", "B")])
    import_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200

    csv_content = _build_timetable_csv(
        [
            ("owner@school.edu", class_name, "Math", "Monday", "08:00", "09:00", "R12", "G1"),
            ("owner@school.edu", class_name, "Math", "Wednesday", "10:00", "11:00", "R12", "G1"),
        ]
    )
    apply_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "append_new_slots",
            "effective_from": "2026-09-01",
            "create_missing_classes": "false",
        },
        files={"file": ("emploi.csv", csv_content, "text/csv")},
    )
    assert apply_resp.status_code == 200

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "exam", "title": "Rationals Unit"},
    )
    assert unit_resp.status_code == 201
    unit_id = int(unit_resp.json()["id"])

    add_item_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/items",
        headers=headers,
        json={"title": "I- Multiplication", "item_kind": "section", "parent_item_id": None},
    )
    assert add_item_resp.status_code == 201

    previous_resp = client.post(
        f"/workflow/classes/{class_id}/sessions",
        headers=headers,
        json={
            "unit_id": unit_id,
            "session_date": "2026-09-07",
            "start_time": "08:00:00",
            "end_time": "09:00:00",
            "note": "Live unit session",
        },
    )
    assert previous_resp.status_code == 201

    start_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/start",
        headers=headers,
        json={"absent_student_ids": []},
    )
    assert start_resp.status_code == 201
    payload = start_resp.json()
    assert int(payload["unit_id"]) == unit_id
    assert str(payload["session_date"]) == "2026-09-09"
    assert str(payload["start_time"]) == "10:00:00"
    assert payload["end_time"] is None


def test_workflow_start_session_skips_existing_future_slot_and_uses_next_available_one(client):
    headers = _auth_headers(client)
    class_name = f"NEXT-SESSION-SKIP-{uuid.uuid4().hex[:6]}"
    class_resp = client.post("/classes", json={"name": class_name, "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    roster = _build_roster_file([("STD940", "A"), ("STD941", "B")])
    import_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200

    csv_content = _build_timetable_csv(
        [
            ("owner@school.edu", class_name, "Math", "Monday", "08:00", "09:00", "R12", "G1"),
            ("owner@school.edu", class_name, "Math", "Wednesday", "10:00", "11:00", "R12", "G1"),
        ]
    )
    apply_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "append_new_slots",
            "effective_from": "2026-09-01",
            "create_missing_classes": "false",
        },
        files={"file": ("emploi.csv", csv_content, "text/csv")},
    )
    assert apply_resp.status_code == 200

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "exam", "title": "Rationals Unit"},
    )
    assert unit_resp.status_code == 201
    unit_id = int(unit_resp.json()["id"])

    add_item_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/items",
        headers=headers,
        json={"title": "I- Multiplication", "item_kind": "section", "parent_item_id": None},
    )
    assert add_item_resp.status_code == 201

    previous_resp = client.post(
        f"/workflow/classes/{class_id}/sessions",
        headers=headers,
        json={
            "unit_id": unit_id,
            "session_date": "2026-09-07",
            "start_time": "08:00:00",
            "end_time": "09:00:00",
            "note": "Current session",
        },
    )
    assert previous_resp.status_code == 201

    existing_future_resp = client.post(
        f"/workflow/classes/{class_id}/sessions",
        headers=headers,
        json={
            "unit_id": unit_id,
            "session_date": "2026-09-09",
            "start_time": "10:00:00",
            "end_time": "11:00:00",
            "note": "Already planned next session",
        },
    )
    assert existing_future_resp.status_code == 201

    start_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/start",
        headers=headers,
        json={"absent_student_ids": []},
    )
    assert start_resp.status_code == 201
    payload = start_resp.json()
    assert int(payload["unit_id"]) == unit_id
    assert str(payload["session_date"]) == "2026-09-14"
    assert str(payload["start_time"]) == "08:00:00"


def test_workflow_class_setup_creates_class_students_and_timetable(client):
    headers = _auth_headers(client)
    setup_resp = client.post(
        "/workflow/class-setup",
        headers=headers,
        json={
            "class_name": f"SETUP-INIT-{uuid.uuid4().hex[:6]}",
            "subject": "Math",
            "level": "2BAC",
            "student_mode": "append_new",
            "students": [
                {"full_name": "Sara Zahra"},
                {"student_code": "STD900", "full_name": "Nabil Idrissi"},
            ],
            "timetable_mode": "replace_future_from_date",
            "effective_from": "2026-09-01",
            "timetable_rows": [
                {
                    "weekday": 1,
                    "start_time": "08:00:00",
                    "end_time": "09:00:00",
                    "subject": "Algebra",
                    "room": "R12",
                    "group": "A",
                },
                {
                    "weekday": 3,
                    "start_time": "10:00:00",
                    "end_time": "11:00:00",
                    "subject": "Geometry",
                    "room": "R12",
                    "group": "A",
                },
            ],
        },
    )
    assert setup_resp.status_code == 200
    payload = setup_resp.json()
    class_id = int(payload["class_id"])
    assert payload["created_class"] is True
    assert int(payload["students_created"]) == 2
    assert int(payload["students_total"]) == 2
    assert int(payload["timetable_applied_rows"]) == 2

    students_resp = client.get(f"/classes/{class_id}/students", headers=headers)
    assert students_resp.status_code == 200
    students = students_resp.json()
    assert len(students) == 2
    assert any(str(row["student_code"]).startswith("AUTO") for row in students)
    assert any(str(row["student_code"]) == "STD900" for row in students)

    rules_resp = client.get(f"/workflow/classes/{class_id}/timetable-rules", headers=headers)
    assert rules_resp.status_code == 200
    rules = rules_resp.json()
    assert len(rules) == 2
    assert all(str(row.get("source") or "") == "class-setup-form" for row in rules)


def test_workflow_class_setup_resubmits_timetable_on_existing_class(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": f"SETUP-RESUBMIT-{uuid.uuid4().hex[:6]}", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    first_setup_resp = client.post(
        "/workflow/class-setup",
        headers=headers,
        json={
            "class_id": class_id,
            "student_mode": "ignore",
            "timetable_mode": "replace_future_from_date",
            "effective_from": "2026-09-01",
            "timetable_rows": [
                {
                    "weekday": 1,
                    "start_time": "08:00:00",
                    "end_time": "09:00:00",
                    "subject": "Algebra",
                    "room": "R10",
                    "group": "A",
                }
            ],
        },
    )
    assert first_setup_resp.status_code == 200
    assert int(first_setup_resp.json()["timetable_applied_rows"]) == 1

    second_setup_resp = client.post(
        "/workflow/class-setup",
        headers=headers,
        json={
            "class_id": class_id,
            "student_mode": "ignore",
            "timetable_mode": "replace_future_from_date",
            "effective_from": "2026-10-01",
            "timetable_rows": [
                {
                    "weekday": 4,
                    "start_time": "14:00:00",
                    "end_time": "15:00:00",
                    "subject": "Functions",
                    "room": "R10",
                    "group": "A",
                }
            ],
        },
    )
    assert second_setup_resp.status_code == 200
    second_payload = second_setup_resp.json()
    assert int(second_payload["timetable_applied_rows"]) == 1
    assert int(second_payload["timetable_replaced_existing_count"]) >= 1

    rules_resp = client.get(f"/workflow/classes/{class_id}/timetable-rules", headers=headers)
    assert rules_resp.status_code == 200
    rules = rules_resp.json()
    assert len(rules) == 2
    first_rule = next((row for row in rules if row["effective_from"] == "2026-09-01"), None)
    latest_rule = next((row for row in rules if row["effective_from"] == "2026-10-01"), None)
    assert first_rule is not None
    assert latest_rule is not None
    assert first_rule["effective_to"] == "2026-09-30"
    assert latest_rule["effective_to"] is None


def test_workflow_auto_setup_from_document_creates_sessions_and_checks_progress(client):
    headers = _auth_headers(client)
    class_name = f"DOC-AUTO-SETUP-{uuid.uuid4().hex[:6]}"
    class_resp = client.post("/classes", json={"name": class_name, "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    roster = _build_roster_file([("STD501", "Lina"), ("STD502", "Ayoub")])
    import_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200

    csv_content = _build_timetable_csv(
        [
            ("owner@school.edu", class_name, "Math", "Monday", "08:00", "09:00", "R20", "A"),
            ("owner@school.edu", class_name, "Math", "Wednesday", "08:00", "09:00", "R20", "A"),
        ]
    )
    apply_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "append_new_slots",
            "effective_from": "2026-09-01",
            "create_missing_classes": "false",
        },
        files={"file": ("emploi.csv", csv_content, "text/csv")},
    )
    assert apply_resp.status_code == 200
    assert int(apply_resp.json()["applied_rows"]) >= 2

    auto_setup_resp = client.post(
        f"/workflow/classes/{class_id}/auto-setup-from-doc",
        headers=headers,
        data={
            "unit_type": "chapter",
            "unit_title": "Fractions and Operations",
            "source_text": "Chapter Fractions\nDefinition of fraction\nEquivalent fractions\nSimplify fractions\nExercises",
            "start_date": "2026-09-07",
            "session_count": "2",
            "auto_check_items": "true",
        },
    )
    assert auto_setup_resp.status_code == 200
    payload = auto_setup_resp.json()
    assert payload["action"] == "plan_document_unit"
    assert int(payload["created_count"]) == 2
    assert int(payload["target_unit_id"]) > 0
    assert len(payload.get("created_sessions") or []) == 2
    assert sum(int(row.get("checked_items_count") or 0) for row in payload["created_sessions"]) >= 1

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_resp.status_code == 200
    workspace_payload = workspace_resp.json()
    active_unit = workspace_payload.get("active_unit")
    assert active_unit is not None
    assert int(active_unit.get("progress_done") or 0) >= 1


def test_workflow_auto_setup_from_document_fills_empty_sessions_with_exercises(client):
    headers = _auth_headers(client)
    class_name = f"DOC-AUTO-FILL-{uuid.uuid4().hex[:6]}"
    class_resp = client.post("/classes", json={"name": class_name, "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    roster = _build_roster_file([("STD601", "Sara"), ("STD602", "Yassine")])
    import_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200

    csv_content = _build_timetable_csv(
        [
            ("owner@school.edu", class_name, "Math", "Monday", "08:00", "09:00", "R30", "A"),
            ("owner@school.edu", class_name, "Math", "Wednesday", "08:00", "09:00", "R30", "A"),
        ]
    )
    apply_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "append_new_slots",
            "effective_from": "2026-09-01",
            "create_missing_classes": "false",
        },
        files={"file": ("emploi.csv", csv_content, "text/csv")},
    )
    assert apply_resp.status_code == 200
    assert int(apply_resp.json()["applied_rows"]) >= 2

    auto_setup_resp = client.post(
        f"/workflow/classes/{class_id}/auto-setup-from-doc",
        headers=headers,
        data={
            "unit_type": "chapter",
            "unit_title": "Very short chapter",
            "source_text": "Chapter tiny content",
            "start_date": "2026-09-07",
            "session_count": "4",
            "auto_check_items": "true",
        },
    )
    assert auto_setup_resp.status_code == 200
    payload = auto_setup_resp.json()
    assert payload["action"] == "plan_document_unit"
    assert int(payload["created_count"]) == 4
    created_sessions = payload.get("created_sessions") or []
    assert len(created_sessions) == 4
    assert all(int(row.get("checked_items_count") or 0) >= 1 for row in created_sessions)


def test_timetable_version_snapshot_compare_restore(client):
    headers = _auth_headers(client)
    class_name = f"TT-VERSION-{uuid.uuid4().hex[:6]}"
    class_resp = client.post("/classes", json={"name": class_name, "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    initial_csv = _build_timetable_csv(
        [
            ("owner@school.edu", class_name, "Math", "Monday", "08:00", "09:00", "R12", "G1"),
        ]
    )
    initial_apply_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "append_new_slots",
            "effective_from": "2026-09-01",
            "create_missing_classes": "false",
        },
        files={"file": ("initial.csv", initial_csv, "text/csv")},
    )
    assert initial_apply_resp.status_code == 200
    assert int(initial_apply_resp.json()["applied_rows"]) >= 1

    snapshot_resp = client.post(
        f"/workflow/classes/{class_id}/timetable-versions",
        headers=headers,
        json={"label": "Baseline Snapshot", "source": "test"},
    )
    assert snapshot_resp.status_code == 201
    snapshot_payload = snapshot_resp.json()
    version_id = int(snapshot_payload["id"])
    assert snapshot_payload["is_active"] is True
    assert int(snapshot_payload["rules_count"]) >= 1

    replacement_csv = _build_timetable_csv(
        [
            ("owner@school.edu", class_name, "Math", "Tuesday", "10:00", "11:00", "R14", "G1"),
        ]
    )
    replace_apply_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "replace_future_from_date",
            "effective_from": "2026-09-01",
            "create_missing_classes": "false",
        },
        files={"file": ("replacement.csv", replacement_csv, "text/csv")},
    )
    assert replace_apply_resp.status_code == 200
    assert int(replace_apply_resp.json()["applied_rows"]) >= 1

    compare_resp = client.get(
        f"/workflow/classes/{class_id}/timetable-versions/{version_id}/compare-current",
        headers=headers,
    )
    assert compare_resp.status_code == 200
    compare_payload = compare_resp.json()
    assert int(compare_payload["snapshot_only_rules_count"]) >= 1
    assert int(compare_payload["current_only_rules_count"]) >= 1

    restore_resp = client.post(
        f"/workflow/classes/{class_id}/timetable-versions/{version_id}/restore",
        headers=headers,
    )
    assert restore_resp.status_code == 200
    restore_payload = restore_resp.json()
    assert int(restore_payload["active_version_id"]) == version_id
    assert int(restore_payload["restored_rules_count"]) >= 1

    rules_resp = client.get(f"/workflow/classes/{class_id}/timetable-rules", headers=headers)
    assert rules_resp.status_code == 200
    restored_rules = rules_resp.json()
    assert any(int(row["weekday"]) == 1 and row["start_time"] == "08:00:00" for row in restored_rules)

    versions_resp = client.get(f"/workflow/classes/{class_id}/timetable-versions", headers=headers)
    assert versions_resp.status_code == 200
    versions_payload = versions_resp.json()
    active_versions = [row for row in versions_payload if row.get("is_active")]
    assert len(active_versions) == 1
    assert int(active_versions[0]["id"]) == version_id


def test_morocco_holidays_list_and_block_session_creation(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Holiday Block Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    holidays_resp = client.get("/workflow/holidays", headers=headers, params={"year": 2026, "country_code": "MA"})
    assert holidays_resp.status_code == 200
    holidays_rows = holidays_resp.json()
    assert len(holidays_rows) >= 1
    independence_row = next((row for row in holidays_rows if row["holiday_date"] == "2026-11-18"), None)
    assert independence_row is not None
    assert independence_row["is_blocked"] is True

    blocked_generic_resp = client.post(
        f"/classes/{class_id}/sessions",
        headers=headers,
        json={"session_date": "2026-11-18", "start_time": "08:00:00", "end_time": "09:00:00"},
    )
    assert blocked_generic_resp.status_code == 409

    blocked_workflow_resp = client.post(
        f"/workflow/classes/{class_id}/sessions",
        headers=headers,
        json={"session_date": "2026-11-18", "start_time": "08:00:00", "end_time": "09:00:00"},
    )
    assert blocked_workflow_resp.status_code == 409

    override_workflow_resp = client.post(
        f"/workflow/classes/{class_id}/sessions",
        headers=headers,
        json={
            "session_date": "2026-11-18",
            "start_time": "08:00:00",
            "end_time": "09:00:00",
            "allow_on_holiday": True,
        },
    )
    assert override_workflow_resp.status_code == 201

    unblocked_resp = client.patch(
        f"/workflow/holidays/{independence_row['id']}",
        headers=headers,
        json={"is_blocked": False},
    )
    assert unblocked_resp.status_code == 200
    assert unblocked_resp.json()["is_blocked"] is False

    allowed_generic_resp = client.post(
        f"/classes/{class_id}/sessions",
        headers=headers,
        json={"session_date": "2026-11-18", "start_time": "10:00:00", "end_time": "11:00:00"},
    )
    assert allowed_generic_resp.status_code == 201


def test_morocco_academic_holidays_seed_exact_2025_2026_calendar(client):
    headers = _auth_headers(client)

    holidays_2025_resp = client.get("/workflow/holidays", headers=headers, params={"year": 2025, "country_code": "MA"})
    assert holidays_2025_resp.status_code == 200
    holidays_2025 = holidays_2025_resp.json()
    assert len(holidays_2025) == 28
    by_date_2025 = {row["holiday_date"]: row for row in holidays_2025}
    assert by_date_2025["2025-09-04"]["name"] == "Prophet's Birthday"
    assert by_date_2025["2025-10-19"]["name"] == "First Mid-Term Break"
    assert by_date_2025["2025-11-06"]["name"] == "Green March Anniversary"
    assert by_date_2025["2025-12-14"]["name"] == "Second Mid-Term Break"

    holidays_2026_resp = client.get("/workflow/holidays", headers=headers, params={"year": 2026, "country_code": "MA"})
    assert holidays_2026_resp.status_code == 200
    holidays_2026 = holidays_2026_resp.json()
    assert len(holidays_2026) == 38
    by_date_2026 = {row["holiday_date"]: row for row in holidays_2026}
    assert by_date_2026["2026-01-14"]["name"] == "Amazigh New Year (Yennayer)"
    assert by_date_2026["2026-03-18"]["name"] == "Third Mid-Term Break / Eid al-Fitr"
    assert by_date_2026["2026-05-27"]["name"] == "Eid al-Adha"
    assert by_date_2026["2026-06-16"]["name"] == "Islamic New Year (1 Muharram)"


def test_owner_can_download_holiday_template(client):
    headers = _auth_headers(client)

    response = client.get("/workflow/holidays/template.xlsx", headers=headers)
    assert response.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers["content-type"]

    workbook = load_workbook(filename=BytesIO(response.content), data_only=True)
    assert "holiday_import" in workbook.sheetnames
    sheet = workbook["holiday_import"]
    header = [cell.value for cell in sheet[1]]
    assert header[:4] == ["holiday", "start_date", "end_date", "is_blocked"]
    first_row = [sheet.cell(2, idx).value for idx in range(1, 5)]
    assert first_row == ["Prophet's Birthday", "2025-09-04", "2025-09-05", True]


def test_owner_can_export_current_holidays_excel(client):
    headers = _auth_headers(client)

    response = client.get(
        "/workflow/holidays/export.xlsx",
        headers=headers,
        params={"year": 2026, "country_code": "MA"},
    )
    assert response.status_code == 200
    workbook = load_workbook(filename=BytesIO(response.content), data_only=True)
    assert "holiday_export" in workbook.sheetnames
    sheet = workbook["holiday_export"]
    header = [cell.value for cell in sheet[1]]
    assert header == ["holiday", "start_date", "end_date", "number_of_days", "is_blocked", "source"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert ("Mid-Year School Break", "2026-01-25", "2026-02-01", 8, True, "morocco-academic-2025-2026") in rows
    assert ("New Year's Day", "2026-01-01", "2026-01-01", 1, True, "morocco-fixed / morocco-academic-2025-2026") in rows


def test_owner_holiday_import_adds_to_standard_calendar_and_preserves_existing_custom_rows(client):
    headers = _auth_headers(client)

    first_import = _build_holiday_file(
        ("holiday", "dates", "number_of_days", "is_blocked"),
        [
            ("Green March Day", "November 6, 2025", 1, True),
        ],
    )

    first_import_resp = client.post(
        "/workflow/holidays/import",
        headers=headers,
        files={"file": ("holidays.xlsx", first_import, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert first_import_resp.status_code == 200
    first_payload = first_import_resp.json()
    assert first_payload["rows"] == 1
    assert first_payload["holiday_dates"] == 1
    assert first_payload["years"] == [2025]

    holidays_2025_resp = client.get("/workflow/holidays", headers=headers, params={"year": 2025, "country_code": "MA"})
    assert holidays_2025_resp.status_code == 200
    holidays_2025 = holidays_2025_resp.json()
    green_march_row = next((row for row in holidays_2025 if row["holiday_date"] == "2025-11-06"), None)
    prophet_row = next((row for row in holidays_2025 if row["holiday_date"] == "2025-09-04"), None)
    assert green_march_row is not None
    assert green_march_row["name"] == "Green March Day"
    assert "morocco-fixed" in str(green_march_row["source"] or "")
    assert "owner-academic-upload" in str(green_march_row["source"] or "")
    assert prophet_row is not None
    assert prophet_row["name"] == "Prophet's Birthday"

    second_import = _build_holiday_file(
        ("holiday", "start_date", "end_date", "is_blocked"),
        [
            ("School Open House", "2025-10-02", "2025-10-02", True),
        ],
    )
    second_import_resp = client.post(
        "/workflow/holidays/import",
        headers=headers,
        files={"file": ("holidays-update.xlsx", second_import, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert second_import_resp.status_code == 200
    second_payload = second_import_resp.json()
    assert second_payload["rows"] == 1
    assert second_payload["holiday_dates"] == 1
    assert second_payload["years"] == [2025]

    refreshed_2025_resp = client.get("/workflow/holidays", headers=headers, params={"year": 2025, "country_code": "MA"})
    assert refreshed_2025_resp.status_code == 200
    refreshed_2025 = refreshed_2025_resp.json()
    refreshed_green_march_row = next((row for row in refreshed_2025 if row["holiday_date"] == "2025-11-06"), None)
    open_house_row = next((row for row in refreshed_2025 if row["holiday_date"] == "2025-10-02"), None)
    assert refreshed_green_march_row is not None
    assert refreshed_green_march_row["name"] == "Green March Day"
    assert open_house_row is not None
    assert open_house_row["name"] == "School Open House"
    assert open_house_row["source"] == "owner-academic-upload"

    patch_resp = client.patch(
        f"/workflow/holidays/{open_house_row['id']}",
        headers=headers,
        json={"is_blocked": False},
    )
    assert patch_resp.status_code == 200
    assert patch_resp.json()["is_blocked"] is False

    final_2025_resp = client.get("/workflow/holidays", headers=headers, params={"year": 2025, "country_code": "MA"})
    assert final_2025_resp.status_code == 200
    final_2025 = final_2025_resp.json()
    final_open_house_row = next((row for row in final_2025 if row["holiday_date"] == "2025-10-02"), None)
    assert final_open_house_row is not None
    assert final_open_house_row["is_blocked"] is False


def test_workflow_unit_start_persists_blueprint(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Blueprint Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Factorisation",
            "source_text": (
                "1.3 Factorisation\n"
                "1.3.2 Factorisation a l'aide des identites remarquables\n"
                "Propriete: Utiliser les identites remarquables.\n"
                "Exemple: Factoriser une expression simple.\n"
            ),
        },
    )
    assert unit_resp.status_code == 201
    unit_id = unit_resp.json()["id"]

    blueprint_resp = client.get(f"/workflow/classes/{class_id}/units/{unit_id}/blueprint", headers=headers)
    assert blueprint_resp.status_code == 200
    blueprint = blueprint_resp.json()
    assert blueprint["unit_id"] == unit_id
    assert blueprint["status"] == "ready"
    assert blueprint["reviewed"] is False
    assert blueprint["blueprint_json"]["unit_title"] == "Factorisation"
    assert isinstance(blueprint["blueprint_json"]["items"], list)
    assert len(blueprint["blueprint_json"]["items"]) >= 1
    assert isinstance(blueprint["unit_map_json"], dict)
    assert blueprint["unit_map_json"]["unit_title"] == "Factorisation"
    assert blueprint["unit_map_json"]["unit_type"] == "chapter"
    assert isinstance(blueprint["unit_map_json"]["ordered_outline"], list)
    assert "ask_unit" in blueprint["unit_map_json"]["future_actions"]

    unit_payload = unit_resp.json()
    assert unit_payload["extraction_reviewed"] is False


def test_workflow_unit_review_endpoint_updates_review_state(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Review State Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Fractions",
            "source_text": "Fractions\nDefinition\nExercices\n",
        },
    )
    assert unit_resp.status_code == 201
    unit_id = unit_resp.json()["id"]
    assert unit_resp.json()["extraction_reviewed"] is False

    approve_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/review",
        headers=headers,
        json={"reviewed": True},
    )
    assert approve_resp.status_code == 200
    approved_unit = approve_resp.json()
    assert approved_unit["extraction_reviewed"] is True
    assert approved_unit["extraction_reviewed_at"] is not None

    blueprint_resp = client.get(f"/workflow/classes/{class_id}/units/{unit_id}/blueprint", headers=headers)
    assert blueprint_resp.status_code == 200
    blueprint = blueprint_resp.json()
    assert blueprint["reviewed"] is True
    assert blueprint["reviewed_at"] is not None

    reopen_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/review",
        headers=headers,
        json={"reviewed": False},
    )
    assert reopen_resp.status_code == 200
    reopened_unit = reopen_resp.json()
    assert reopened_unit["extraction_reviewed"] is False


def test_workflow_confirm_can_generate_saved_session_writeup(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Writeup Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    roster_content = _build_roster_file([("A1", "Student One")])
    roster_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("roster.xlsx", roster_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert roster_resp.status_code == 200

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_resp.status_code == 200
    workspace_payload = workspace_resp.json()
    active_unit = workspace_payload.get("active_unit")
    if isinstance(active_unit, dict) and active_unit.get("id"):
        close_active_resp = client.post(
            f"/workflow/classes/{class_id}/units/{active_unit['id']}/close",
            headers=headers,
        )
        assert close_active_resp.status_code == 200
    active_session = workspace_payload.get("active_session")
    if isinstance(active_session, dict) and active_session.get("id"):
        end_active_session_resp = client.post(
            f"/workflow/classes/{class_id}/sessions/{active_session['id']}/end",
            headers=headers,
            json={},
        )
        assert end_active_session_resp.status_code == 200

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Identites remarquables",
            "source_text": (
                "1. Identites remarquables\n"
                "1.1 Somme et difference\n"
                "Propriete: Reconnaitre une identite remarquable.\n"
                "Exemple: a2-b2.\n"
                "Application: Exercices de factorisation.\n"
            ),
        },
    )
    assert unit_resp.status_code == 201

    start_session_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/start",
        headers=headers,
        json={"absent_student_ids": []},
    )
    assert start_session_resp.status_code == 201
    session_id = start_session_resp.json()["id"]

    confirm_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/{session_id}/confirm",
        headers=headers,
        json={"generate_session_writeup": True},
    )
    assert confirm_resp.status_code == 200
    confirm_payload = confirm_resp.json()
    assert confirm_payload["writeup_generated"] is True
    assert confirm_payload["session"]["has_saved_writeup"] is True

    writeup_resp = client.get(
        f"/workflow/classes/{class_id}/sessions/{session_id}/writeup",
        headers=headers,
    )
    assert writeup_resp.status_code == 200
    writeup = writeup_resp.json()
    assert writeup["status"] == "ready"
    assert writeup["provider"] in {"fallback", "openai"}
    assert len(writeup["learning_focus"]) >= 1
    assert len(writeup["teaching_content"]) >= 1
    assert len(writeup["practice_items"]) >= 1


def test_workflow_confirm_rejects_already_confirmed_session(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Confirmed Once Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    roster_content = _build_roster_file([("B1", "Student Two")])
    roster_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("roster.xlsx", roster_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert roster_resp.status_code == 200

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_resp.status_code == 200
    active_unit = workspace_resp.json().get("active_unit")
    if isinstance(active_unit, dict) and active_unit.get("id"):
        close_active_resp = client.post(
            f"/workflow/classes/{class_id}/units/{active_unit['id']}/close",
            headers=headers,
        )
        assert close_active_resp.status_code == 200

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Trigonometrie",
            "source_text": (
                "1. Trigonometrie\n"
                "1.1 Activite\n"
                "Definition: Sinus et cosinus.\n"
                "Exemple: Triangle rectangle.\n"
                "Exercices: Calculs.\n"
            ),
        },
    )
    assert unit_resp.status_code == 201

    start_session_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/start",
        headers=headers,
        json={"absent_student_ids": []},
    )
    assert start_session_resp.status_code == 201
    session_id = start_session_resp.json()["id"]

    first_confirm = client.post(
        f"/workflow/classes/{class_id}/sessions/{session_id}/confirm",
        headers=headers,
        json={"generate_session_writeup": True},
    )
    assert first_confirm.status_code == 200

    second_confirm = client.post(
        f"/workflow/classes/{class_id}/sessions/{session_id}/confirm",
        headers=headers,
        json={"generate_session_writeup": True},
    )
    assert second_confirm.status_code == 409
    assert second_confirm.json()["detail"] == "This session is already confirmed."


def test_workflow_blueprint_records_requested_provider_when_falling_back(client, monkeypatch):
    from app import config as app_config
    from app.services import workflow_generation

    monkeypatch.setattr(app_config, "UNIT_PLANNER_PROVIDER", "notebooklm")
    async def _fake_client_unavailable():
        return None
    monkeypatch.setattr(workflow_generation, "_create_notebooklm_client", _fake_client_unavailable)
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Blueprint Provider Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Calcul litteral",
            "source_text": "1. Developpement\n1.1 Distributivite\nPropriete: k(a+b)=ka+kb\n",
        },
    )
    assert unit_resp.status_code == 201
    unit_id = unit_resp.json()["id"]

    blueprint_resp = client.get(f"/workflow/classes/{class_id}/units/{unit_id}/blueprint", headers=headers)
    assert blueprint_resp.status_code == 200
    blueprint = blueprint_resp.json()
    assert blueprint["provider"] == "fallback"
    assert blueprint["status"] == "degraded"
    assert blueprint["error_message"] == "notebooklm_client_unavailable"
    assert blueprint["raw_provider_response"]["requested_provider"] == "notebooklm"
    assert blueprint["raw_provider_response"]["error_message"] == "notebooklm_client_unavailable"


def test_workflow_unit_start_cleans_up_after_processing_failure(client, monkeypatch):
    from app.routers import workflow as workflow_router

    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Start Failure Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    def _boom(*args, **kwargs):
        raise RuntimeError("provider exploded")

    monkeypatch.setattr(workflow_router, "generate_unit_checklist", _boom)

    start_unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Failure chapter",
            "source_text": "1. Introduction\n2. Practice",
        },
    )
    assert start_unit_resp.status_code == 500
    assert "Unit start failed while processing the source content (RuntimeError)." == start_unit_resp.json()["detail"]

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_resp.status_code == 200
    assert workspace_resp.json().get("active_unit") is None


def test_workflow_unit_reextract_updates_checklist_and_blueprint(client, monkeypatch):
    from app.routers import workflow as workflow_router

    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Reextract Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Les nombres rationnels",
            "source_text": (
                "Les nombres rationnels : Somme et difference\n"
                "Objectifs d'apprentissage\n"
                "Contenu de la lecon\n"
            ),
        },
    )
    assert unit_resp.status_code == 201
    unit_id = unit_resp.json()["id"]

    monkeypatch.setattr(
        workflow_router,
        "generate_unit_checklist",
        lambda **kwargs: {
            "source": "notebooklm",
            "requested_provider": "notebooklm",
            "model": "notebooklm-py",
            "status": "ready",
            "error_message": None,
                "provider_context": {
                    "provider": "notebooklm",
                    "notebook_id": "nb-rerun-1",
                    "source_ids": ["src-rerun-1"],
                    "notebook_title": "Teacher Progress - Les nombres rationnels",
                    "notebook_role": "chapter_outline",
                },
            "unit_map": {
                "unit_title": "Les nombres rationnels : Somme et difference",
                "unit_type": "chapter",
                "source_mode": "notebooklm-unit-map",
                "teaching_goals": ["Identifier la progression de l'unite"],
                "prerequisites": [],
                "teacher_resources": [],
                "activity_blocks": ["Activites"],
                "assessment_blocks": ["Evaluation"],
                "pedagogy_notes": ["Commencer par les activites avant la lecon."],
                "ordered_outline": [
                    {
                        "title": "Les nombres rationnels : Somme et difference",
                        "kind": "chapter",
                        "children": [
                            {"title": "Activites", "kind": "section", "children": []},
                            {
                                "title": "Contenu de la lecon",
                                "kind": "section",
                                "children": [
                                    {"title": "I- Addition", "kind": "subsection", "children": []},
                                    {"title": "Evaluation", "kind": "exercise", "children": []},
                                ],
                            },
                        ],
                    }
                ],
                "future_actions": ["checklist", "session_writeup", "ask_unit", "slide_outline"],
            },
            "raw_provider_response": {
                "responses": [
                    {
                        "variant": "outline",
                        "prompt": "Extract the pedagogical outline only.",
                        "answer": "- Les nombres rationnels : Somme et difference\n  - Activites\n  - Contenu de la lecon\n",
                    }
                ]
            },
            "items": [
                {
                    "title": "Les nombres rationnels : Somme et difference",
                    "kind": "chapter",
                    "children": [
                        {"title": "Activites", "kind": "section", "children": []},
                        {
                            "title": "Contenu de la lecon",
                            "kind": "section",
                            "children": [
                                {"title": "I- Addition", "kind": "subsection", "children": []},
                                {"title": "Evaluation", "kind": "exercise", "children": []},
                            ],
                        },
                    ],
                }
            ],
        },
    )
    monkeypatch.setattr(workflow_router, "delete_provider_unit_context", lambda provider_context: True)

    reextract_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/reextract",
        headers=headers,
    )
    assert reextract_resp.status_code == 200
    payload = reextract_resp.json()
    assert payload["extraction_source"] == "notebooklm"
    assert payload["extraction_model"] == "notebooklm-py"
    assert payload["checklist"][0]["title"] == "Les nombres rationnels : Somme et difference"
    assert payload["checklist"][0]["children"][0]["title"] == "Activites"
    assert payload["checklist"][0]["children"][1]["children"][0]["title"] == "I- Addition"

    blueprint_resp = client.get(f"/workflow/classes/{class_id}/units/{unit_id}/blueprint", headers=headers)
    assert blueprint_resp.status_code == 200
    blueprint = blueprint_resp.json()
    assert blueprint["provider"] == "notebooklm"
    assert blueprint["model"] == "notebooklm-py"
    assert blueprint["blueprint_json"]["provider_context"]["notebook_id"] == "nb-rerun-1"
    assert blueprint["blueprint_json"]["provider_context"]["notebook_role"] == "chapter_outline"
    assert blueprint["unit_map_json"]["source_mode"] == "notebooklm-unit-map"
    assert blueprint["unit_map_json"]["activity_blocks"] == ["Activites"]
    assert blueprint["raw_provider_response"]["raw_provider_response"]["responses"][0]["prompt"] == "Extract the pedagogical outline only."


def test_workflow_unit_assistant_returns_guided_notebook_response(client, monkeypatch):
    from app.routers import workflow as workflow_router

    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Assistant Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Les nombres rationnels",
            "source_text": (
                "Les nombres rationnels : Somme et difference\n"
                "Activites\n"
                "Contenu de la lecon\n"
                "Evaluation\n"
            ),
        },
    )
    assert unit_resp.status_code == 201
    unit_id = unit_resp.json()["id"]

    captured: dict[str, object] = {}

    def _fake_generate_unit_assistant_package(**kwargs):
        captured.update(kwargs)
        return {
            "provider": "notebooklm",
            "requested_provider": "notebooklm",
            "model": "notebooklm-py",
            "status": "ready",
            "section_title": "1) Les denominateurs sont les memes",
            "section_path": ["I- Addition", "1) Les denominateurs sont les memes"],
            "action": "generate_harder_practice",
            "title": "Practice extension",
            "answer_rows": [
                "Propose two harder fraction additions with unlike denominators.",
                "Ask students to justify the common denominator they choose.",
            ],
            "suggested_followups": [
                "Generate a quick correction for these harder tasks.",
                "Prepare one oral warm-up before the harder exercises.",
            ],
            "source_payload": {
                "teacher_request": "Give me harder practice for this section.",
                "section_title": "1) Les denominateurs sont les memes",
                "section_path": ["I- Addition", "1) Les denominateurs sont les memes"],
                "action": "generate_harder_practice",
            },
            "raw_provider_response": {"answer": "{\"title\":\"Practice extension\"}"},
            "error_message": None,
        }

    monkeypatch.setattr(workflow_router, "generate_unit_assistant_package", _fake_generate_unit_assistant_package)

    assistant_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/assistant",
        headers=headers,
        json={
            "section_title": "1) Les denominateurs sont les memes",
            "section_path": ["I- Addition", "1) Les denominateurs sont les memes"],
            "action": "generate_harder_practice",
            "teacher_request": "Give me harder practice for this section.",
        },
    )
    assert assistant_resp.status_code == 200
    payload = assistant_resp.json()
    assert payload["provider"] == "notebooklm"
    assert payload["status"] == "ready"
    assert payload["action"] == "generate_harder_practice"
    assert payload["title"] == "Practice extension"
    assert len(payload["answer_rows"]) == 2
    assert payload["suggested_followups"][0].startswith("Generate a quick correction")

    assert captured["provider"] == "notebooklm"
    assert captured["teacher_request"] == "Give me harder practice for this section."
    assert captured["section_title"] == "1) Les denominateurs sont les memes"
    assert captured["section_path"] == ["I- Addition", "1) Les denominateurs sont les memes"]
    assert isinstance(captured["unit_map"], dict)


def test_workflow_unit_assistant_artifact_save_and_download(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Assistant Artifact Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Les nombres rationnels",
            "source_text": (
                "Les nombres rationnels : Somme et difference\n"
                "Activites\n"
                "Contenu de la lecon\n"
                "Evaluation\n"
            ),
        },
    )
    assert unit_resp.status_code == 201
    unit_id = unit_resp.json()["id"]

    save_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/assistant/artifacts",
        headers=headers,
        json={
            "artifact_kind": "guided_practice",
            "provider": "notebooklm",
            "model": "notebooklm-py",
            "section_title": "1) Les denominateurs sont les memes",
            "section_path": ["I- Addition", "1) Les denominateurs sont les memes"],
            "action": "generate_harder_practice",
            "title": "Practice extension",
            "answer_rows": [
                "Propose two harder fraction additions with unlike denominators.",
                "Ask students to justify the common denominator they choose.",
            ],
            "suggested_followups": [
                "Generate a quick correction for these harder tasks.",
            ],
            "source_payload": {"teacher_request": "Give me harder practice for this section."},
            "raw_provider_response": {"answer": "{\"title\":\"Practice extension\"}"},
        },
    )
    assert save_resp.status_code == 200
    artifact = save_resp.json()
    artifact_id = artifact["id"]
    assert artifact["artifact_kind"] == "guided_practice"
    assert artifact["section_title"] == "1) Les denominateurs sont les memes"
    assert "## Guidance" in artifact["content_markdown"]

    list_resp = client.get(
        f"/workflow/classes/{class_id}/units/{unit_id}/assistant/artifacts",
        headers=headers,
    )
    assert list_resp.status_code == 200
    rows = list_resp.json()
    assert len(rows) == 1
    assert rows[0]["artifact_kind"] == "guided_practice"

    download_resp = client.get(
        f"/workflow/classes/{class_id}/units/{unit_id}/assistant/artifacts/{artifact_id}/download",
        headers=headers,
    )
    assert download_resp.status_code == 200
    assert "text/markdown" in str(download_resp.headers.get("content-type") or "").lower()
    assert "guided-practice" in str(download_resp.headers.get("content-disposition") or "").lower()
    assert "Propose two harder fraction additions" in download_resp.text

    delete_resp = client.delete(
        f"/workflow/classes/{class_id}/units/{unit_id}/assistant/artifacts/{artifact_id}",
        headers=headers,
    )
    assert delete_resp.status_code == 200
    assert delete_resp.json() == {"ok": True}

    list_after_delete_resp = client.get(
        f"/workflow/classes/{class_id}/units/{unit_id}/assistant/artifacts",
        headers=headers,
    )
    assert list_after_delete_resp.status_code == 200
    assert list_after_delete_resp.json() == []


def test_workflow_session_writeup_can_import_saved_assistant_guidance(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Assistant Import Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    roster_content = _build_roster_file([("A1", "Student One")])
    roster_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("roster.xlsx", roster_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert roster_resp.status_code == 200

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Les nombres rationnels",
            "source_text": (
                "Les nombres rationnels : Somme et difference\n"
                "I- Addition\n"
                "1) Les denominateurs sont les memes\n"
                "Exemples\n"
                "Exercices\n"
            ),
        },
    )
    assert unit_resp.status_code == 201
    unit_id = unit_resp.json()["id"]

    start_session_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/start",
        headers=headers,
        json={"absent_student_ids": []},
    )
    assert start_session_resp.status_code == 201
    session_id = start_session_resp.json()["id"]

    save_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/assistant/artifacts",
        headers=headers,
        json={
            "artifact_kind": "guided_practice",
            "provider": "notebooklm",
            "model": "notebooklm-py",
            "section_title": "1) Les denominateurs sont les memes",
            "section_path": ["I- Addition", "1) Les denominateurs sont les memes"],
            "action": "generate_harder_practice",
            "title": "Harder fraction practice",
            "answer_rows": [
                "Calcule 7/12 + 5/18 en justifiant le PPCM choisi.",
                "Calcule 11/15 + 7/10 puis simplifie le resultat.",
            ],
            "suggested_followups": ["Prepare a short correction for these items."],
            "source_payload": {"teacher_request": "Give me harder practice."},
            "raw_provider_response": {"answer": "{\"title\":\"Harder fraction practice\"}"},
        },
    )
    assert save_resp.status_code == 200
    artifact_id = save_resp.json()["id"]

    import_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/{session_id}/writeup/import-assistant-artifact",
        headers=headers,
        json={"artifact_id": artifact_id},
    )
    assert import_resp.status_code == 200
    writeup = import_resp.json()
    assert writeup["approved"] is False
    assert "1) Les denominateurs sont les memes" in writeup["learning_focus"]
    assert "Calcule 7/12 + 5/18 en justifiant le PPCM choisi." in writeup["practice_items"]
    assert writeup["source_payload"]["imported_assistant_artifacts"][0]["artifact_id"] == artifact_id


def test_workflow_unit_material_generation_persists_study_guide(client, monkeypatch):
    from app.routers import workflow as workflow_router

    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Material Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Les nombres rationnels",
            "source_text": (
                "Les nombres rationnels : Somme et difference\n"
                "Activites\n"
                "Contenu de la lecon\n"
                "Evaluation\n"
            ),
        },
    )
    assert unit_resp.status_code == 201
    unit_id = unit_resp.json()["id"]

    captured: dict[str, object] = {}

    def _fake_generate_unit_material_package(**kwargs):
        captured.update(kwargs)
        return {
            "provider": "notebooklm",
            "requested_provider": "notebooklm",
            "model": "notebooklm-py",
            "status": "ready",
            "material_type": "study_guide",
            "title": "Study guide",
            "notebook_artifact_id": "artifact-study-guide-1",
            "source_payload": {
                "provider_context": {"provider": "notebooklm", "notebook_id": "nb-unit-1", "source_ids": ["src-1"]},
            },
            "content_markdown": "# Guide d'etude\n\n## Section 1\n- Point cle",
            "raw_provider_response": {"completion": {"status": "completed"}},
            "error_message": None,
        }

    monkeypatch.setattr(workflow_router, "generate_unit_material_package", _fake_generate_unit_material_package)

    generate_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/materials/generate",
        headers=headers,
        json={"material_type": "study_guide"},
    )
    assert generate_resp.status_code == 200
    payload = generate_resp.json()
    assert payload["provider"] == "notebooklm"
    assert payload["material_type"] == "study_guide"
    assert payload["notebook_artifact_id"] == "artifact-study-guide-1"
    assert payload["content_markdown"].startswith("# Guide d'etude")

    materials_resp = client.get(
        f"/workflow/classes/{class_id}/units/{unit_id}/materials",
        headers=headers,
    )
    assert materials_resp.status_code == 200
    rows = materials_resp.json()
    assert len(rows) == 1
    assert rows[0]["material_type"] == "study_guide"
    assert rows[0]["content_markdown"].startswith("# Guide d'etude")

    assert captured["provider"] == "notebooklm"
    assert captured["material_type"] == "study_guide"
    assert isinstance(captured["unit_map"], dict)


def test_workflow_unit_material_generation_persists_formative_quiz(client, monkeypatch):
    from app.routers import workflow as workflow_router

    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Quiz Material Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Les nombres rationnels",
            "source_text": (
                "Les nombres rationnels : Somme et difference\n"
                "Activites\n"
                "Contenu de la lecon\n"
                "Evaluation\n"
            ),
        },
    )
    assert unit_resp.status_code == 201
    unit_id = unit_resp.json()["id"]

    captured: dict[str, object] = {}

    def _fake_generate_unit_material_package(**kwargs):
        captured.update(kwargs)
        return {
            "provider": "notebooklm",
            "requested_provider": "notebooklm",
            "model": "notebooklm-py",
            "status": "ready",
            "material_type": "formative_quiz",
            "title": "Formative quiz",
            "notebook_artifact_id": "artifact-quiz-1",
            "source_payload": {
                "provider_context": {"provider": "notebooklm", "notebook_id": "nb-unit-1", "source_ids": ["src-1"]},
                "notebooklm_method": "generate_quiz",
            },
            "content_markdown": "# Quiz formatif\n\n- Question 1",
            "raw_provider_response": {"completion": {"status": "completed"}},
            "error_message": None,
        }

    monkeypatch.setattr(workflow_router, "generate_unit_material_package", _fake_generate_unit_material_package)

    generate_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/materials/generate",
        headers=headers,
        json={"material_type": "formative_quiz"},
    )
    assert generate_resp.status_code == 200
    payload = generate_resp.json()
    assert payload["material_type"] == "formative_quiz"
    assert payload["notebook_artifact_id"] == "artifact-quiz-1"
    assert payload["content_markdown"].startswith("# Quiz formatif")

    materials_resp = client.get(
        f"/workflow/classes/{class_id}/units/{unit_id}/materials",
        headers=headers,
    )
    assert materials_resp.status_code == 200
    rows = materials_resp.json()
    assert len(rows) == 1
    assert rows[0]["material_type"] == "formative_quiz"
    assert rows[0]["source_payload"]["notebooklm_method"] == "generate_quiz"

    assert captured["material_type"] == "formative_quiz"
    assert "content_blocks" in captured


def test_workflow_unit_material_download_returns_markdown(client, monkeypatch):
    from app.routers import workflow as workflow_router

    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Material Download Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Les nombres rationnels",
            "source_text": (
                "Les nombres rationnels : Somme et difference\n"
                "Activites\n"
                "Contenu de la lecon\n"
                "Evaluation\n"
            ),
        },
    )
    assert unit_resp.status_code == 201
    unit_id = unit_resp.json()["id"]

    def _fake_generate_unit_material_package(**kwargs):
        return {
            "provider": "notebooklm",
            "requested_provider": "notebooklm",
            "model": "notebooklm-py",
            "status": "ready",
            "material_type": "study_guide",
            "title": "Study guide",
            "notebook_artifact_id": "artifact-study-guide-1",
            "source_payload": {
                "provider_context": {"provider": "notebooklm", "notebook_id": "nb-unit-1", "source_ids": ["src-1"]},
            },
            "content_markdown": "# Guide d'etude\n\n## Section 1\n- Point cle",
            "raw_provider_response": {"completion": {"status": "completed"}},
            "error_message": None,
        }

    monkeypatch.setattr(workflow_router, "generate_unit_material_package", _fake_generate_unit_material_package)

    generate_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/materials/generate",
        headers=headers,
        json={"material_type": "study_guide"},
    )
    assert generate_resp.status_code == 200
    material_id = generate_resp.json()["id"]

    download_resp = client.get(
        f"/workflow/classes/{class_id}/units/{unit_id}/materials/{material_id}/download",
        headers=headers,
    )
    assert download_resp.status_code == 200
    assert "text/markdown" in str(download_resp.headers.get("content-type") or "").lower()
    disposition = str(download_resp.headers.get("content-disposition") or "")
    assert "attachment;" in disposition.lower()
    assert ".md" in disposition.lower()
    assert "# Guide d'etude" in download_resp.text


def test_workflow_unit_material_download_returns_slide_deck_file(client, monkeypatch, tmp_path):
    from app.routers import workflow as workflow_router

    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Slide Deck Material Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Les nombres rationnels",
            "source_text": (
                "Les nombres rationnels : Somme et difference\n"
                "Activites\n"
                "Contenu de la lecon\n"
                "Evaluation\n"
            ),
        },
    )
    assert unit_resp.status_code == 201
    unit_id = unit_resp.json()["id"]

    artifact_path = tmp_path / "presenter-slides.pptx"
    artifact_path.write_bytes(b"PPTX-DATA")

    def _fake_generate_unit_material_package(**kwargs):
        return {
            "provider": "notebooklm",
            "requested_provider": "notebooklm",
            "model": "notebooklm-py",
            "status": "ready",
            "material_type": "presenter_slides",
            "title": "Presenter slide deck",
            "notebook_artifact_id": "artifact-slides-1",
            "source_payload": {
                "provider_context": {"provider": "notebooklm", "notebook_id": "nb-unit-1", "source_ids": ["src-1"]},
                "notebooklm_method": "generate_slide_deck",
            },
            "content_markdown": None,
            "file_path": str(artifact_path),
            "file_name": "presenter-slides.pptx",
            "file_content_type": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
            "raw_provider_response": {"completion": {"status": "completed"}},
            "error_message": None,
        }

    monkeypatch.setattr(workflow_router, "generate_unit_material_package", _fake_generate_unit_material_package)

    generate_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/materials/generate",
        headers=headers,
        json={"material_type": "presenter_slides"},
    )
    assert generate_resp.status_code == 200
    payload = generate_resp.json()
    material_id = payload["id"]
    assert payload["material_type"] == "presenter_slides"
    assert payload["file_name"] == "presenter-slides.pptx"
    assert payload["content_markdown"] is None

    download_resp = client.get(
        f"/workflow/classes/{class_id}/units/{unit_id}/materials/{material_id}/download",
        headers=headers,
    )
    assert download_resp.status_code == 200
    assert (
        "application/vnd.openxmlformats-officedocument.presentationml.presentation"
        in str(download_resp.headers.get("content-type") or "").lower()
    )
    disposition = str(download_resp.headers.get("content-disposition") or "")
    assert "presenter-slides.pptx" in disposition
    assert download_resp.content == b"PPTX-DATA"


def test_workflow_unit_reextract_is_blocked_after_teaching_starts(client):
    from app.database import SessionLocal
    from app.models import ClassSession

    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Blocked Reextract Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    roster_content = _build_roster_file([("A1", "Student One")])
    roster_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("roster.xlsx", roster_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert roster_resp.status_code == 200

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Fractions",
            "source_text": "Fractions\nObjectifs\nExercices\n",
        },
    )
    assert unit_resp.status_code == 201
    unit_id = unit_resp.json()["id"]

    db = SessionLocal()
    try:
        db.add(
            ClassSession(
                class_id=int(class_id),
                unit_id=int(unit_id),
                unit_session_number=1,
                session_date=date(2026, 5, 17),
                start_time=None,
                end_time=None,
                note="Linked for reextract guard test",
            )
        )
        db.commit()
    finally:
        db.close()

    reextract_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/reextract",
        headers=headers,
    )
    assert reextract_resp.status_code == 409
    assert "before teaching starts" in reextract_resp.json()["detail"]


def test_owner_can_run_notebooklm_smoke_test(client, monkeypatch):
    from app.routers import ops as ops_router

    headers = _auth_headers(client)
    monkeypatch.setattr(
        ops_router,
        "notebooklm_smoke_test",
        lambda: {
            "ok": True,
            "provider": "notebooklm",
            "model": "notebooklm-py",
            "error_message": None,
            "answer": "OK",
            "notebook_id": "nb-test",
            "source_ids": ["src-test"],
        },
    )

    resp = client.post("/ops/notebooklm/smoke-test", headers=headers)
    assert resp.status_code == 200
    payload = resp.json()
    assert "status" in payload
    assert "smoke" in payload
    assert payload["smoke"]["ok"] is True
    assert payload["smoke"]["answer"] == "OK"


def test_owner_can_cleanup_notebooklm_temp_notebooks(client, monkeypatch):
    from app.routers import ops as ops_router

    headers = _auth_headers(client)
    monkeypatch.setattr(
        ops_router,
        "notebooklm_cleanup_temp_notebooks",
        lambda: {
            "ok": True,
            "error_message": None,
            "deleted_count": 2,
            "deleted_titles": ["Teacher Progress Smoke Test", "Teacher Progress Smoke Test"],
        },
    )

    resp = client.post("/ops/notebooklm/cleanup-temp", headers=headers)
    assert resp.status_code == 200
    payload = resp.json()
    assert "status" in payload
    assert "cleanup" in payload
    assert payload["cleanup"]["ok"] is True
    assert payload["cleanup"]["deleted_count"] == 2


def test_workflow_writeup_requires_notebooklm_auth_when_requested(client, monkeypatch):
    from app import config as app_config

    monkeypatch.setattr(app_config, "SESSION_WRITER_PROVIDER", "notebooklm")
    headers = _auth_headers(client)
    _close_any_active_unit(client, headers)
    class_resp = client.post("/classes", json={"name": "Writeup Provider Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    roster_content = _build_roster_file([("A1", "Student One")])
    roster_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("roster.xlsx", roster_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert roster_resp.status_code == 200

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Factorisation",
            "source_text": "1. Factorisation\n1.1 Mise en facteur commun\nApplication: Exercices de base\n",
        },
    )
    assert unit_resp.status_code == 201

    start_session_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/start",
        headers=headers,
        json={"absent_student_ids": []},
    )
    assert start_session_resp.status_code == 201
    session_id = start_session_resp.json()["id"]

    writeup_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/{session_id}/writeup/generate",
        headers=headers,
        json={"regenerate": True},
    )
    assert writeup_resp.status_code == 409
    assert "NotebookLM" in str(writeup_resp.json().get("detail", ""))


def test_workflow_writeup_can_be_edited_after_generation(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Writeup Edit Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    roster_content = _build_roster_file([("A1", "Student One")])
    roster_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("roster.xlsx", roster_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert roster_resp.status_code == 200

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Theoreme de Pythagore",
            "source_text": "1. Theoreme de Pythagore\n1.1 Enonce\nExemple: Triangle rectangle\nApplication: Exercices directs\n",
        },
    )
    assert unit_resp.status_code == 201

    start_session_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/start",
        headers=headers,
        json={"absent_student_ids": []},
    )
    assert start_session_resp.status_code == 201
    session_id = start_session_resp.json()["id"]

    generate_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/{session_id}/writeup/generate",
        headers=headers,
        json={"regenerate": True},
    )
    assert generate_resp.status_code == 200

    update_resp = client.patch(
        f"/workflow/classes/{class_id}/sessions/{session_id}/writeup",
        headers=headers,
        json={
            "title": "Seance 1 - Pythagore en triangle rectangle",
            "learning_focus": ["Enoncer le theoreme", "Identifier le cote oppose a l'angle droit"],
            "teaching_content": [
                "La seance a introduit le theoreme de Pythagore dans le cas du triangle rectangle.",
                "Les eleves ont applique la relation a des exemples simples avant un entrainement guide.",
            ],
            "practice_items": ["Exercices directs sur le calcul d'une longueur"],
            "approved": False,
        },
    )
    assert update_resp.status_code == 200
    updated = update_resp.json()
    assert updated["title"] == "Seance 1 - Pythagore en triangle rectangle"
    assert updated["learning_focus"][0] == "Enoncer le theoreme"
    assert updated["practice_items"] == ["Exercices directs sur le calcul d'une longueur"]
    assert updated["approved"] is False


def test_owner_can_read_notebooklm_status(client):
    headers = _auth_headers(client)
    resp = client.get('/ops/notebooklm/status', headers=headers)
    assert resp.status_code == 200
    payload = resp.json()
    assert "installed" in payload
    assert "ready" in payload
    assert "auth_path" in payload
    assert "profile" in payload
    assert "runtime_health" in payload


def test_owner_can_download_notebooklm_refresh_helper(client):
    headers = _auth_headers(client)
    resp = client.get('/ops/notebooklm/refresh-helper.cmd', headers=headers)
    assert resp.status_code == 200
    content = resp.text
    assert "python -m notebooklm login" in content
    assert "/ops/notebooklm/auth/upload" in content
    assert "/ops/notebooklm/smoke-test" in content
    assert "refresh_notebooklm.cmd" in str(resp.headers.get("content-disposition", ""))


def test_owner_can_upload_and_clear_notebooklm_auth_file(client, monkeypatch, tmp_path):
    import app.routers.ops as ops_router
    from app.services import workflow_generation

    auth_path = tmp_path / "notebooklm" / "profiles" / "default" / "storage_state.json"
    monkeypatch.setattr(ops_router, "NOTEBOOKLM_AUTH_PATH", str(auth_path))
    monkeypatch.setattr(workflow_generation.app_config, "NOTEBOOKLM_AUTH_PATH", str(auth_path))
    headers = _auth_headers(client)

    upload_resp = client.post(
        '/ops/notebooklm/auth/upload',
        headers=headers,
        files={"file": ("storage_state.json", b'{"cookies": [], "origins": []}', "application/json")},
    )
    assert upload_resp.status_code == 200
    payload = upload_resp.json()
    assert payload["auth_file_exists"] is True
    assert payload["auth_file_valid"] is True
    assert payload["runtime_health"]["refresh_required"] is False
    assert payload["runtime_health"]["last_manual_refresh_at"] is not None
    assert auth_path.exists() is True

    clear_resp = client.post('/ops/notebooklm/auth/clear', headers=headers)
    assert clear_resp.status_code == 200
    cleared = clear_resp.json()
    assert cleared["auth_file_exists"] is False
    assert cleared["runtime_health"]["refresh_required"] is True
    assert cleared["runtime_health"]["last_manual_clear_at"] is not None
    assert auth_path.exists() is False


def test_checklist_postprocess_splits_verbose_extraction_items():
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    raw_items = [
        {
            "title": "introduction-aux-nombres-rationnels-cours-ma",
            "kind": "chapter",
            "children": [
                {
                    "title": (
                        "Chapitre1:Introduction des nombres rationnels Nombre rationnel 1: "
                        "Définition Un nombre rationnel est un nombre qui peut s'exprimer sous la forme "
                        "du quotient de deux nombres entiers. 2.Exemples les nombres 7-3 et 7/3. "
                        "3.Remarques: Tout nombre entier relatif est un nombre rationnel."
                    ),
                    "kind": "example",
                    "children": [],
                }
            ],
        }
    ]

    normalized = workflow_generation._postprocess_checklist_items(
        raw_items,
        unit_type=WorkflowUnitType.CHAPTER,
        unit_title="introduction-aux-nombres-rationnels-cours-ma",
    )

    assert normalized
    titles = [str(item["title"]) for item in normalized]
    assert titles[0] != "introduction-aux-nombres-rationnels-cours-ma"
    assert any("Definition" in title for title in titles)
    assert any("Exemples" in title for title in titles)
    assert any("Remarques" in title for title in titles)
    assert all(len(title) <= 120 for title in titles)


def test_notebooklm_prompt_omits_noisy_source_hint_when_pdf_is_attached():
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    prompt = workflow_generation._build_notebooklm_checklist_prompt(
        unit_type=WorkflowUnitType.CHAPTER,
        title="Nombres rationnels",
        source_hint="",
        session_count=6,
    )

    assert "Indice textuel de secours" not in prompt
    assert "Retourne la liste hierarchique de tous les headlines pedagogiques visibles utiles pour enseigner." in prompt
    assert "Garde uniquement les headlines visibles; ne reorganise pas l'ordre" in prompt
    assert "garde-la seulement si elle apparait comme un vrai headline du document" in prompt
    assert "N'inclus pas les rubriques meta enseignant" in prompt
    assert "indentation de deux espaces par niveau" in prompt
    assert "aucun commentaire avant ou apres la liste" in prompt


def test_notebooklm_role_mapping_separates_unit_types():
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    assert workflow_generation._notebooklm_role_for_unit_type(WorkflowUnitType.CHAPTER) == "chapter_outline"
    assert workflow_generation._notebooklm_role_for_unit_type(WorkflowUnitType.EXERCISE_SERIES) == "exercise_outline"
    assert workflow_generation._notebooklm_role_for_unit_type(WorkflowUnitType.EXAM) == "exam_outline"
    assert workflow_generation._notebooklm_role_for_unit_type(WorkflowUnitType.EXAM_CORRECTION) == "correction_outline"


def test_normalize_exercise_series_headline_title_keeps_compound_exercise_names():
    from app.services import workflow_generation

    assert workflow_generation._normalize_exercise_series_headline_title("EXERCICE 2") == "Exercice 2"
    assert workflow_generation._normalize_exercise_series_headline_title("EXERCICE2B.3-POLYNESIE2001") == "Exercice 2B.3 - POLYNESIE 2001"
    assert workflow_generation._normalize_exercise_series_headline_title(
        "EXERCICE 1 Calculer en donnant le résultat en écriture fractionnaire:"
    ) == "Exercice 1"


def test_notebooklm_prompt_for_exercise_series_preserves_exact_visible_exercise_titles():
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    prompt = workflow_generation._build_notebooklm_checklist_prompt(
        unit_type=WorkflowUnitType.EXERCISE_SERIES,
        title="Triangle, milieux et paralleles - exercices",
        source_hint="",
        session_count=6,
    )

    assert "Retourne uniquement le titre de la serie comme racine puis les headlines explicites des exercices visibles." in prompt
    assert "garde uniquement les headlines explicites des exercices visibles" in prompt.lower()
    assert "preserve le headline visible exact" in prompt.lower()
    assert "N'ajoute pas de sous-noeud sous un exercice." in prompt
    assert "N'inclus pas d'activites, definitions, remarques, exemples" in prompt
    assert "A l'interieur de chaque concept ou section" not in prompt


def test_notebooklm_prompt_omits_slug_like_title_hint():
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    prompt = workflow_generation._build_notebooklm_checklist_prompt(
        unit_type=WorkflowUnitType.CHAPTER,
        title="introduction-aux-nombres-rationnels-cours-ma",
        source_hint="",
        session_count=6,
    )

    assert "Titre fourni si utile:" not in prompt


def test_parse_notebooklm_outline_response_preserves_heading_tree():
    from app.models import WorkflowChecklistItemKind, WorkflowUnitType
    from app.services import workflow_generation

    answer = "\n".join(
        [
            "- **Chapitre 5 : Les nombres relatifs** [1]",
            "  - **5.1 Somme de deux nombres relatifs** [1]",
            "    - **5.1.1 Les deux nombres sont de même signe** [1]",
            "      - **Propriete** [1]",
            "      - **Exemples** [1]",
            "    - **5.1.2 Les deux nombres sont de signes contraires** [2]",
            "      - **Exemples** [2]",
            "  - **5.2 Différence de deux nombres relatifs** [3]",
            "    - **Activités** [3]",
        ]
    )

    items = workflow_generation._parse_notebooklm_outline_response(
        answer,
        unit_type=WorkflowUnitType.CHAPTER,
        unit_title="Les nombres relatifs",
    )

    assert items
    assert len(items) == 1
    chapter = items[0]
    assert chapter["kind"] == WorkflowChecklistItemKind.CHAPTER.value
    assert "Chapitre 5" in chapter["title"]
    assert len(chapter["children"]) == 2
    assert chapter["children"][0]["kind"] == WorkflowChecklistItemKind.SECTION.value
    assert chapter["children"][0]["title"].startswith("5.1")
    assert chapter["children"][0]["children"][0]["kind"] == WorkflowChecklistItemKind.SUBSECTION.value
    assert chapter["children"][0]["children"][0]["title"].startswith("5.1.1")
    assert chapter["children"][0]["children"][0]["children"][0]["kind"] == WorkflowChecklistItemKind.PROPERTY.value
    assert chapter["children"][0]["children"][0]["children"][1]["kind"] == WorkflowChecklistItemKind.EXAMPLE.value
    assert chapter["children"][1]["title"].startswith("5.2")
    assert "**" not in chapter["title"]
    assert "[" not in chapter["title"]


def test_parse_notebooklm_outline_response_keeps_document_order_without_resequencing():
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    answer = "\n".join(
        [
            "- Chapitre 1 : Produit et division",
            "  - II- Division de nombres rationnels :",
            "  - Activite 1 :",
            "  - I- Multiplication de nombres rationnels :",
        ]
    )

    items = workflow_generation._parse_notebooklm_outline_response(
        answer,
        unit_type=WorkflowUnitType.CHAPTER,
        unit_title="Produit et division",
    )

    assert items
    chapter = items[0]
    child_titles = [str(row.get("title") or "") for row in chapter.get("children", [])]
    assert child_titles == [
        "II- Division de nombres rationnels :",
        "I- Multiplication de nombres rationnels :",
    ]



def test_parse_notebooklm_outline_response_collapses_duplicate_parent_child_titles():
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    answer = "\n".join(
        [
            "- Les nombres rationnels : Produit et division",
            "  - I- Multiplication de nombres rationnels :",
            "    - I- Multiplication de nombres rationnels :",
            "      - 1) Inverse d'un nombre rationnel :",
        ]
    )

    items = workflow_generation._parse_notebooklm_outline_response(
        answer,
        unit_type=WorkflowUnitType.CHAPTER,
        unit_title="Produit et division",
    )

    assert items
    chapter = items[0]
    child_titles = [str(row.get("title") or "") for row in chapter.get("children", [])]
    assert child_titles == ["I- Multiplication de nombres rationnels :"]
    assert chapter["children"][0]["children"][0]["title"] == "1) Inverse d'un nombre rationnel :"


def test_parse_notebooklm_outline_response_drops_generic_buckets_and_itemized_rows_for_chapters():
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    answer = "\n".join(
        [
            "- Les nombres rationnels : Somme et difference",
            "  - Activites",
            "    - Activite 1 :",
            "    - Activite 2 :",
            "  - Contenu de la lecon",
            "    - I- Addition et soustraction de deux nombres rationnels :",
            "      - 1) Les denominateurs sont les memes :",
            "  - Evaluation",
            "    - Exercice 1 :",
        ]
    )

    items = workflow_generation._parse_notebooklm_outline_response(
        answer,
        unit_type=WorkflowUnitType.CHAPTER,
        unit_title="Les nombres rationnels : Somme et difference",
    )

    assert items
    chapter = items[0]
    child_titles = [str(row.get("title") or "") for row in chapter.get("children", [])]
    assert "Activites" not in child_titles
    assert "Contenu de la lecon" not in child_titles
    assert "Evaluation" not in child_titles
    assert "Activite 1 :" not in child_titles
    assert "Exercice 1 :" not in child_titles
    assert child_titles == ["I- Addition et soustraction de deux nombres rationnels :"]


def test_parse_notebooklm_outline_response_keeps_itemized_rows_when_nested_under_real_concepts():
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    answer = "\n".join(
        [
            "- Les nombres rationnels : Produit et division",
            "  - I- Multiplication de nombres rationnels :",
            "    - Activite 1 :",
            "    - Regle :",
            "    - Exercice 1 :",
            "  - II- Division de nombres rationnels :",
            "    - 1) Inverse d'un nombre rationnel :",
            "      - Activite 2 :",
            "      - Definition :",
            "      - Exercice 2 :",
        ]
    )

    items = workflow_generation._parse_notebooklm_outline_response(
        answer,
        unit_type=WorkflowUnitType.CHAPTER,
        unit_title="Les nombres rationnels : Produit et division",
    )

    assert items
    chapter = items[0]
    first_children = [str(row.get("title") or "") for row in chapter["children"][0]["children"]]
    second_children = [str(row.get("title") or "") for row in chapter["children"][1]["children"][0]["children"]]
    assert first_children == ["Activite 1 :", "Regle :", "Exercice 1 :"]
    assert second_children == ["Activite 2 :", "Definition :", "Exercice 2 :"]


def test_parse_notebooklm_outline_response_minimizes_exercise_series_to_title_and_exercise_headlines():
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    answer = "\n".join(
        [
            "- Triangle, milieux et paralleles - exercices",
            "  - Triangle, milieux et paralleles - exercices",
            "  - Exercice 1 :",
            "    - Exercice 1 :",
            "  - Exercice 2 : (Ex13/14/17/18 - p48)",
            "    - Exercice 2 :",
            "  - Remarque :",
            "  - Exemple :",
        ]
    )

    items = workflow_generation._parse_notebooklm_outline_response(
        answer,
        unit_type=WorkflowUnitType.EXERCISE_SERIES,
        unit_title="Triangle, milieux et paralleles - exercices",
    )

    assert items
    assert len(items) == 1
    root = items[0]
    assert root["title"] == "Triangle, milieux et paralleles - exercices"
    assert root["kind"] == "section"
    child_titles = [str(row.get("title") or "") for row in root.get("children", [])]
    child_kinds = [str(row.get("kind") or "") for row in root.get("children", [])]
    assert child_titles == ["Exercice 1", "Exercice 2"]
    assert child_kinds == ["exercise", "exercise"]


def test_postprocess_checklist_preserves_exercise_series_root():
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    normalized = workflow_generation._postprocess_checklist_items(
        [
            {
                "title": "Triangle, milieux et paralleles - exercices",
                "kind": "section",
                "children": [
                    {"title": "Exercice 1", "kind": "exercise", "children": []},
                    {"title": "Exercice 2", "kind": "exercise", "children": []},
                ],
            }
        ],
        unit_type=WorkflowUnitType.EXERCISE_SERIES,
        unit_title="Triangle, milieux et paralleles - exercices",
    )

    assert len(normalized) == 1
    assert normalized[0]["title"] == "Triangle, milieux et paralleles - exercices"
    assert [str(row.get("title") or "") for row in normalized[0].get("children", [])] == ["Exercice 1", "Exercice 2"]


def test_select_reference_outline_seed_prefers_richer_layout_for_exercise_series():
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    layout_seed = [
        {
            "title": "Serie",
            "kind": "section",
            "children": [
                {"title": "Exercice 1", "kind": "exercise", "children": []},
                {"title": "Exercice 2", "kind": "exercise", "children": []},
                {"title": "Exercice 3", "kind": "exercise", "children": []},
            ],
        }
    ]
    outline_seed = [
        {
            "title": "Serie",
            "kind": "section",
            "children": [
                {"title": "Exercice 1", "kind": "exercise", "children": []},
                {"title": "Exercice 2", "kind": "exercise", "children": []},
            ],
        }
    ]

    selected = workflow_generation._select_reference_outline_seed(
        layout_seed=layout_seed,
        outline_seed=outline_seed,
        unit_type=WorkflowUnitType.EXERCISE_SERIES,
        unit_title="Serie",
    )

    assert selected == layout_seed


def test_generate_unit_checklist_package_repairs_weak_exercise_series_outline_with_pdf_layout_seed(monkeypatch):
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    weak_items = [
        {
            "title": "Nombres relatifs en ecriture fractionnaire exercice 2B",
            "kind": "section",
            "children": [
                {"title": "Exercice 1", "kind": "exercise", "children": []},
                {"title": "Exercice 2", "kind": "exercise", "children": []},
            ],
        }
    ]
    repair_outline = [
        {
            "title": "Nombres relatifs en ecriture fractionnaire exercice 2B",
            "kind": "section",
            "children": [
                {"title": "Exercice 1", "kind": "exercise", "children": []},
                {"title": "Exercice 2", "kind": "exercise", "children": []},
                {"title": "Exercice 2B.3 - POLYNESIE 2001", "kind": "exercise", "children": []},
                {"title": "Exercice 2B.4 - AFRIQUE DU NORD 2001", "kind": "exercise", "children": []},
            ],
        }
    ]

    monkeypatch.setattr(
        workflow_generation,
        "_notebooklm_generate_checklist",
        lambda **kwargs: (
            weak_items,
            {"provider": "notebooklm", "notebook_id": "nb-test", "source_ids": ["src-1"], "notebook_role": "exercise_outline"},
            {"response_mode": "outline_only"},
            None,
        ),
    )
    monkeypatch.setattr(
        workflow_generation,
        "_build_pdf_layout_outline_seed",
        lambda **kwargs: repair_outline,
    )

    package = workflow_generation.generate_unit_checklist_package(
        unit_type=WorkflowUnitType.EXERCISE_SERIES,
        title="Nombres relatifs en ecriture fractionnaire exercice 2B",
        source_text="weak text",
        session_count=4,
        provider="notebooklm",
        document_path="fake.pdf",
    )

    repaired_children = package["items"][0]["children"]
    assert [str(row.get("title") or "") for row in repaired_children] == [
        "Exercice 1",
        "Exercice 2",
        "Exercice 2B.3 - POLYNESIE 2001",
        "Exercice 2B.4 - AFRIQUE DU NORD 2001",
    ]
    assert package["raw_provider_response"]["selected_structure_source"] == "pdf_layout_seed"


def test_candidate_needs_structural_repair_for_richer_exercise_series_titles():
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    candidate_items = [
        {
            "title": "Nombres relatifs en ecriture fractionnaire exercice 2B",
            "kind": "section",
            "children": [
                {"title": "Exercice 1", "kind": "exercise", "children": []},
                {"title": "Exercice 2", "kind": "exercise", "children": []},
                {"title": "Exercice 2B.3", "kind": "exercise", "children": []},
                {"title": "Exercice 2B.4", "kind": "exercise", "children": []},
            ],
        }
    ]
    reference_outline = [
        {
            "title": "Nombres relatifs en ecriture fractionnaire exercice 2B",
            "kind": "section",
            "children": [
                {"title": "Exercice 1", "kind": "exercise", "children": []},
                {"title": "Exercice 2", "kind": "exercise", "children": []},
                {"title": "Exercice 2B.3 - POLYNESIE 2001", "kind": "exercise", "children": []},
                {"title": "Exercice 2B.4 - AFRIQUE DU NORD 2001", "kind": "exercise", "children": []},
            ],
        }
    ]

    assert workflow_generation._candidate_needs_structural_repair(
        candidate_items,
        reference_outline=reference_outline,
        unit_type=WorkflowUnitType.EXERCISE_SERIES,
        unit_title="Nombres relatifs en ecriture fractionnaire exercice 2B",
    ) is True


def test_generate_unit_checklist_package_repairs_weaker_exact_exercise_titles_with_pdf_layout_seed(monkeypatch):
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    weak_exact_items = [
        {
            "title": "Nombres relatifs en ecriture fractionnaire exercice 2B",
            "kind": "section",
            "children": [
                {"title": "Exercice 1", "kind": "exercise", "children": []},
                {"title": "Exercice 2", "kind": "exercise", "children": []},
                {"title": "Exercice 2B.3", "kind": "exercise", "children": []},
                {"title": "Exercice 2B.4", "kind": "exercise", "children": []},
            ],
        }
    ]
    repair_outline = [
        {
            "title": "Nombres relatifs en ecriture fractionnaire exercice 2B",
            "kind": "section",
            "children": [
                {"title": "Exercice 1", "kind": "exercise", "children": []},
                {"title": "Exercice 2", "kind": "exercise", "children": []},
                {"title": "Exercice 2B.3 - POLYNESIE 2001", "kind": "exercise", "children": []},
                {"title": "Exercice 2B.4 - AFRIQUE DU NORD 2001", "kind": "exercise", "children": []},
            ],
        }
    ]

    monkeypatch.setattr(
        workflow_generation,
        "_notebooklm_generate_checklist",
        lambda **kwargs: (
            weak_exact_items,
            {"provider": "notebooklm", "notebook_id": "nb-test", "source_ids": ["src-1"], "notebook_role": "exercise_outline"},
            {"response_mode": "outline_only"},
            None,
        ),
    )
    monkeypatch.setattr(
        workflow_generation,
        "_build_pdf_layout_outline_seed",
        lambda **kwargs: repair_outline,
    )

    package = workflow_generation.generate_unit_checklist_package(
        unit_type=WorkflowUnitType.EXERCISE_SERIES,
        title="Nombres relatifs en ecriture fractionnaire exercice 2B",
        source_text="weak text",
        session_count=4,
        provider="notebooklm",
        document_path="fake.pdf",
    )

    repaired_children = package["items"][0]["children"]
    assert [str(row.get("title") or "") for row in repaired_children] == [
        "Exercice 1",
        "Exercice 2",
        "Exercice 2B.3 - POLYNESIE 2001",
        "Exercice 2B.4 - AFRIQUE DU NORD 2001",
    ]
    assert package["raw_provider_response"]["selected_structure_source"] == "pdf_layout_seed"


def test_select_best_notebooklm_outline_candidate_prefers_richer_exercise_titles():
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    primary_items = [
        {
            "title": "Serie",
            "kind": "section",
            "children": [
                {"title": "Exercice 1", "kind": "exercise", "children": []},
                {"title": "Exercice 2", "kind": "exercise", "children": []},
                {"title": "Exercice 2B.3", "kind": "exercise", "children": []},
                {"title": "Exercice 2B.4", "kind": "exercise", "children": []},
            ],
        }
    ]
    review_items = [
        {
            "title": "Serie",
            "kind": "section",
            "children": [
                {"title": "Exercice 1", "kind": "exercise", "children": []},
                {"title": "Exercice 2", "kind": "exercise", "children": []},
                {"title": "Exercice 2B.3 - POLYNESIE 2001", "kind": "exercise", "children": []},
                {"title": "Exercice 2B.4 - AFRIQUE DU NORD 2001", "kind": "exercise", "children": []},
            ],
        }
    ]

    name, items = workflow_generation._select_best_notebooklm_outline_candidate(
        [
            ("primary", primary_items),
            ("completeness_review", review_items),
        ],
        source_text="",
        unit_type=WorkflowUnitType.EXERCISE_SERIES,
        unit_title="Serie",
    )

    assert name == "completeness_review"
    assert items == review_items


def test_notebooklm_generate_checklist_surfaces_runtime_error_detail(monkeypatch):
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    def raise_value_error(coro):
        try:
            coro.close()
        except Exception:
            pass
        raise ValueError("test notebooklm runtime failure")

    monkeypatch.setattr(workflow_generation.asyncio, "run", raise_value_error)

    items, provider_context, raw_provider_response, error_message = workflow_generation._notebooklm_generate_checklist(
        unit_type=WorkflowUnitType.EXERCISE_SERIES,
        title="Serie",
        source_text="",
        session_count=4,
        document_path=None,
        outline_hint_lines=None,
    )

    assert items is None
    assert provider_context is None
    assert raw_provider_response is None
    assert error_message == "notebooklm_runtime_error:ValueError:test notebooklm runtime failure"


def test_pdf_text_extraction_preserves_line_break_structure():
    from app.services import workflow as workflow_service

    pdf_bytes = _build_pdf_file(
        [
            "Chapitre 1 : Operations sur les nombres relatifs",
            "1) Pour additionner deux nombres relatifs de meme signe",
            "Exemples - Somme de deux nombres positifs",
            "2) Pour additionner deux nombres de signes contraires",
        ]
    )
    with NamedTemporaryFile(suffix=".pdf", delete=False) as temp:
        temp.write(pdf_bytes)
        temp_path = temp.name
    try:
        extracted = workflow_service.extract_text_from_document(temp_path)
    finally:
        os.remove(temp_path)

    rows = [row.strip() for row in extracted.splitlines() if row.strip()]
    assert any("Chapitre 1" in row for row in rows)
    assert any("Pour additionner deux nombres relatifs de meme signe" in row for row in rows)
    assert any("Exemples - Somme de deux nombres positifs" in row for row in rows)
    assert len(rows) >= 4


def test_outline_seed_filters_metadata_and_keeps_teaching_structure():
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    source_text = "\n".join(
        [
            "Exemples - Mathématiques",
            "Exemples - 3eme Annee College Seance 1 (Operations sur les nombres relatifs) Professeur : MR BENGHANI",
            "1) Pour additionner deux nombres relatifs de meme signe : On garde le signe commun",
            "Exemples - Somme de deux nombres positifs",
            "2) Pour additionner deux nombres de signes contraires",
            "Exemples - On ecrit le signe du nombre qui a la plus grande distance a zero",
            "1) Pour multiplier deux nombres relatifs",
            "On effectue le produit des distances a zero",
        ]
    )

    items = workflow_generation._build_outline_seed(
        unit_type=WorkflowUnitType.CHAPTER,
        title="Operations sur les nombres relatifs",
        source_text=source_text,
    )
    flat = _flatten_checklist(items)
    titles = [str(row.get("title", "")) for row in flat]
    lower_titles = [title.lower() for title in titles]

    assert not any("prof" in title for title in lower_titles)
    assert not any("math" in title for title in lower_titles)
    assert any("pour additionner deux nombres relatifs de meme signe" in title.lower() for title in titles)
    assert any(title.startswith("Exemples -") for title in titles)
    assert any(title.startswith("Propriete -") for title in titles)


def test_generate_unit_checklist_package_prefers_structured_outline_over_noisy_provider(monkeypatch):
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    noisy_items = [
        {"title": "Exemples - Mathématiques", "kind": "example", "children": []},
        {
            "title": "Exemples - 3eme Annee College Seance 1 (Operations sur les nombres relatifs) Professeur : MR BENGHANI",
            "kind": "example",
            "children": [],
        },
    ]

    monkeypatch.setattr(
        workflow_generation,
        "_notebooklm_generate_checklist",
        lambda **kwargs: (noisy_items, {"provider": "notebooklm"}, {"answer": "{}"}, None),
    )

    package = workflow_generation.generate_unit_checklist_package(
        unit_type=WorkflowUnitType.CHAPTER,
        title="Operations sur les nombres relatifs",
        source_text="\n".join(
            [
                "1) Pour additionner deux nombres relatifs de meme signe : On garde le signe commun",
                "Exemples - Somme de deux nombres positifs",
                "2) Pour additionner deux nombres de signes contraires",
                "Exemples - On ecrit le signe du nombre qui a la plus grande distance a zero",
            ]
        ),
        provider="notebooklm",
    )

    flat = _flatten_checklist(package["items"])
    titles = [str(row.get("title", "")) for row in flat]
    assert any("pour additionner deux nombres relatifs de meme signe" in title.lower() for title in titles)
    assert not any("prof" in title.lower() for title in titles)
    assert not any(title.lower().startswith("exemples - mathematiques") for title in titles)


def test_pdf_layout_heading_selection_preserves_outline_order(monkeypatch, tmp_path):
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    fake_rows = [
        {"page": 1, "y": 640.0, "x": 56.7, "text": "Chapitre 5", "max_size": 24.8, "avg_size": 24.8, "bold": False},
        {"page": 1, "y": 594.0, "x": 56.7, "text": "Les nombres relatifs : addition et", "max_size": 24.8, "avg_size": 24.8, "bold": False},
        {"page": 1, "y": 564.0, "x": 56.7, "text": "soustraction", "max_size": 24.8, "avg_size": 24.8, "bold": False},
        {"page": 1, "y": 502.0, "x": 56.7, "text": "5.1 Somme de deux nombres relatifs", "max_size": 17.2, "avg_size": 17.2, "bold": False},
        {"page": 1, "y": 472.0, "x": 56.7, "text": "5.1.1 Les deux nombres sont de même signe", "max_size": 14.3, "avg_size": 14.3, "bold": False},
        {"page": 1, "y": 446.0, "x": 87.8, "text": "La somme de deux nombres positifs est un nombre positif", "max_size": 12.0, "avg_size": 12.0, "bold": False},
        {"page": 1, "y": 383.0, "x": 56.7, "text": "Exemples", "max_size": 12.0, "avg_size": 12.0, "bold": False},
        {"page": 1, "y": 323.0, "x": 56.7, "text": "5.1.2 Les deux nombres sont de signes contraires", "max_size": 14.3, "avg_size": 14.3, "bold": False},
    ]

    monkeypatch.setattr(workflow_generation, "_extract_pdf_layout_rows", lambda source: fake_rows)
    pdf_path = tmp_path / "outline.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n%test")

    items = workflow_generation._build_pdf_layout_outline_seed(
        unit_type=WorkflowUnitType.CHAPTER,
        title="Les nombres relatifs",
        document_path=str(pdf_path),
    )
    flat = _flatten_checklist(items)
    titles = [str(row.get("title", "")) for row in flat]

    assert titles[0].startswith("Chapitre 5")
    assert "addition et soustraction" in titles[0].lower()
    assert any(title.startswith("5.1") and "somme de deux nombres relatifs" in title.lower() for title in titles)
    assert any(title.startswith("5.1.1") and "même signe" in title.lower() for title in titles)
    assert any(title == "Exemples" for title in titles)
    assert not any("nombre positif" in title.lower() for title in titles)


def test_generate_unit_checklist_package_promotes_openai_shadow_when_notebooklm_is_weaker(monkeypatch):
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    poor_notebook_items = [
        {"title": "Exemples - Mathématiques", "kind": "example", "children": []},
        {"title": "5) heures", "kind": "section", "children": []},
    ]
    stronger_openai_items = [
        {
            "title": "Chapitre 5: Les nombres relatifs",
            "kind": "chapter",
            "children": [
                {"title": "5.1 Somme de deux nombres relatifs", "kind": "section", "children": []},
                {"title": "Exemples", "kind": "example", "children": []},
            ],
        }
    ]

    monkeypatch.setattr(workflow_generation.app_config, "OPENAI_API_KEY", "test-key")
    monkeypatch.setattr(
        workflow_generation,
        "_notebooklm_generate_checklist",
        lambda **kwargs: (poor_notebook_items, {"provider": "notebooklm"}, {"answer": "poor"}, None),
    )
    monkeypatch.setattr(
        workflow_generation,
        "_openai_generate_checklist",
        lambda **kwargs: (stronger_openai_items, {"answer": "strong"}, None),
    )
    monkeypatch.setattr(
        workflow_generation,
        "_build_pdf_layout_outline_seed",
        lambda **kwargs: [
            {
                "title": "Chapitre 5: Les nombres relatifs",
                "kind": "chapter",
                "children": [{"title": "5.1 Somme de deux nombres relatifs", "kind": "section", "children": []}],
            }
        ],
    )

    package = workflow_generation.generate_unit_checklist_package(
        unit_type=WorkflowUnitType.CHAPTER,
        title="Les nombres relatifs",
        source_text="Chapitre 5\n5.1 Somme de deux nombres relatifs\nExemples",
        provider="notebooklm",
        document_path="dummy.pdf",
    )

    flat = _flatten_checklist(package["items"])
    titles = [str(row.get("title", "")) for row in flat]
    assert package["source"] == "openai"
    assert any("chapitre 5" in title.lower() for title in titles)
    assert any("5.1" in title.lower() and "somme de deux nombres relatifs" in title.lower() for title in titles)
    assert not any("5) heures" in title.lower() for title in titles)


def test_generate_unit_checklist_package_trusts_notebooklm_outline_without_openai_shadow(monkeypatch):
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    outline_items = workflow_generation._parse_notebooklm_outline_response(
        "\n".join(
            [
                "- Chapitre 5 : Les nombres relatifs",
                "  - 5.1 Somme de deux nombres relatifs",
                "    - 5.1.1 Les deux nombres sont de même signe",
                "      - Propriete",
                "      - Exemples",
                "  - 5.2 Différence de deux nombres relatifs",
            ]
        ),
        unit_type=WorkflowUnitType.CHAPTER,
        unit_title="Les nombres relatifs",
    )
    assert outline_items

    monkeypatch.setattr(workflow_generation.app_config, "OPENAI_API_KEY", "test-key")
    monkeypatch.setattr(
        workflow_generation,
        "_notebooklm_generate_checklist",
        lambda **kwargs: (outline_items, {"provider": "notebooklm"}, {"answer": "outline", "response_mode": "outline"}, None),
    )
    monkeypatch.setattr(
        workflow_generation,
        "_openai_generate_checklist",
        lambda **kwargs: (_ for _ in ()).throw(RuntimeError("shadow_should_not_run")),
    )

    package = workflow_generation.generate_unit_checklist_package(
        unit_type=WorkflowUnitType.CHAPTER,
        title="Les nombres relatifs",
        source_text="Texte source minimal",
        provider="notebooklm",
    )

    flat = _flatten_checklist(package["items"])
    titles = [str(row.get("title", "")) for row in flat]
    assert package["source"] == "notebooklm"
    assert titles[0].startswith("Chapitre 5")
    assert any(title.startswith("5.1") for title in titles)
    assert any(title == "Exemples" for title in titles)
    assert package["unit_map"]["selected_outline_source"] in {"outline_response", "unit_map"}


def test_notebooklm_outline_normalizer_drops_teacher_meta_sections():
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    items = workflow_generation._normalize_notebooklm_outline_items(
        [
            {
                "title": "Les nombres rationnels : Somme et difference",
                "kind": "chapter",
                "children": [
                    {"title": "Objectifs d'apprentissage", "kind": "section", "children": []},
                    {"title": "Prerequis", "kind": "section", "children": []},
                    {"title": "Outils didactiques", "kind": "section", "children": []},
                    {"title": "Gestion du temps", "kind": "section", "children": []},
                    {"title": "Activites", "kind": "section", "children": []},
                    {"title": "Contenu de la lecon", "kind": "section", "children": []},
                    {"title": "Evaluation", "kind": "section", "children": []},
                ],
            }
        ],
        unit_type=WorkflowUnitType.CHAPTER,
        unit_title="Les nombres rationnels : Somme et difference",
    )

    titles = [str(row.get("title", "")) for row in _flatten_checklist(items)]
    assert "Activites" in titles
    assert "Contenu de la lecon" in titles
    assert "Evaluation" in titles
    assert "Objectifs d'apprentissage" not in titles
    assert "Prerequis" not in titles
    assert "Outils didactiques" not in titles
    assert "Gestion du temps" not in titles


def test_notebooklm_outline_normalizer_reorders_student_flow_and_drops_broader_teacher_meta():
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    items = workflow_generation._normalize_notebooklm_outline_items(
        [
            {
                "title": "Les nombres rationnels : Somme et difference",
                "kind": "chapter",
                "children": [
                    {"title": "Evaluation", "kind": "section", "children": []},
                    {"title": "Competences visees", "kind": "section", "children": []},
                    {
                        "title": "Contenu de la lecon",
                        "kind": "section",
                        "children": [
                            {
                                "title": "1) Les denominateurs sont les memes",
                                "kind": "section",
                                "children": [
                                    {"title": "Exercices", "kind": "exercise", "children": []},
                                    {"title": "Exemples", "kind": "example", "children": []},
                                    {"title": "Regle", "kind": "property", "children": []},
                                ],
                            }
                        ],
                    },
                    {"title": "Activites", "kind": "other", "children": []},
                    {"title": "Demarche pedagogique", "kind": "section", "children": []},
                ],
            }
        ],
        unit_type=WorkflowUnitType.CHAPTER,
        unit_title="Les nombres rationnels : Somme et difference",
    )

    root_children = items[0]["children"]
    child_titles = [str(row.get("title", "")) for row in root_children]
    assert child_titles == ["Activites", "Contenu de la lecon", "Evaluation"]
    nested_titles = [str(row.get("title", "")) for row in root_children[1]["children"][0]["children"]]
    assert nested_titles == ["Regle", "Exemples", "Exercices"]


def test_generate_unit_checklist_package_aligns_checklist_with_stronger_unit_map_outline(monkeypatch):
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    outline_items = [
        {
            "title": "Les nombres rationnels : Somme et difference",
            "kind": "chapter",
            "children": [
                {"title": "Activites", "kind": "section", "children": []},
                {"title": "Contenu de la lecon", "kind": "section", "children": []},
            ],
        }
    ]
    stronger_unit_map = {
        "unit_title": "Les nombres rationnels : Somme et difference",
        "ordered_outline": [
            {
                "title": "Les nombres rationnels : Somme et difference",
                "kind": "chapter",
                "children": [
                    {"title": "Objectifs d'apprentissage", "kind": "section", "children": []},
                    {"title": "Activites", "kind": "section", "children": []},
                    {
                        "title": "Contenu de la lecon",
                        "kind": "section",
                        "children": [
                            {"title": "I- Addition", "kind": "subsection", "children": []},
                        ],
                    },
                    {"title": "Evaluation", "kind": "section", "children": []},
                ],
            }
        ],
        "teaching_goals": ["Comprendre la progression de l'unite"],
    }

    monkeypatch.setattr(workflow_generation.app_config, "OPENAI_API_KEY", "test-key")
    monkeypatch.setattr(
        workflow_generation,
        "_notebooklm_generate_checklist",
        lambda **kwargs: (
            outline_items,
            {"provider": "notebooklm"},
            {"response_mode": "outline", "unit_map": stronger_unit_map, "responses": []},
            None,
        ),
    )
    monkeypatch.setattr(
        workflow_generation,
        "_openai_generate_checklist",
        lambda **kwargs: (_ for _ in ()).throw(RuntimeError("shadow_should_not_run")),
    )

    package = workflow_generation.generate_unit_checklist_package(
        unit_type=WorkflowUnitType.CHAPTER,
        title="Les nombres rationnels : Somme et difference",
        source_text="Objectifs d'apprentissage\nActivites\nContenu de la lecon\nI- Addition\nEvaluation",
        provider="notebooklm",
    )

    flat = _flatten_checklist(package["items"])
    titles = [str(row.get("title", "")) for row in flat]
    assert package["source"] == "notebooklm"
    assert "Evaluation" in titles
    assert "Objectifs d'apprentissage" not in titles
    assert "Comprendre la progression de l'unite" in package["unit_map"]["teaching_goals"]
    assert package["unit_map"]["selected_outline_source"] == "unit_map"
    assert package["unit_map"]["ordered_outline"][0]["children"][0]["title"] == "Activites"


def test_unit_map_section_plans_capture_delivery_sequence(monkeypatch):
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    outline_items = [
        {
            "title": "Les nombres rationnels : Somme et difference",
            "kind": "chapter",
            "children": [
                {
                    "title": "Contenu de la lecon",
                    "kind": "section",
                    "children": [
                        {
                            "title": "1) Les denominateurs sont les memes",
                            "kind": "section",
                            "children": [
                                {"title": "Exercices", "kind": "exercise", "children": []},
                                {"title": "Exemples", "kind": "example", "children": []},
                                {"title": "Regle", "kind": "property", "children": []},
                            ],
                        }
                    ],
                }
            ],
        }
    ]

    monkeypatch.setattr(workflow_generation.app_config, "OPENAI_API_KEY", "test-key")
    monkeypatch.setattr(
        workflow_generation,
        "_notebooklm_generate_checklist",
        lambda **kwargs: (
            outline_items,
            {"provider": "notebooklm"},
            {"response_mode": "outline", "responses": []},
            None,
        ),
    )
    monkeypatch.setattr(
        workflow_generation,
        "_openai_generate_checklist",
        lambda **kwargs: (_ for _ in ()).throw(RuntimeError("shadow_should_not_run")),
    )

    package = workflow_generation.generate_unit_checklist_package(
        unit_type=WorkflowUnitType.CHAPTER,
        title="Les nombres rationnels : Somme et difference",
        source_text="Contenu de la lecon\n1) Les denominateurs sont les memes\nRegle\nExemples\nExercices",
        provider="notebooklm",
    )

    section_plans = package["unit_map"]["section_plans"]
    matching = next(plan for plan in section_plans if plan["section_title"] == "1) Les denominateurs sont les memes")
    assert matching["delivery_sequence"] == ["Regle", "Exemples", "Exercices"]
    assert matching["content_titles"] == ["Regle"]
    assert matching["example_titles"] == ["Exemples"]
    assert matching["exercise_titles"] == ["Exercices"]


def test_generate_unit_checklist_package_keeps_notebooklm_content_blocks(monkeypatch):
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    outline_items = [
        {
            "title": "Les nombres rationnels : Somme et difference",
            "kind": "chapter",
            "children": [
                {"title": "Objectifs d'apprentissage", "kind": "section", "children": []},
                {"title": "Contenu de la lecon", "kind": "section", "children": []},
                {"title": "Evaluation", "kind": "section", "children": []},
            ],
        }
    ]
    content_pack = {
        "content_blocks": [
            {
                "section_title": "1) Les denominateurs sont les memes",
                "section_path": ["I- Addition et soustraction", "1) Les denominateurs sont les memes"],
                "kind": "activity",
                "teaching_phase": "activity",
                "title": "Activite 1 : Calculer",
                "source_excerpt": "Calcule puis compare les resultats.",
                "teaching_material": "Faire calculer les eleves en binomes avant la mise en commun.",
                "student_visible": True,
                "teacher_only": False,
                "order_index": 1,
            },
            {
                "section_title": "1) Les denominateurs sont les memes",
                "section_path": ["I- Addition et soustraction", "1) Les denominateurs sont les memes"],
                "kind": "property",
                "teaching_phase": "content",
                "title": "Regle",
                "source_excerpt": "Les denominateurs sont les memes.",
                "teaching_material": "Institutionnaliser la regle avant les applications longues.",
                "student_visible": True,
                "teacher_only": False,
                "order_index": 2,
            },
        ]
    }

    monkeypatch.setattr(workflow_generation.app_config, "OPENAI_API_KEY", "test-key")
    monkeypatch.setattr(
        workflow_generation,
        "_notebooklm_generate_checklist",
        lambda **kwargs: (
            outline_items,
            {"provider": "notebooklm"},
            {"response_mode": "outline", "content_pack": content_pack, "responses": []},
            None,
        ),
    )
    monkeypatch.setattr(
        workflow_generation,
        "_openai_generate_checklist",
        lambda **kwargs: (_ for _ in ()).throw(RuntimeError("shadow_should_not_run")),
    )

    package = workflow_generation.generate_unit_checklist_package(
        unit_type=WorkflowUnitType.CHAPTER,
        title="Les nombres rationnels : Somme et difference",
        source_text="Activite 1\nRegle\nEvaluation",
        provider="notebooklm",
    )

    assert package["source"] == "notebooklm"
    assert package["unit_map"]["future_actions"][1] == "content_pack"
    assert package["content_blocks"][0]["title"] == "Activite 1 : Calculer"
    assert package["content_blocks"][0]["kind"] == "activity"
    assert package["content_blocks"][0]["section_path"] == ["I- Addition et soustraction", "1) Les denominateurs sont les memes"]
    assert package["content_blocks"][1]["teaching_material"].startswith("Institutionnaliser")
    section_plans = package["unit_map"]["section_plans"]
    activity_plan = next(plan for plan in section_plans if plan["section_title"] == "1) Les denominateurs sont les memes")
    assert activity_plan["delivery_sequence"][0] == "Activite 1 : Calculer"
    assert activity_plan["blocks"][0]["teaching_material"].startswith("Faire calculer")
    assert activity_plan["content_titles"] == ["Regle"]
    teacher_playbook = package["unit_map"]["teacher_playbook"]
    playbook_entry = next(entry for entry in teacher_playbook if entry["section_title"] == "1) Les denominateurs sont les memes")
    assert "generate_harder_practice" in playbook_entry["available_actions"]
    assert any("plus difficiles" in row for row in playbook_entry["suggested_requests"])
    material_studio = package["unit_map"]["material_studio"]
    assert any(row["id"] == "study_guide" for row in material_studio["unit_artifacts"])
    assert any(row["artifact_type"] == "slide_deck" for row in material_studio["unit_artifacts"])
    assert any(row["section_title"] == "1) Les denominateurs sont les memes" for row in material_studio["teacher_artifacts"])
    rendered_titles = {node["title"] for node in workflow_generation._flatten_checklist_nodes(package["items"])}
    assert "Objectifs d'apprentissage" not in rendered_titles
    assert "Activite 1 : Calculer" in rendered_titles


def test_build_checklist_from_section_plans_prefers_precise_section_paths():
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    section_plans = [
        {
            "section_title": "1) Les denominateurs sont les memes",
            "section_path": ["I- Addition et soustraction", "1) Les denominateurs sont les memes"],
            "delivery_sequence": ["Activite 1 : Calculer", "Regle", "Exemples", "Exercices"],
            "blocks": [
                {"title": "Activite 1 : Calculer", "kind": "activity", "order_index": 1, "student_visible": True, "teacher_only": False},
                {"title": "Regle", "kind": "property", "order_index": 2, "student_visible": True, "teacher_only": False},
                {"title": "Exemples", "kind": "example", "order_index": 3, "student_visible": True, "teacher_only": False},
                {"title": "Exercices", "kind": "exercise", "order_index": 4, "student_visible": True, "teacher_only": False},
            ],
        }
    ]

    items = workflow_generation._build_checklist_from_section_plans(
        section_plans,
        unit_title="Les nombres rationnels : Somme et difference",
        unit_type=WorkflowUnitType.CHAPTER,
        fallback_outline=[
            {
                "title": "Les nombres rationnels : Somme et difference",
                "kind": "chapter",
                "children": [{"title": "Activites", "kind": "section", "children": []}],
            }
        ],
    )

    flat_titles = [str(row.get("title", "")) for row in workflow_generation._flatten_checklist_nodes(items)]
    assert flat_titles[0] == "Les nombres rationnels : Somme et difference"
    assert "I- Addition et soustraction" in flat_titles
    assert "1) Les denominateurs sont les memes" in flat_titles
    assert "Activites" not in flat_titles
    assert flat_titles.index("Activite 1 : Calculer") < flat_titles.index("Regle") < flat_titles.index("Exemples") < flat_titles.index("Exercices")


def test_normalize_content_blocks_payload_retargets_generic_paths_to_precise_sections():
    from app.services import workflow_generation

    normalized = workflow_generation._normalize_content_blocks_payload(
        {
            "content_blocks": [
                {
                    "section_title": "Activites",
                    "section_path": ["Activites"],
                    "kind": "activity",
                    "teaching_phase": "activity",
                    "title": "Activite 1 : Calculer",
                    "source_excerpt": "Calcule",
                    "teaching_material": "Calculer en binomes",
                    "student_visible": True,
                    "teacher_only": False,
                    "order_index": 1,
                },
                {
                    "section_title": "1) Les denominateurs sont les memes",
                    "section_path": ["I- Addition et soustraction", "1) Les denominateurs sont les memes"],
                    "kind": "property",
                    "teaching_phase": "content",
                    "title": "Regle",
                    "source_excerpt": "Les denominateurs sont les memes",
                    "teaching_material": "Institutionnaliser la regle",
                    "student_visible": True,
                    "teacher_only": False,
                    "order_index": 2,
                },
            ]
        },
        unit_map=None,
        fallback_outline=[],
    )

    assert normalized[0]["section_path"] == ["I- Addition et soustraction", "1) Les denominateurs sont les memes"]
    assert normalized[0]["section_title"] == "1) Les denominateurs sont les memes"


def test_normalize_content_blocks_payload_accepts_refined_sections_schema():
    from app.services import workflow_generation

    normalized = workflow_generation._normalize_content_blocks_payload(
        {
            "unit_title": "Les nombres rationnels : Somme et difference",
            "sections": [
                {
                    "section_title": "1) Les denominateurs sont les memes",
                    "section_path": ["I- Addition et soustraction", "1) Les denominateurs sont les memes"],
                    "order_index": 1,
                    "blocks": [
                        {
                            "kind": "property",
                            "title": "Regle",
                            "exact_text": "On garde le meme denominateur.\nOn additionne les numerateurs.",
                            "order_index": 1,
                        },
                        {
                            "kind": "example",
                            "title": "Exemple 1",
                            "exact_text": "3/7 + 2/7 = 5/7",
                            "order_index": 2,
                        },
                    ],
                }
            ],
        },
        unit_map=None,
        fallback_outline=[],
    )

    assert [row["title"] for row in normalized] == ["Regle", "Exemple 1"]
    assert normalized[0]["section_path"] == ["I- Addition et soustraction", "1) Les denominateurs sont les memes"]
    assert normalized[0]["source_excerpt_raw"].startswith("On garde le meme denominateur.")
    assert normalized[0]["teaching_material_raw"].startswith("On garde le meme denominateur.")
    assert normalized[1]["kind"] == "example"


def test_generate_unit_checklist_package_accepts_refined_notebooklm_sections(monkeypatch):
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    outline_items = [
        {
            "title": "Les nombres rationnels : Somme et difference",
            "kind": "chapter",
            "children": [
                {"title": "I- Addition et soustraction", "kind": "section", "children": []},
            ],
        }
    ]
    refined_content_pack = {
        "unit_title": "Les nombres rationnels : Somme et difference",
        "sections": [
            {
                "section_title": "1) Les denominateurs sont les memes",
                "section_path": ["I- Addition et soustraction", "1) Les denominateurs sont les memes"],
                "order_index": 1,
                "blocks": [
                    {
                        "kind": "activity",
                        "title": "Activite 1 : Calculer",
                        "exact_text": "Calcule 3/7 + 2/7 puis compare avec 5/7.",
                        "order_index": 1,
                    },
                    {
                        "kind": "property",
                        "title": "Regle",
                        "exact_text": "On garde le meme denominateur.\nOn additionne les numerateurs.",
                        "order_index": 2,
                    },
                ],
            }
        ],
    }

    monkeypatch.setattr(workflow_generation.app_config, "OPENAI_API_KEY", "test-key")
    monkeypatch.setattr(
        workflow_generation,
        "_notebooklm_generate_checklist",
        lambda **kwargs: (
            outline_items,
            {"provider": "notebooklm"},
            {"response_mode": "outline", "content_pack": refined_content_pack, "responses": []},
            None,
        ),
    )
    monkeypatch.setattr(
        workflow_generation,
        "_openai_generate_checklist",
        lambda **kwargs: (_ for _ in ()).throw(RuntimeError("shadow_should_not_run")),
    )

    package = workflow_generation.generate_unit_checklist_package(
        unit_type=WorkflowUnitType.CHAPTER,
        title="Les nombres rationnels : Somme et difference",
        source_text="Activite 1\nRegle",
        provider="notebooklm",
    )

    assert package["source"] == "notebooklm"
    assert package["content_blocks"][0]["title"] == "Activite 1 : Calculer"
    assert package["content_blocks"][0]["teaching_material_raw"] == "Calcule 3/7 + 2/7 puis compare avec 5/7."
    assert package["content_blocks"][1]["title"] == "Regle"
    assert package["content_blocks"][1]["source_excerpt_raw"].startswith("On garde le meme denominateur.")
    section_plans = package["unit_map"]["section_plans"]
    activity_plan = next(plan for plan in section_plans if plan["section_title"] == "1) Les denominateurs sont les memes")
    assert activity_plan["delivery_sequence"] == ["Activite 1 : Calculer", "Regle"]
    assert activity_plan["blocks"][1]["teaching_material"].startswith("On garde le meme denominateur.")


def test_notebooklm_generate_checklist_async_uses_outline_only_flow(monkeypatch):
    import asyncio

    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    recorded_prompts: list[str] = []

    class _FakeNotebook:
        id = "nb-test-123"

    class _FakeNotebooks:
        async def create(self, _title):
            return _FakeNotebook()

    class _FakeOpened:
        def __init__(self):
            self.notebooks = _FakeNotebooks()

        async def __aenter__(self):
            return self

        async def __aexit__(self, exc_type, exc, tb):
            return False

    class _FakeAnswer:
        def __init__(self, answer, conversation_id):
            self.answer = answer
            self.conversation_id = conversation_id

    async def _fake_create_client():
        return _FakeOpened()

    monkeypatch.setattr(workflow_generation, "_create_notebooklm_client", _fake_create_client)
    monkeypatch.setattr(
        workflow_generation,
        "_notebooklm_attach_source",
        lambda **kwargs: asyncio.sleep(0, result=["source-1"]),
    )

    async def _fake_ask(**kwargs):
        prompt = str(kwargs.get("prompt") or "")
        recorded_prompts.append(prompt)
        return _FakeAnswer(
            "- Les nombres rationnels : Somme et difference\n  - I- Addition et soustraction\n    - 1) Les denominateurs sont les memes",
            f"conv-{len(recorded_prompts)}",
        )

    monkeypatch.setattr(workflow_generation, "_ask_notebooklm_with_source_retry", _fake_ask)

    items, provider_context, raw_result, error_message = asyncio.run(
        workflow_generation._notebooklm_generate_checklist_async(
            unit_type=WorkflowUnitType.CHAPTER,
            title="Les nombres rationnels : Somme et difference",
            source_text="source",
            session_count=6,
            document_path="",
            outline_hint_lines=None,
        )
    )

    assert error_message is None
    assert items
    assert provider_context["notebook_id"] == "nb-test-123"
    assert provider_context["notebook_role"] == "chapter_outline"
    assert raw_result["response_mode"] == "outline_only"
    assert "content_pack" not in raw_result
    assert len(recorded_prompts) == 2


def test_workflow_unit_start_returns_clear_error_when_notebooklm_refresh_is_required(client, monkeypatch):
    from app.routers import workflow as workflow_router
    from app.services.workflow_generation import NotebookLMGenerationUnavailableError

    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "NotebookLM Guard Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    monkeypatch.setattr(
        workflow_router,
        "generate_unit_checklist",
        lambda **kwargs: (_ for _ in ()).throw(
            NotebookLMGenerationUnavailableError("NotebookLM login refresh is required before unit extraction.")
        ),
    )

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Fractions",
            "source_text": "Fractions\nActivites\nContenu de la lecon\n",
        },
    )
    assert unit_resp.status_code == 409
    assert "NotebookLM login refresh is required" in unit_resp.json()["detail"]


def test_select_best_notebooklm_outline_candidate_prefers_more_complete_tree():
    from app.models import WorkflowUnitType
    from app.services import workflow_generation

    weaker = [
        {
            "title": "Les nombres rationnels : Somme et différence",
            "kind": "chapter",
            "children": [
                {"title": "Activités", "kind": "other", "children": [{"title": "Activité 1 : Calculer :", "kind": "other", "children": []}]},
                {"title": "Evaluation", "kind": "section", "children": [{"title": "Exercice 1 : Calcule puis simplifie :", "kind": "exercise", "children": []}]},
            ],
        }
    ]
    stronger = [
        {
            "title": "Les nombres rationnels : Somme et différence",
            "kind": "chapter",
            "children": [
                {
                    "title": "Activités",
                    "kind": "other",
                    "children": [
                        {"title": "Activité 1 : Calculer :", "kind": "other", "children": []},
                        {"title": "Activité 2 : Calculer :", "kind": "other", "children": []},
                        {"title": "Activité 3 : Calculer :", "kind": "other", "children": []},
                    ],
                },
                {
                    "title": "Evaluation",
                    "kind": "section",
                    "children": [
                        {"title": "Exercice 1 : Calcule puis simplifie :", "kind": "exercise", "children": []},
                        {"title": "Exercice 2 : Calcule puis simplifie :", "kind": "exercise", "children": []},
                        {"title": "Exercice 3 : Calcule puis simplifie :", "kind": "exercise", "children": []},
                    ],
                },
            ],
        }
    ]
    source_text = "\n".join(
        [
            "Les nombres rationnels :",
            "Somme et différence",
            "Activités",
            "Activité 1 : Calculer :",
            "Activité 2 : Calculer :",
            "Activité 3 : Calculer :",
            "Evaluation",
            "Exercice 1 : Calcule puis simplifie :",
            "Exercice 2 : Calcule puis simplifie :",
            "Exercice 3 : Calcule puis simplifie :",
        ]
    )

    variant, items = workflow_generation._select_best_notebooklm_outline_candidate(
        [("primary", weaker), ("completeness_review", stronger)],
        source_text=source_text,
        unit_type=WorkflowUnitType.CHAPTER,
        unit_title="Les nombres rationnels : Somme et différence",
    )

    assert variant == "completeness_review"
    assert items == stronger


def test_generate_unit_checklist_prefers_notebooklm_for_pdf_when_ready(monkeypatch):
    from app.models import WorkflowUnitType
    from app.services import workflow as workflow_service

    captured: dict[str, object] = {}

    monkeypatch.setattr(workflow_service.app_config, "UNIT_PLANNER_PROVIDER", "openai")
    monkeypatch.setattr(workflow_service, "OPENAI_API_KEY", "")
    monkeypatch.setattr(workflow_service, "notebooklm_provider_ready", lambda: True)

    def fake_package(**kwargs):
        captured.update(kwargs)
        return {
            "source": "notebooklm",
            "requested_provider": kwargs.get("provider"),
            "model": "notebooklm-py",
            "status": "ready",
            "items": [{"title": "Chapitre test", "kind": "chapter", "children": []}],
            "raw_provider_response": {"answer": "- Chapitre test"},
            "error_message": None,
            "provider_context": {"provider": "notebooklm"},
        }

    monkeypatch.setattr(workflow_service, "generate_unit_checklist_package", fake_package)

    payload = workflow_service.generate_unit_checklist(
        unit_type=WorkflowUnitType.CHAPTER,
        title="Chapitre test",
        source_text="Texte source",
        session_count=4,
        document_path="dummy.pdf",
    )

    assert captured["provider"] == "notebooklm"
    assert payload["source"] == "notebooklm"
    assert payload["model"] == "notebooklm-py"


def test_create_notebooklm_client_uses_resolved_storage_state_path(monkeypatch, tmp_path):
    import asyncio
    import sys
    import types

    from app.services import workflow_generation

    storage_path = tmp_path / "profiles" / "default" / "storage_state.json"
    storage_path.parent.mkdir(parents=True, exist_ok=True)
    storage_path.write_text('{"cookies": [], "origins": []}', encoding="utf-8")

    captured: dict[str, object] = {}

    class FakeNotebookLMClient:
        @classmethod
        async def from_storage(cls, **kwargs):
            captured.update(kwargs)
            return object()

    monkeypatch.setattr(workflow_generation.app_config, "NOTEBOOKLM_HOME", str(tmp_path))
    monkeypatch.setattr(workflow_generation.app_config, "NOTEBOOKLM_AUTH_PATH", "")
    monkeypatch.setattr(workflow_generation.app_config, "NOTEBOOKLM_PROFILE", "default")
    monkeypatch.setitem(sys.modules, "notebooklm", types.SimpleNamespace(NotebookLMClient=FakeNotebookLMClient))

    client = asyncio.run(workflow_generation._create_notebooklm_client())

    assert client is not None
    assert str(captured.get("path")) == str(storage_path)


def test_unit_creation_replaces_slug_title_with_first_meaningful_generated_heading(client, monkeypatch):
    import app.routers.workflow as workflow_router

    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Slug Title Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    monkeypatch.setattr(
        workflow_router,
        "generate_unit_checklist",
        lambda **kwargs: {
            "source": "notebooklm",
            "requested_provider": "notebooklm",
            "model": "notebooklm-py",
            "status": "ready",
            "items": [
                {
                    "title": "Chapitre 1: Introduction des nombres rationnels",
                    "kind": "chapter",
                    "children": [
                        {"title": "Definition - Nombre rationnel", "kind": "definition", "children": []},
                    ],
                }
            ],
            "raw_provider_response": {"answer": "{}"},
            "error_message": None,
            "provider_context": None,
        },
    )

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "introduction-aux-nombres-rationnels-cours-ma",
            "source_text": "dummy source",
        },
    )

    assert unit_resp.status_code == 201
    payload = unit_resp.json()
    assert payload["title"] == "Chapitre 1: Introduction des nombres rationnels"


def test_workflow_notebooklm_provider_persists_context_and_writeup(client, monkeypatch):
    from app import config as app_config
    from app.services import workflow as workflow_service
    from app.services import workflow_generation

    monkeypatch.setattr(app_config, "UNIT_PLANNER_PROVIDER", "notebooklm")
    monkeypatch.setattr(app_config, "SESSION_WRITER_PROVIDER", "notebooklm")
    monkeypatch.setattr(workflow_service, "ensure_notebooklm_generation_ready", lambda **kwargs: None)
    monkeypatch.setattr(workflow_generation, "ensure_notebooklm_generation_ready", lambda **kwargs: None)

    monkeypatch.setattr(
        workflow_generation,
        "_notebooklm_generate_checklist",
        lambda **kwargs: (
            [
                {
                    "title": "1. Factorisation",
                    "kind": "chapter",
                    "children": [
                        {"title": "1.1 Mise en facteur commun", "kind": "section", "children": [], "session_number": 1},
                        {"title": "1.2 Exercices d'application", "kind": "exercise", "children": [], "session_number": 2},
                    ],
                }
            ],
            {
                "provider": "notebooklm",
                "notebook_id": "nb-unit-1",
                "source_ids": ["src-1"],
                "notebook_title": "Teacher Progress - Factorisation",
            },
            {
                "answer": "{\"items\": []}",
                "unit_map": {
                    "unit_title": "Factorisation",
                    "ordered_outline": [
                        {
                            "title": "Factorisation",
                            "kind": "chapter",
                            "children": [
                                {
                                    "title": "1.1 Mise en facteur commun",
                                    "kind": "section",
                                    "children": [
                                        {"title": "Exemple guide", "kind": "example", "children": []},
                                        {"title": "Exercices d'application", "kind": "exercise", "children": []},
                                    ],
                                }
                            ],
                        }
                    ],
                    "section_plans": [
                        {
                            "section_title": "1.1 Mise en facteur commun",
                            "section_path": ["Factorisation", "1.1 Mise en facteur commun"],
                            "delivery_sequence": ["Exemple guide", "Exercices d'application"],
                            "activity_titles": [],
                            "content_titles": ["1.1 Mise en facteur commun"],
                            "example_titles": ["Exemple guide"],
                            "exercise_titles": ["Exercices d'application"],
                            "blocks": [
                                {
                                    "title": "Exemple guide",
                                    "kind": "example",
                                    "teaching_material": "Montrer comment isoler le facteur commun.",
                                    "source_excerpt": "Exemple guide",
                                    "student_visible": True,
                                    "teacher_only": False,
                                    "order_index": 1,
                                },
                                {
                                    "title": "Exercices d'application",
                                    "kind": "exercise",
                                    "teaching_material": "Faire pratiquer la factorisation sur des expressions simples.",
                                    "source_excerpt": "Exercices d'application",
                                    "student_visible": True,
                                    "teacher_only": False,
                                    "order_index": 2,
                                },
                            ],
                        }
                    ],
                },
                "content_blocks": [
                    {
                        "section_title": "1.1 Mise en facteur commun",
                        "section_path": ["Factorisation", "1.1 Mise en facteur commun"],
                        "kind": "lesson",
                        "teaching_phase": "content",
                        "title": "1.1 Mise en facteur commun",
                        "source_excerpt": "Mise en facteur commun",
                        "teaching_material": "Introduire la mise en facteur commun avec une progression claire.",
                        "student_visible": True,
                        "teacher_only": False,
                        "order_index": 1,
                    },
                    {
                        "section_title": "1.1 Mise en facteur commun",
                        "section_path": ["Factorisation", "1.1 Mise en facteur commun"],
                        "kind": "example",
                        "teaching_phase": "example",
                        "title": "Exemple guide",
                        "source_excerpt": "Exemple guide",
                        "teaching_material": "Montrer comment isoler le facteur commun.",
                        "student_visible": True,
                        "teacher_only": False,
                        "order_index": 2,
                    },
                    {
                        "section_title": "1.1 Mise en facteur commun",
                        "section_path": ["Factorisation", "1.1 Mise en facteur commun"],
                        "kind": "exercise",
                        "teaching_phase": "practice",
                        "title": "Exercices d'application",
                        "source_excerpt": "Exercices d'application",
                        "teaching_material": "Faire pratiquer la factorisation sur des expressions simples.",
                        "student_visible": True,
                        "teacher_only": False,
                        "order_index": 3,
                    },
                ],
            },
            None,
        ),
    )
    captured_writeup_kwargs: dict[str, object] = {}
    monkeypatch.setattr(
        workflow_generation,
        "_notebooklm_generate_session_writeup",
        lambda **kwargs: (
            captured_writeup_kwargs.update(kwargs)
            or {
                "provider": "notebooklm",
                "requested_provider": "notebooklm",
                "model": "notebooklm-py",
                "status": "ready",
                "title": "Seance 1 - Mise en facteur commun",
                "checked_item_ids": kwargs["checked_item_ids"],
                "checked_item_titles": kwargs["checked_item_titles"],
                "learning_focus": ["Reconna?tre un facteur commun."],
                "teaching_content": ["La seance a introduit la mise en facteur commun avec des exemples progressifs."],
                "practice_items": ["Exercices d'application sur des expressions algebriques."],
                "teacher_note_snapshot": kwargs["note_text"] or None,
                "raw_provider_response": {"answer": "{\"title\":\"ok\"}"},
                "error_message": None,
            }
        ),
    )

    headers = _auth_headers(client)
    _close_any_active_unit(client, headers)
    class_resp = client.post("/classes", json={"name": "NotebookLM Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    roster_content = _build_roster_file([("A1", "Student One")])
    roster_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("roster.xlsx", roster_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert roster_resp.status_code == 200

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Factorisation",
            "source_text": "1. Factorisation\n1.1 Mise en facteur commun\n1.2 Exercices d'application\n",
        },
    )
    assert unit_resp.status_code == 201
    unit_id = unit_resp.json()["id"]

    blueprint_resp = client.get(f"/workflow/classes/{class_id}/units/{unit_id}/blueprint", headers=headers)
    assert blueprint_resp.status_code == 200
    blueprint = blueprint_resp.json()
    assert blueprint["provider"] == "notebooklm"
    assert blueprint["blueprint_json"]["provider_context"]["notebook_id"] == "nb-unit-1"

    start_session_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/start",
        headers=headers,
        json={"absent_student_ids": []},
    )
    assert start_session_resp.status_code == 201
    session_id = start_session_resp.json()["id"]

    writeup_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/{session_id}/writeup/generate",
        headers=headers,
        json={"regenerate": True},
    )
    assert writeup_resp.status_code == 200
    writeup = writeup_resp.json()
    assert writeup["provider"] == "notebooklm"
    assert writeup["title"] == "Seance 1 - Mise en facteur commun"
    assert writeup["learning_focus"] == ["Reconna?tre un facteur commun."]
    assert isinstance(captured_writeup_kwargs.get("unit_map"), dict)
    assert isinstance(captured_writeup_kwargs.get("content_blocks"), list)
    assert captured_writeup_kwargs["unit_map"]["section_plans"][0]["section_title"] == "1.1 Mise en facteur commun"
    assert any(
        row.get("title") == "1.1 Mise en facteur commun"
        or row.get("section_title") == "1.1 Mise en facteur commun"
        or "1.1 Mise en facteur commun" in [str(value) for value in (row.get("section_path") or [])]
        for row in captured_writeup_kwargs["content_blocks"]
        if isinstance(row, dict)
    )


def test_generate_session_writeup_package_uses_unit_brain_in_fallback():
    from app.services.workflow_generation import generate_session_writeup_package

    package = generate_session_writeup_package(
        unit_title="Les nombres rationnels",
        unit_type=None,
        session_number=2,
        checked_item_ids=[1, 2],
        checked_item_titles=["1) Les denominateurs sont differents", "Exemples :"],
        note_text="",
        source_text="",
        provider="fallback",
        unit_map={
            "section_plans": [
                {
                    "section_title": "1) Les denominateurs sont differents",
                    "section_path": ["Les nombres rationnels", "1) Les denominateurs sont differents"],
                    "delivery_sequence": ["Propriete", "Exemples", "Exercices d'application"],
                    "activity_titles": [],
                    "content_titles": ["Propriete"],
                    "example_titles": ["Exemples"],
                    "exercise_titles": ["Exercices d'application"],
                }
            ]
        },
        content_blocks=[
            {
                "section_title": "1) Les denominateurs sont differents",
                "section_path": ["Les nombres rationnels", "1) Les denominateurs sont differents"],
                "kind": "property",
                "teaching_phase": "content",
                "title": "Propriete",
                "source_excerpt": "Propriete",
                "teaching_material": "Rappeler la propriete utilisee pour additionner deux fractions de denominateurs differents.",
                "student_visible": True,
                "teacher_only": False,
                "order_index": 1,
            },
            {
                "section_title": "1) Les denominateurs sont differents",
                "section_path": ["Les nombres rationnels", "1) Les denominateurs sont differents"],
                "kind": "example",
                "teaching_phase": "example",
                "title": "Exemples",
                "source_excerpt": "Exemples",
                "teaching_material": "Montrer un exemple guide avec PPCM puis simplification.",
                "student_visible": True,
                "teacher_only": False,
                "order_index": 2,
            },
            {
                "section_title": "1) Les denominateurs sont differents",
                "section_path": ["Les nombres rationnels", "1) Les denominateurs sont differents"],
                "kind": "exercise",
                "teaching_phase": "practice",
                "title": "Exercices d'application",
                "source_excerpt": "Exercices d'application",
                "teaching_material": "Faire pratiquer plusieurs additions de fractions de denominateurs differents.",
                "student_visible": True,
                "teacher_only": False,
                "order_index": 3,
            },
        ],
        saved_guidance=[
            {
                "artifact_kind": "guided_practice",
                "section_title": "1) Les denominateurs sont differents",
                "section_path": ["Les nombres rationnels", "1) Les denominateurs sont differents"],
                "title": "Practice ladder",
                "content_markdown": "Commencer par un exemple de PPCM puis donner trois calculs progressifs.",
            }
        ],
    )

    assert package["provider"] == "fallback"
    assert package["source_payload"]["unit_brain_used"] is True
    assert "1) Les denominateurs sont differents" in package["source_payload"]["matched_section_titles"]
    assert "Exercices d'application" in package["source_payload"]["matched_block_titles"]
    assert "Practice ladder" in package["source_payload"]["matched_guidance_titles"]
    assert any(row == "Exercices d'application." for row in package["practice_items"])
    assert any("PPCM" in row for row in package["teaching_content"])

def test_timetable_import_preview_csv(client):
    headers = _auth_headers(client)
    csv_content = _build_timetable_csv(
        [
            ("owner@school.edu", "2APIC-1", "Math", "Monday", "08:00", "09:00", "R12", "G1"),
            ("owner@school.edu", "2APIC-2", "Math", "Friday", "10:30", "09:30", "R15", "G2"),
        ]
    )

    preview_resp = client.post(
        "/workflow/timetable/import/preview",
        headers=headers,
        files={"file": ("emploi.csv", csv_content, "text/csv")},
    )
    assert preview_resp.status_code == 200
    payload = preview_resp.json()
    assert payload["total_rows"] == 2
    assert payload["valid_rows"] == 1
    assert payload["invalid_rows"] == 1
    assert len(payload["rows"]) == 2
    assert payload["rows"][0]["is_valid"] is True
    assert payload["rows"][0]["weekday"] == 1
    assert payload["rows"][0]["start_time"] == "08:00:00"
    assert payload["rows"][1]["is_valid"] is False
    assert any("end_time" in issue for issue in payload["rows"][1]["issues"])


def test_timetable_import_preview_ics(client):
    headers = _auth_headers(client)
    ics_content = _build_timetable_ics(
        [
            {
                "summary": "2APIC-1 | Math",
                "dtstart": "20260907T080000",
                "dtend": "20260907T090000",
                "rrule": "FREQ=WEEKLY;BYDAY=MO,WE",
                "location": "R12",
                "description": "Group:G1",
            },
            {
                "summary": "2APIC-2 | Physics",
                "dtstart": "20260908T110000",
                "dtend": "20260908T103000",
                "location": "R15",
            },
        ]
    )

    preview_resp = client.post(
        "/workflow/timetable/import/preview",
        headers=headers,
        files={"file": ("emploi.ics", ics_content, "text/calendar")},
    )
    assert preview_resp.status_code == 200
    payload = preview_resp.json()
    assert payload["total_rows"] == 3
    assert payload["valid_rows"] == 2
    assert payload["invalid_rows"] == 1
    assert payload["rows"][0]["class_name"] == "2APIC-1"
    assert payload["rows"][0]["subject"] == "Math"
    assert payload["rows"][0]["weekday"] == 1
    assert payload["rows"][0]["room"] == "R12"
    assert payload["rows"][0]["group"] == "G1"
    assert payload["rows"][1]["weekday"] == 3
    assert payload["rows"][2]["is_valid"] is False
    assert any("end_time" in issue for issue in payload["rows"][2]["issues"])


def test_timetable_import_apply_append_and_replace(client):
    headers = _auth_headers(client)
    existing_class_resp = client.post(
        "/classes",
        json={"name": "2APIC-1", "subject": "Math"},
        headers=headers,
    )
    assert existing_class_resp.status_code == 201
    existing_class_id = existing_class_resp.json()["id"]

    csv_content = _build_timetable_csv(
        [
            ("owner@school.edu", "2APIC-1", "Math", "Monday", "08:00", "09:00", "R12", "G1"),
            ("owner@school.edu", "2APIC-2", "Math", "Tuesday", "10:00", "11:00", "R15", "G2"),
        ]
    )

    dry_run_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "dry_run_only",
            "effective_from": "2026-09-01",
            "create_missing_classes": "false",
        },
        files={"file": ("emploi.csv", csv_content, "text/csv")},
    )
    assert dry_run_resp.status_code == 200
    dry_run_payload = dry_run_resp.json()
    assert dry_run_payload["planned_apply_rows"] == 1
    assert dry_run_payload["applied_rows"] == 0
    assert "2APIC-2" in dry_run_payload["unresolved_class_names"]

    append_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "append_new_slots",
            "effective_from": "2026-09-01",
            "create_missing_classes": "true",
        },
        files={"file": ("emploi.csv", csv_content, "text/csv")},
    )
    assert append_resp.status_code == 200
    append_payload = append_resp.json()
    assert append_payload["applied_rows"] == 2
    assert append_payload["created_classes_count"] == 1

    append_again_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "append_new_slots",
            "effective_from": "2026-09-01",
            "create_missing_classes": "false",
        },
        files={"file": ("emploi.csv", csv_content, "text/csv")},
    )
    assert append_again_resp.status_code == 200
    append_again_payload = append_again_resp.json()
    assert append_again_payload["applied_rows"] == 0
    assert append_again_payload["skipped_duplicate_rows"] == 2

    rules_existing_resp = client.get(f"/workflow/classes/{existing_class_id}/timetable-rules", headers=headers)
    assert rules_existing_resp.status_code == 200
    rules_existing = rules_existing_resp.json()
    assert len(rules_existing) == 1
    assert rules_existing[0]["start_time"] == "08:00:00"

    classes_resp = client.get("/classes", headers=headers)
    assert classes_resp.status_code == 200
    created_class = next((row for row in classes_resp.json() if row["name"] == "2APIC-2"), None)
    assert created_class is not None
    created_class_id = int(created_class["id"])

    replace_csv = _build_timetable_csv(
        [
            ("owner@school.edu", "2APIC-1", "Math", "Monday", "11:00", "12:00", "R12", "G1"),
        ]
    )
    replace_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "replace_future_from_date",
            "effective_from": "2026-09-01",
            "create_missing_classes": "false",
        },
        files={"file": ("emploi-replace.csv", replace_csv, "text/csv")},
    )
    assert replace_resp.status_code == 200
    assert replace_resp.json()["applied_rows"] == 1

    rules_existing_after_resp = client.get(f"/workflow/classes/{existing_class_id}/timetable-rules", headers=headers)
    assert rules_existing_after_resp.status_code == 200
    rules_existing_after = rules_existing_after_resp.json()
    assert len(rules_existing_after) == 1
    assert rules_existing_after[0]["start_time"] == "11:00:00"

    rules_created_resp = client.get(f"/workflow/classes/{created_class_id}/timetable-rules", headers=headers)
    assert rules_created_resp.status_code == 200
    rules_created = rules_created_resp.json()
    assert len(rules_created) == 1


def test_timetable_import_apply_ics(client):
    headers = _auth_headers(client)
    class_resp = client.post(
        "/classes",
        json={"name": "2APIC-1", "subject": "Math"},
        headers=headers,
    )
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    ics_content = _build_timetable_ics(
        [
            {
                "summary": "2APIC-1 | Math",
                "dtstart": "20260907T080000",
                "dtend": "20260907T090000",
                "rrule": "FREQ=WEEKLY;BYDAY=MO",
                "location": "R12",
                "description": "Group:G1",
            },
        ]
    )

    apply_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "append_new_slots",
            "effective_from": "2026-09-01",
            "create_missing_classes": "false",
        },
        files={"file": ("emploi.ics", ics_content, "text/calendar")},
    )
    assert apply_resp.status_code == 200
    payload = apply_resp.json()
    assert payload["applied_rows"] == 1
    assert payload["unresolved_class_names"] == []

    rules_resp = client.get(f"/workflow/classes/{class_id}/timetable-rules", headers=headers)
    assert rules_resp.status_code == 200
    rules = rules_resp.json()
    assert len(rules) == 1
    assert rules[0]["weekday"] == 1
    assert rules[0]["start_time"] == "08:00:00"
    assert rules[0]["end_time"] == "09:00:00"


def test_timetable_rule_exception_create_list_delete(client):
    headers = _auth_headers(client)
    class_name = f"2APIC-EXC-{uuid.uuid4().hex[:6]}"
    class_resp = client.post(
        "/classes",
        json={"name": class_name, "subject": "Math"},
        headers=headers,
    )
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    csv_content = _build_timetable_csv(
        [
            ("owner@school.edu", class_name, "Math", "Monday", "08:00", "09:00", "R12", "G1"),
        ]
    )
    apply_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "append_new_slots",
            "effective_from": "2026-09-01",
            "create_missing_classes": "false",
        },
        files={"file": ("emploi.csv", csv_content, "text/csv")},
    )
    assert apply_resp.status_code == 200
    assert apply_resp.json()["applied_rows"] == 1

    rules_resp = client.get(f"/workflow/classes/{class_id}/timetable-rules", headers=headers)
    assert rules_resp.status_code == 200
    rules = rules_resp.json()
    assert len(rules) == 1
    rule_id = int(rules[0]["id"])

    create_exception_resp = client.post(
        f"/workflow/classes/{class_id}/timetable-exceptions",
        headers=headers,
        json={
            "rule_id": rule_id,
            "exception_date": "2026-09-07",
            "exception_type": "cancel",
            "note": "Teacher absent",
        },
    )
    assert create_exception_resp.status_code == 201
    exception_row = create_exception_resp.json()
    assert int(exception_row["class_id"]) == class_id
    assert int(exception_row["rule_id"]) == rule_id
    assert exception_row["exception_type"] == "cancel"
    assert exception_row["exception_date"] == "2026-09-07"

    create_duplicate_resp = client.post(
        f"/workflow/classes/{class_id}/timetable-exceptions",
        headers=headers,
        json={
            "rule_id": rule_id,
            "exception_date": "2026-09-07",
            "exception_type": "cancel",
            "note": "Updated note",
        },
    )
    assert create_duplicate_resp.status_code == 201
    duplicate_row = create_duplicate_resp.json()
    assert int(duplicate_row["id"]) == int(exception_row["id"])
    assert duplicate_row["note"] == "Updated note"

    update_cancel_resp = client.patch(
        f"/workflow/timetable-exceptions/{exception_row['id']}",
        headers=headers,
        json={
            "exception_date": "2026-09-14",
            "note": "Updated by patch",
        },
    )
    assert update_cancel_resp.status_code == 200
    updated_cancel_row = update_cancel_resp.json()
    assert updated_cancel_row["exception_date"] == "2026-09-14"
    assert updated_cancel_row["note"] == "Updated by patch"

    list_resp = client.get(
        f"/workflow/classes/{class_id}/timetable-exceptions?date_from=2026-09-01&date_to=2026-09-30",
        headers=headers,
    )
    assert list_resp.status_code == 200
    rows = list_resp.json()
    assert len(rows) == 1
    assert int(rows[0]["id"]) == int(exception_row["id"])

    existing_session_resp = client.post(
        f"/workflow/classes/{class_id}/sessions",
        headers=headers,
        json={
            "session_date": "2026-09-10",
            "start_time": "12:00:00",
            "end_time": "13:00:00",
            "note": "Existing real session",
        },
    )
    assert existing_session_resp.status_code == 201

    move_exception_resp = client.post(
        f"/workflow/classes/{class_id}/timetable-exceptions",
        headers=headers,
        json={
            "rule_id": rule_id,
            "exception_date": "2026-10-05",
            "exception_type": "move",
            "target_date": "2026-09-08",
            "target_start_time": "10:00:00",
            "target_end_time": "11:00:00",
            "note": "Shifted to Tuesday",
        },
    )
    assert move_exception_resp.status_code == 201
    move_row = move_exception_resp.json()
    assert move_row["exception_type"] == "move"
    assert move_row["target_date"] == "2026-09-08"
    assert move_row["target_start_time"] == "10:00:00"
    assert move_row["target_end_time"] == "11:00:00"

    move_list_resp = client.get(
        f"/workflow/classes/{class_id}/timetable-exceptions?date_from=2026-09-08&date_to=2026-09-08",
        headers=headers,
    )
    assert move_list_resp.status_code == 200
    move_rows = move_list_resp.json()
    assert len(move_rows) == 1
    assert int(move_rows[0]["id"]) == int(move_row["id"])

    create_overlap_blocked_resp = client.post(
        f"/workflow/classes/{class_id}/timetable-exceptions",
        headers=headers,
        json={
            "rule_id": rule_id,
            "exception_date": "2026-10-12",
            "exception_type": "move",
            "target_date": "2026-09-10",
            "target_start_time": "12:00:00",
            "target_end_time": "13:00:00",
            "note": "Should be blocked without override",
        },
    )
    assert create_overlap_blocked_resp.status_code == 409
    assert "allow_overlap=true" in str(create_overlap_blocked_resp.json().get("detail", ""))

    create_overlap_allowed_resp = client.post(
        f"/workflow/classes/{class_id}/timetable-exceptions",
        headers=headers,
        json={
            "rule_id": rule_id,
            "exception_date": "2026-10-12",
            "exception_type": "move",
            "target_date": "2026-09-10",
            "target_start_time": "12:00:00",
            "target_end_time": "13:00:00",
            "allow_overlap": True,
            "note": "Allowed with explicit override",
        },
    )
    assert create_overlap_allowed_resp.status_code == 201
    overlap_move_row = create_overlap_allowed_resp.json()
    assert overlap_move_row["target_date"] == "2026-09-10"
    assert overlap_move_row["target_start_time"] == "12:00:00"

    update_overlap_note_only_resp = client.patch(
        f"/workflow/timetable-exceptions/{overlap_move_row['id']}",
        headers=headers,
        json={
            "note": "Note-only patch should not require override",
        },
    )
    assert update_overlap_note_only_resp.status_code == 200
    assert update_overlap_note_only_resp.json()["note"] == "Note-only patch should not require override"

    update_move_blocked_resp = client.patch(
        f"/workflow/timetable-exceptions/{move_row['id']}",
        headers=headers,
        json={
            "target_date": "2026-09-10",
            "target_start_time": "12:00:00",
            "target_end_time": "13:00:00",
            "note": "Should fail without override",
        },
    )
    assert update_move_blocked_resp.status_code == 409
    assert "allow_overlap=true" in str(update_move_blocked_resp.json().get("detail", ""))

    update_move_resp = client.patch(
        f"/workflow/timetable-exceptions/{move_row['id']}",
        headers=headers,
        json={
            "target_date": "2026-09-10",
            "target_start_time": "12:00:00",
            "target_end_time": "13:00:00",
            "allow_overlap": True,
            "note": "Patched move",
        },
    )
    assert update_move_resp.status_code == 200
    updated_move_row = update_move_resp.json()
    assert updated_move_row["target_date"] == "2026-09-10"
    assert updated_move_row["target_start_time"] == "12:00:00"
    assert updated_move_row["target_end_time"] == "13:00:00"
    assert updated_move_row["note"] == "Patched move"

    delete_resp = client.delete(f"/workflow/timetable-exceptions/{exception_row['id']}", headers=headers)
    assert delete_resp.status_code == 204
    delete_move_resp = client.delete(f"/workflow/timetable-exceptions/{move_row['id']}", headers=headers)
    assert delete_move_resp.status_code == 204
    delete_overlap_move_resp = client.delete(f"/workflow/timetable-exceptions/{overlap_move_row['id']}", headers=headers)
    assert delete_overlap_move_resp.status_code == 204

    list_after_delete = client.get(f"/workflow/classes/{class_id}/timetable-exceptions", headers=headers)
    assert list_after_delete.status_code == 200
    assert list_after_delete.json() == []


def test_timetable_import_apply_with_class_mappings(client):
    headers = _auth_headers(client)
    class_resp = client.post(
        "/classes",
        json={"name": "2APIC-1", "subject": "Math"},
        headers=headers,
    )
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    csv_content = _build_timetable_csv(
        [
            ("owner@school.edu", "2APIC-LEGACY", "Math", "Monday", "08:00", "09:00", "R12", "G1"),
        ]
    )
    mapping_payload = json.dumps({"2APIC-LEGACY": class_id})

    dry_run_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "dry_run_only",
            "effective_from": "2026-09-01",
            "create_missing_classes": "false",
            "class_mappings_json": mapping_payload,
        },
        files={"file": ("emploi.csv", csv_content, "text/csv")},
    )
    assert dry_run_resp.status_code == 200
    dry_run_payload = dry_run_resp.json()
    assert dry_run_payload["planned_apply_rows"] == 1
    assert dry_run_payload["applied_rows"] == 0
    assert dry_run_payload["unresolved_class_names"] == []
    assert any(
        row["action"] == "dry_run_ready" and int(row.get("class_id") or 0) == class_id
        for row in dry_run_payload["rows"]
    )

    append_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "append_new_slots",
            "effective_from": "2026-09-01",
            "create_missing_classes": "false",
            "class_mappings_json": mapping_payload,
        },
        files={"file": ("emploi.csv", csv_content, "text/csv")},
    )
    assert append_resp.status_code == 200
    append_payload = append_resp.json()
    assert append_payload["applied_rows"] == 1
    assert append_payload["unresolved_class_names"] == []

    rules_resp = client.get(f"/workflow/classes/{class_id}/timetable-rules", headers=headers)
    assert rules_resp.status_code == 200
    rows = rules_resp.json()
    assert len(rows) == 1
    assert rows[0]["start_time"] == "08:00:00"


def test_timetable_import_saved_class_aliases_reused_and_manageable(client):
    headers = _auth_headers(client)
    class_resp = client.post(
        "/classes",
        json={"name": "2APIC-1", "subject": "Math"},
        headers=headers,
    )
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])
    class_resp_2 = client.post(
        "/classes",
        json={"name": "2APIC-2", "subject": "Math"},
        headers=headers,
    )
    assert class_resp_2.status_code == 201
    class_id_2 = int(class_resp_2.json()["id"])

    csv_content = _build_timetable_csv(
        [
            ("owner@school.edu", "2APIC LEGACY", "Math", "Monday", "08:00", "09:00", "R12", "G1"),
        ]
    )
    mapping_payload = json.dumps({"2APIC LEGACY": class_id})

    apply_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "append_new_slots",
            "effective_from": "2026-09-01",
            "create_missing_classes": "false",
            "class_mappings_json": mapping_payload,
            "save_class_mappings": "true",
        },
        files={"file": ("emploi.csv", csv_content, "text/csv")},
    )
    assert apply_resp.status_code == 200
    assert apply_resp.json()["applied_rows"] == 1

    list_resp = client.get("/workflow/timetable/class-mappings", headers=headers)
    assert list_resp.status_code == 200
    rows = list_resp.json()
    target_row = next((row for row in rows if str(row.get("alias_name")) == "2APIC LEGACY"), None)
    assert target_row is not None
    assert int(target_row["class_id"]) == class_id
    mapping_id = int(target_row["id"])

    update_resp = client.patch(
        f"/workflow/timetable/class-mappings/{mapping_id}",
        headers=headers,
        json={"class_id": class_id_2},
    )
    assert update_resp.status_code == 200
    update_payload = update_resp.json()
    assert int(update_payload["id"]) == mapping_id
    assert int(update_payload["class_id"]) == class_id_2

    dry_run_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "dry_run_only",
            "effective_from": "2026-10-01",
            "create_missing_classes": "false",
        },
        files={"file": ("emploi.csv", csv_content, "text/csv")},
    )
    assert dry_run_resp.status_code == 200
    dry_run_payload = dry_run_resp.json()
    assert dry_run_payload["unresolved_class_names"] == []
    assert dry_run_payload["planned_apply_rows"] == 1
    assert any(
        row["action"] == "dry_run_ready" and int(row.get("class_id") or 0) == class_id_2
        for row in dry_run_payload["rows"]
    )

    delete_resp = client.delete(f"/workflow/timetable/class-mappings/{mapping_id}", headers=headers)
    assert delete_resp.status_code == 204

    list_after_delete = client.get("/workflow/timetable/class-mappings", headers=headers)
    assert list_after_delete.status_code == 200
    assert all(int(row["id"]) != mapping_id for row in list_after_delete.json())


def test_timetable_class_mapping_bulk_save(client):
    headers = _auth_headers(client)
    class_1_resp = client.post("/classes", json={"name": "2APIC-1", "subject": "Math"}, headers=headers)
    assert class_1_resp.status_code == 201
    class_1_id = int(class_1_resp.json()["id"])
    class_2_resp = client.post("/classes", json={"name": "2APIC-2", "subject": "Math"}, headers=headers)
    assert class_2_resp.status_code == 201
    class_2_id = int(class_2_resp.json()["id"])

    bulk_save_resp = client.post(
        "/workflow/timetable/class-mappings/bulk-save",
        headers=headers,
        json={
            "mappings": {
                "2APIC LEGACY": class_1_id,
                "   ": class_1_id,
            },
        },
    )
    assert bulk_save_resp.status_code == 200
    bulk_save_payload = bulk_save_resp.json()
    assert bulk_save_payload["saved_count"] == 1
    assert bulk_save_payload["skipped_count"] == 1
    assert len(bulk_save_payload["rows"]) == 1
    assert int(bulk_save_payload["rows"][0]["class_id"]) == class_1_id

    bulk_update_resp = client.post(
        "/workflow/timetable/class-mappings/bulk-save",
        headers=headers,
        json={
            "mappings": {
                "2APIC LEGACY": class_2_id,
            },
        },
    )
    assert bulk_update_resp.status_code == 200
    bulk_update_payload = bulk_update_resp.json()
    assert bulk_update_payload["saved_count"] == 1
    assert int(bulk_update_payload["rows"][0]["class_id"]) == class_2_id

    csv_content = _build_timetable_csv(
        [
            ("owner@school.edu", "2APIC LEGACY", "Math", "Monday", "08:00", "09:00", "R12", "G1"),
        ]
    )
    dry_run_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "dry_run_only",
            "effective_from": "2026-10-01",
            "create_missing_classes": "false",
        },
        files={"file": ("emploi.csv", csv_content, "text/csv")},
    )
    assert dry_run_resp.status_code == 200
    dry_run_payload = dry_run_resp.json()
    assert dry_run_payload["unresolved_class_names"] == []
    assert any(
        row["action"] == "dry_run_ready" and int(row.get("class_id") or 0) == class_2_id
        for row in dry_run_payload["rows"]
    )


def test_timetable_import_apply_supports_effective_to_window(client):
    headers = _auth_headers(client)
    class_resp = client.post(
        "/classes",
        json={"name": "2APIC-Window", "subject": "Math"},
        headers=headers,
    )
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    csv_content = _build_timetable_csv(
        [
            ("owner@school.edu", "2APIC-Window", "Math", "Monday", "08:00", "09:00", "R12", "G1"),
        ]
    )

    apply_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "append_new_slots",
            "effective_from": "2026-09-01",
            "effective_to": "2026-12-31",
            "create_missing_classes": "false",
        },
        files={"file": ("emploi.csv", csv_content, "text/csv")},
    )
    assert apply_resp.status_code == 200
    payload = apply_resp.json()
    assert payload["applied_rows"] == 1
    assert payload["effective_to"] == "2026-12-31"

    rules_resp = client.get(f"/workflow/classes/{class_id}/timetable-rules", headers=headers)
    assert rules_resp.status_code == 200
    rows = rules_resp.json()
    assert len(rows) == 1
    assert rows[0]["effective_from"] == "2026-09-01"
    assert rows[0]["effective_to"] == "2026-12-31"

    invalid_window_resp = client.post(
        "/workflow/timetable/import/apply",
        headers=headers,
        data={
            "mode": "append_new_slots",
            "effective_from": "2026-09-01",
            "effective_to": "2026-08-31",
            "create_missing_classes": "false",
        },
        files={"file": ("emploi.csv", csv_content, "text/csv")},
    )
    assert invalid_window_resp.status_code == 400


def test_workflow_checklist_reorder_supports_reparent_and_reposition(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Workflow Reorder 2APIC", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    start_unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "chapter", "title": "Reorder Unit", "source_text": "Chapter reorder seed"},
    )
    assert start_unit_resp.status_code == 201
    unit = start_unit_resp.json()

    root_a_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items",
        headers=headers,
        json={"title": "Root A", "item_kind": "chapter", "parent_item_id": None},
    )
    assert root_a_resp.status_code == 201
    root_a_id = root_a_resp.json()["id"]

    root_b_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items",
        headers=headers,
        json={"title": "Root B", "item_kind": "chapter", "parent_item_id": None},
    )
    assert root_b_resp.status_code == 201
    root_b_id = root_b_resp.json()["id"]

    child_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items",
        headers=headers,
        json={"title": "Child A1", "item_kind": "exercise", "parent_item_id": root_a_id},
    )
    assert child_resp.status_code == 201
    child_id = child_resp.json()["id"]

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_resp.status_code == 200
    flat_before = _flatten_checklist(workspace_resp.json()["active_unit"]["checklist"])
    rows_by_id = {
        int(item["id"]): {
            "id": int(item["id"]),
            "parent_item_id": item.get("parent_item_id"),
            "position": int(item.get("position", 0)),
        }
        for item in flat_before
    }

    rows_by_id[root_b_id]["position"] = 0
    rows_by_id[root_a_id]["position"] = 1
    rows_by_id[child_id]["parent_item_id"] = root_b_id
    rows_by_id[child_id]["position"] = 0

    reorder_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items/reorder",
        headers=headers,
        json={"items": list(rows_by_id.values())},
    )
    assert reorder_resp.status_code == 200
    assert reorder_resp.json()["updated"] >= len(rows_by_id)
    assert reorder_resp.json()["moved"] >= 1

    workspace_after = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_after.status_code == 200
    active_unit = workspace_after.json()["active_unit"]
    flat_after = _flatten_checklist(active_unit["checklist"])
    by_id = {int(item["id"]): item for item in flat_after}
    assert int(by_id[child_id]["parent_item_id"]) == root_b_id

    roots = [row for row in active_unit["checklist"] if row.get("parent_item_id") is None]
    root_ids = [int(row["id"]) for row in roots]
    assert root_ids.index(root_b_id) < root_ids.index(root_a_id)
    root_b_children = [int(row["id"]) for row in by_id[root_b_id].get("children", [])]
    assert child_id in root_b_children


def test_workflow_leaf_toggle_updates_calendar_checked_item_ids(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Workflow Cascade 2APIC", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]
    roster = _build_roster_file([("STD001", "Alice"), ("STD002", "Bob")])
    import_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200

    start_unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "chapter", "title": "Cascade Unit", "source_text": "Seed lesson"},
    )
    assert start_unit_resp.status_code == 201
    unit = start_unit_resp.json()

    parent_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items",
        headers=headers,
        json={"title": "Parent topic", "item_kind": "chapter", "parent_item_id": None},
    )
    assert parent_resp.status_code == 201
    parent_id = parent_resp.json()["id"]

    child_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items",
        headers=headers,
        json={"title": "Child practice", "item_kind": "exercise", "parent_item_id": parent_id},
    )
    assert child_resp.status_code == 201
    child_id = child_resp.json()["id"]

    start_session_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/start",
        headers=headers,
        json={"absent_student_ids": []},
    )
    assert start_session_resp.status_code == 201
    session_id = start_session_resp.json()["id"]

    toggle_parent_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/{session_id}/items/{parent_id}/toggle",
        headers=headers,
        json={"checked": True},
    )
    assert toggle_parent_resp.status_code == 409

    toggle_child_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/{session_id}/items/{child_id}/toggle",
        headers=headers,
        json={"checked": True},
    )
    assert toggle_child_resp.status_code == 200
    assert toggle_child_resp.json()["is_completed"] is True

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_resp.status_code == 200
    flat = _flatten_checklist(workspace_resp.json()["active_unit"]["checklist"])
    by_id = {int(row["id"]): row for row in flat}
    assert by_id[child_id]["is_completed"] is True
    assert by_id[parent_id]["is_completed"] is True

    end_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/{session_id}/end",
        headers=headers,
        json={"end_time": "23:00:00"},
    )
    assert end_resp.status_code == 200

    calendar_resp = client.get(f"/workflow/classes/{class_id}/calendar", headers=headers)
    assert calendar_resp.status_code == 200
    event = next((row for row in calendar_resp.json() if row["session_id"] == session_id), None)
    assert event is not None
    checked_items = [str(value or "") for value in event.get("checked_items", [])]
    checked_text = " | ".join(checked_items).lower()
    assert event.get("checked_items_count", 0) == 1
    assert len(event.get("checked_item_ids", [])) >= 1
    assert all(int(value) > 0 for value in event.get("checked_item_ids", []))
    assert "parent topic" in checked_text
    assert "child practice" in checked_text


def test_workflow_end_session_rejects_end_before_start(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Workflow End Time Validation", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    roster = _build_roster_file([("STD001", "Alice")])
    import_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "chapter", "title": "Time Validation Unit", "source_text": "chapter seed"},
    )
    assert unit_resp.status_code == 201

    session_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/start",
        headers=headers,
        json={"absent_student_ids": []},
    )
    assert session_resp.status_code == 201
    session_id = session_resp.json()["id"]

    invalid_end_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/{session_id}/end",
        headers=headers,
        json={"start_time": "10:00:00", "end_time": "09:00:00"},
    )
    assert invalid_end_resp.status_code == 400
    assert "end_time must be greater than or equal to start_time" in str(invalid_end_resp.json().get("detail", ""))


def test_workflow_end_with_attendance_only_keeps_session_open(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Workflow Attendance Update Open Session", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    roster = _build_roster_file([("STD001", "Alice"), ("STD002", "Bob")])
    import_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200
    students = client.get(f"/classes/{class_id}/students", headers=headers).json()
    absent_id = students[0]["id"]

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "chapter", "title": "Attendance Update Unit", "source_text": "chapter seed"},
    )
    assert unit_resp.status_code == 201

    session_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/start",
        headers=headers,
        json={"absent_student_ids": []},
    )
    assert session_resp.status_code == 201
    session_id = session_resp.json()["id"]

    attendance_only_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/{session_id}/end",
        headers=headers,
        json={"absent_student_ids": [absent_id]},
    )
    assert attendance_only_resp.status_code == 200
    assert attendance_only_resp.json()["end_time"] is None
    assert attendance_only_resp.json()["absent_count"] == 1

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_resp.status_code == 200
    active_session = workspace_resp.json().get("active_session")
    assert active_session is not None
    assert int(active_session["id"]) == session_id

    close_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/{session_id}/end",
        headers=headers,
        json={},
    )
    assert close_resp.status_code == 200
    assert close_resp.json()["end_time"] is not None


def test_workflow_start_unit_accepts_source_text_without_pdf(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Workflow Manual 2APIC", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    start_unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Chapitre 3",
            "source_text": "Chapitre 3\n1.1 Fractions\nDefinition: fraction equivalente",
        },
    )
    assert start_unit_resp.status_code == 201
    unit = start_unit_resp.json()
    assert unit["unit_type"] == "chapter"
    assert unit["progress_total"] >= 1
    assert unit["checklist"]


def test_workflow_start_unit_rejects_blank_title_and_non_positive_planned_hours(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Workflow Validation Unit Start", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    blank_title_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "   ",
            "source_text": "chapter seed",
        },
    )
    assert blank_title_resp.status_code == 400
    assert "Unit title is required" in str(blank_title_resp.json().get("detail", ""))

    invalid_hours_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Chapitre Validation",
            "source_text": "chapter seed",
            "planned_hours": "0",
        },
    )
    assert invalid_hours_resp.status_code == 400
    assert "planned_hours must be greater than zero" in str(invalid_hours_resp.json().get("detail", ""))


def test_workflow_start_unit_from_pdf_generates_todo_tree(client, monkeypatch):
    from app.services import workflow as workflow_service

    monkeypatch.setattr(workflow_service, "OPENAI_API_KEY", "")
    monkeypatch.setattr(
        workflow_service,
        "_ocr_pdf_pages",
        lambda _source, *, max_pages: [
            "Chapitre 2 : Calcul litteral - Identites remarquables",
            "1. Developpement",
            "1.1 Developpement par la distributivite",
            "Propriete : k(a+b)=ka+kb",
            "Exemple : developper des expressions",
            "2. Identites remarquables",
            "2.1 (a+b)^2",
        ],
    )
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Workflow PDF 2APIC", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    pdf_bytes = _build_pdf_file(
        [
            "Chapitre 2 : Calcul litteral - Identites remarquables",
            "1. Developpement",
            "1.1 Developpement par la distributivite",
            "Propriete : k(a+b)=ka+kb",
            "Exemple : developper des expressions",
            "2. Identites remarquables",
            "2.1 (a+b)^2",
        ]
    )
    start_unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "chapter", "title": "Chapitre 2", "planned_hours": "6"},
        files={"file": ("chapter.pdf", pdf_bytes, "application/pdf")},
    )
    assert start_unit_resp.status_code == 201
    unit = start_unit_resp.json()
    assert unit["unit_type"] == "chapter"
    assert unit["progress_total"] >= 4
    checklist = unit["checklist"]
    assert checklist

    def flatten(nodes):
        output = []
        for node in nodes:
            output.append(node)
            output.extend(flatten(node.get("children", [])))
        return output

    flat = flatten(checklist)
    titles = [str(node.get("title", "")).lower() for node in flat]
    assert any("chapitre 2" in title for title in titles)
    assert any("1.1" in title or "distributivite" in title for title in titles)
    assert any(int(node.get("depth", 0)) > 0 for node in flat)


def test_infer_exam_title_from_source_text_prefers_visible_exam_heading():
    from app.routers import workflow as workflow_router

    extracted_text = "\n".join(
        [
            "Devoir surveille N 2 : Fractions",
            "Classe : 3AC",
            "Exercice 1",
        ]
    )

    title = workflow_router._infer_exam_title_from_source_text(
        extracted_text,
        fallback_title="exam-fractions-2026",
        file_name="exam-fractions-2026.pdf",
    )

    assert title == "Devoir surveille N 2 : Fractions"


def test_infer_exam_title_from_source_text_skips_metadata_and_uses_topic_line():
    from app.routers import workflow as workflow_router

    extracted_text = "\n".join(
        [
            "DS N 2",
            "Classe : 3AC",
            "Fractions et operations",
            "Exercice 1",
        ]
    )

    title = workflow_router._infer_exam_title_from_source_text(
        extracted_text,
        fallback_title="exam-ds2-fractions",
        file_name="exam-ds2-fractions.pdf",
    )

    assert title == "DS N 2 - Fractions et operations"


def test_infer_exam_title_from_source_text_keeps_detailed_heading_without_forcing_subtitle():
    from app.routers import workflow as workflow_router

    extracted_text = "\n".join(
        [
            "Controle continu N 3 de mathematiques",
            "Classe : 2AC",
            "Exercice 1",
        ]
    )

    title = workflow_router._infer_exam_title_from_source_text(
        extracted_text,
        fallback_title="controle-continu-maths-3",
        file_name="controle-continu-maths-3.pdf",
    )

    assert title == "Controle continu N 3 de mathematiques"


def _unique_owner_headers(client) -> dict[str, str]:
    from app.database import SessionLocal
    from app.models import User, UserRole
    from app.services.auth import create_access_token, hash_password

    db = SessionLocal()
    try:
        user = User(
            email=f"owner_{uuid.uuid4().hex[:8]}@app.local",
            full_name="Owner",
            password_hash=hash_password("OwnerPass123"),
            role=UserRole.OWNER,
            is_active=True,
        )
        db.add(user)
        db.commit()
        db.refresh(user)
        token = create_access_token(db, user)
        return {"Authorization": f"Bearer {token.token}"}
    finally:
        db.close()


def test_workflow_start_exam_unit_from_pdf_infers_better_title(client):
    headers = _unique_owner_headers(client)
    class_resp = client.post("/classes", json={"name": "Workflow Exam Title", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    pdf_bytes = _build_pdf_file(
        [
            "Devoir surveille N 2 : Fractions",
            "Classe : 3AC",
            "Exercice 1",
            "Exercice 2",
        ]
    )
    start_unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "exam", "title": "exam-fractions-2026"},
        files={"file": ("exam-fractions-2026.pdf", pdf_bytes, "application/pdf")},
    )
    assert start_unit_resp.status_code == 201
    unit = start_unit_resp.json()
    assert unit["unit_type"] == "exam"
    assert unit["title"] == "Devoir surveille N 2 : Fractions"
    assert unit["checklist"]


def test_workflow_start_notebooklm_for_exam_unit_stores_provider_context(client, monkeypatch):
    from app.services import workflow_generation

    monkeypatch.setattr(
        workflow_generation,
        "initialize_unit_notebooklm_context",
        lambda **kwargs: (
            {
                "provider": "notebooklm",
                "notebook_id": "nb-exam-1",
                "source_ids": ["src-exam-1"],
                "notebook_title": "Teacher Progress - Devoir surveille N 2 : Fractions",
                "notebook_role": "exam_outline",
            },
            {
                "provider": "notebooklm",
                "action": "initialize_context",
                "notebook_id": "nb-exam-1",
                "source_ids": ["src-exam-1"],
                "notebook_role": "exam_outline",
            },
        ),
    )

    headers = _unique_owner_headers(client)
    class_resp = client.post("/classes", json={"name": "Workflow Exam NotebookLM", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    pdf_bytes = _build_pdf_file(
        [
            "Devoir surveille N 2 : Fractions",
            "Exercice 1",
            "Exercice 2",
        ]
    )
    start_unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "exam", "title": "exam-fractions-2026"},
        files={"file": ("exam-fractions-2026.pdf", pdf_bytes, "application/pdf")},
    )
    assert start_unit_resp.status_code == 201
    unit_id = start_unit_resp.json()["id"]

    notebook_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/notebooklm/start",
        headers=headers,
    )
    assert notebook_resp.status_code == 200
    payload = notebook_resp.json()
    assert payload["id"] == unit_id
    assert payload["extraction_source"] == "template"
    assert payload["extraction_notebook_role"] == "exam_outline"


def test_workflow_start_unit_rejects_non_pdf_document(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Workflow PDF Strict", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    text_content = b"Chapitre 1\n1.1 Developpement"
    start_unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "chapter", "title": "Chapitre 1"},
        files={"file": ("chapter.txt", text_content, "text/plain")},
    )
    assert start_unit_resp.status_code == 400
    assert "Only PDF documents are supported" in str(start_unit_resp.json().get("detail", ""))


def test_workflow_reopen_closed_unit_flow(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Workflow Reopen Flow", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Reopen Unit",
            "source_text": "chapter seed",
        },
    )
    assert unit_resp.status_code == 201
    first_unit_id = unit_resp.json()["id"]

    close_resp = client.post(f"/workflow/classes/{class_id}/units/{first_unit_id}/close", headers=headers)
    assert close_resp.status_code == 200

    reopen_resp = client.post(f"/workflow/classes/{class_id}/units/{first_unit_id}/reopen", headers=headers)
    assert reopen_resp.status_code == 200
    reopened = reopen_resp.json()
    assert reopened["status"] == "active"
    assert reopened["closed_at"] is None

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_resp.status_code == 200
    assert int(workspace_resp.json()["active_unit"]["id"]) == first_unit_id

    reopen_again_resp = client.post(f"/workflow/classes/{class_id}/units/{first_unit_id}/reopen", headers=headers)
    assert reopen_again_resp.status_code == 409

    close_again_resp = client.post(f"/workflow/classes/{class_id}/units/{first_unit_id}/close", headers=headers)
    assert close_again_resp.status_code == 200

    second_unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Active Other Unit",
            "source_text": "chapter seed 2",
        },
    )
    assert second_unit_resp.status_code == 201

    reopen_blocked_resp = client.post(f"/workflow/classes/{class_id}/units/{first_unit_id}/reopen", headers=headers)
    assert reopen_blocked_resp.status_code == 409
    assert "active unit already exists" in str(reopen_blocked_resp.json().get("detail", "")).lower()


def test_workflow_delete_unit_deletes_linked_sessions(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Workflow Delete Unit", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Delete Me Unit",
            "source_text": "chapter seed",
        },
    )
    assert unit_resp.status_code == 201
    unit_id = int(unit_resp.json()["id"])

    first_session_resp = client.post(
        f"/workflow/classes/{class_id}/sessions",
        headers=headers,
        json={
            "session_date": "2026-03-03",
            "start_time": "08:00:00",
            "end_time": "09:00:00",
            "unit_id": unit_id,
            "absent_student_ids": [],
        },
    )
    assert first_session_resp.status_code == 201
    first_session_id = int(first_session_resp.json()["id"])

    second_session_resp = client.post(
        f"/workflow/classes/{class_id}/sessions",
        headers=headers,
        json={
            "session_date": "2026-03-04",
            "start_time": "08:00:00",
            "unit_id": unit_id,
            "absent_student_ids": [],
        },
    )
    assert second_session_resp.status_code == 201
    second_session_id = int(second_session_resp.json()["id"])

    unit_sessions_before = client.get(f"/workflow/units/{unit_id}/sessions", headers=headers)
    assert unit_sessions_before.status_code == 200
    assert len(unit_sessions_before.json()) >= 2

    delete_unit_resp = client.delete(f"/workflow/classes/{class_id}/units/{unit_id}", headers=headers)
    assert delete_unit_resp.status_code == 200
    delete_payload = delete_unit_resp.json()
    assert int(delete_payload["deleted_unit_id"]) == unit_id
    assert int(delete_payload["deleted_sessions_count"]) >= 2

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_resp.status_code == 200
    workspace = workspace_resp.json()
    assert workspace.get("active_unit") is None
    assert workspace.get("active_session") is None

    unit_sessions_after = client.get(f"/workflow/units/{unit_id}/sessions", headers=headers)
    assert unit_sessions_after.status_code == 404

    first_session_detail = client.get(f"/sessions/{first_session_id}", headers=headers)
    assert first_session_detail.status_code == 404
    second_session_detail = client.get(f"/sessions/{second_session_id}", headers=headers)
    assert second_session_detail.status_code == 404

    calendar_resp = client.get(f"/workflow/classes/{class_id}/calendar", headers=headers)
    assert calendar_resp.status_code == 200
    assert not any(int(row.get("unit_id") or 0) == unit_id for row in calendar_resp.json())


def test_classes_students_sessions_and_extraction(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "2APIC-3", "subject": "Math", "level": "2APIC"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    roster = _build_roster_file([("STD001", "Alice"), ("STD002", "Bob")])
    import_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200
    assert import_resp.json()["created"] == 2

    students = client.get(f"/classes/{class_id}/students", headers=headers).json()
    assert len(students) == 2

    session_resp = client.post(
        f"/classes/{class_id}/sessions",
        headers=headers,
        json={"session_date": "2026-03-02", "start_time": "08:00:00", "end_time": "09:00:00", "note": "chapter progress"},
    )
    assert session_resp.status_code == 201
    session_id = session_resp.json()["id"]

    attendance_payload = [
        {"student_id": students[0]["id"], "status": "present", "minutes_late": 0},
        {"student_id": students[1]["id"], "status": "late", "minutes_late": 10},
    ]
    attendance_resp = client.put(f"/sessions/{session_id}/attendance", json=attendance_payload, headers=headers)
    assert attendance_resp.status_code == 200
    assert len(attendance_resp.json()) == 2
    attendance_summary_resp = client.get(f"/classes/{class_id}/attendance-summary", headers=headers)
    assert attendance_summary_resp.status_code == 200
    assert attendance_summary_resp.json()["total_sessions"] == 1
    student_profile_resp = client.get(f"/classes/{class_id}/students/{students[0]['id']}/profile", headers=headers)
    assert student_profile_resp.status_code == 200
    profile = student_profile_resp.json()
    assert profile["student"]["student_code"] == students[0]["student_code"]
    assert profile["attendance"]["total_rows"] == 1
    student_profile_pdf_resp = client.get(
        f"/classes/{class_id}/students/{students[0]['id']}/reports/profile.pdf",
        headers=headers,
    )
    assert student_profile_pdf_resp.status_code == 200
    assert student_profile_pdf_resp.content.startswith(b"%PDF")
    dashboard_resp = client.get(f"/classes/{class_id}/dashboard", headers=headers)
    assert dashboard_resp.status_code == 200
    assert dashboard_resp.json()["counts"]["students"] == 2
    assert "attendance_trend" in dashboard_resp.json()
    assert "extraction_metrics" in dashboard_resp.json()
    assert "exam_trend" in dashboard_resp.json()

    upload_resp = client.post(
        f"/sessions/{session_id}/uploads",
        headers=headers,
        files={"file": ("board.jpg", b"fake-image", "image/jpeg")},
        data={"raw_text": "1 Algebra\n1.1 Equations\nActivity: Solve by substitution\nExercise 1 page 20"},
    )
    assert upload_resp.status_code == 201
    extraction = upload_resp.json()
    assert extraction["lesson_headings"] == ["1 Algebra", "1.1 Equations"]
    assert len(extraction["activities"]) == 1
    assert len(extraction["exercises"]) == 1
    assert extraction["provider"] in {"heuristic", "openai"}

    confirm_payload = {
        "items": [
            {"item_type": "lesson", "heading": "1 Algebra", "position": 1},
            {"item_type": "lesson", "heading": "1.1 Equations", "position": 2},
            {
                "item_type": "activity",
                "heading": "Activity: Solve by substitution",
                "content": "Activity: Solve by substitution",
                "position": 3,
            },
            {
                "item_type": "exercise",
                "heading": "Exercise 1 page 20",
                "content": "Exercise 1 page 20",
                "position": 4,
            },
        ]
    }
    confirm_resp = client.post(f"/sessions/{session_id}/confirm-extraction", json=confirm_payload, headers=headers)
    assert confirm_resp.status_code == 200
    assert len(confirm_resp.json()) == 4

    class_audit_resp = client.get(f"/classes/{class_id}/audit-logs", headers=headers)
    assert class_audit_resp.status_code == 200
    extraction_entries = [row for row in class_audit_resp.json()["items"] if row["action"] == "extraction.confirm"]
    assert extraction_entries
    extraction_details = extraction_entries[0]["details"] or {}
    assert "before_items" in extraction_details
    assert "after_items" in extraction_details

    update_resp = client.put(
        f"/sessions/{session_id}",
        headers=headers,
        json={"note": "updated chapter progress"},
    )
    assert update_resp.status_code == 200
    assert update_resp.json()["note"] == "updated chapter progress"

    session_detail_resp = client.get(f"/sessions/{session_id}", headers=headers)
    assert session_detail_resp.status_code == 200
    session_detail = session_detail_resp.json()
    assert session_detail["session"]["note"] == "updated chapter progress"
    assert len(session_detail["attendance"]) == 2
    assert len(session_detail["progress_items"]) == 4

    timeline_resp = client.get(f"/classes/{class_id}/timeline", headers=headers)
    assert timeline_resp.status_code == 200
    assert timeline_resp.json()["sessions"][0]["session_id"] == session_id
    timeline_filtered_resp = client.get(
        f"/classes/{class_id}/timeline?note_query=updated&has_progress=true",
        headers=headers,
    )
    assert timeline_filtered_resp.status_code == 200
    assert len(timeline_filtered_resp.json()["sessions"]) == 1
    timeline_reviewed_resp = client.get(
        f"/classes/{class_id}/timeline?has_reviewed_upload=true",
        headers=headers,
    )
    assert timeline_reviewed_resp.status_code == 200
    assert len(timeline_reviewed_resp.json()["sessions"]) == 1

    attendance_csv_resp = client.get(f"/classes/{class_id}/attendance-export.csv", headers=headers)
    assert attendance_csv_resp.status_code == 200
    assert "text/csv" in attendance_csv_resp.headers["content-type"]
    text = attendance_csv_resp.text
    assert "session_date,session_id,student_code,full_name,status,minutes_late,comment" in text
    assert "STD001" in text

    dashboard_after_extraction = client.get(f"/classes/{class_id}/dashboard", headers=headers)
    assert dashboard_after_extraction.status_code == 200
    metrics = dashboard_after_extraction.json()["extraction_metrics"]
    assert metrics["sample_size"] >= 1
    assert isinstance(metrics["average_confidence"], float)


def test_quick_submit_session_flow(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Quick Submit Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    roster = _build_roster_file([("STD701", "Meryem"), ("STD702", "Omar")])
    import_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200

    students = client.get(f"/classes/{class_id}/students", headers=headers).json()
    absent_id = students[0]["id"]
    quick_submit_resp = client.post(
        f"/classes/{class_id}/quick-submit",
        headers=headers,
        files={"file": ("board.jpg", b"fake-image", "image/jpeg")},
        data={
            "absent_student_ids": f"[{absent_id}]",
            "raw_text": "1 Geometry\n1.1 Triangles\nActivity: classify triangles\nExercise 1 page 44",
        },
    )
    assert quick_submit_resp.status_code == 201
    payload = quick_submit_resp.json()
    assert payload["class_id"] == class_id
    assert payload["absent_students"] == 1
    assert payload["lesson_headings_count"] == 2
    assert payload["activities_count"] == 1
    assert payload["exercises_count"] == 1
    assert payload["provider"] in {"heuristic", "openai"}

    session_id = payload["session_id"]
    detail_resp = client.get(f"/sessions/{session_id}", headers=headers)
    assert detail_resp.status_code == 200
    detail = detail_resp.json()
    assert len(detail["attendance"]) == 2
    absent_rows = [row for row in detail["attendance"] if row["status"] == "absent"]
    present_rows = [row for row in detail["attendance"] if row["status"] == "present"]
    assert len(absent_rows) == 1
    assert len(present_rows) == 1
    assert len(detail["progress_items"]) == 4
    assert len(detail["uploads"]) == 1
    assert detail["uploads"][0]["reviewed"] is True


def test_confirm_extraction_append_mode_keeps_existing_progress(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Append Mode Class"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    session_resp = client.post(
        f"/classes/{class_id}/sessions",
        headers=headers,
        json={"session_date": "2026-03-02"},
    )
    assert session_resp.status_code == 201
    session_id = session_resp.json()["id"]

    replace_resp = client.post(
        f"/sessions/{session_id}/confirm-extraction",
        headers=headers,
        json={"items": [{"item_type": "lesson", "heading": "Intro", "position": 99}]},
    )
    assert replace_resp.status_code == 200
    replace_rows = replace_resp.json()
    assert [row["heading"] for row in replace_rows] == ["Intro"]
    assert [row["position"] for row in replace_rows] == [1]

    append_resp = client.post(
        f"/sessions/{session_id}/confirm-extraction",
        headers=headers,
        json={
            "mode": "append",
            "items": [
                {
                    "item_type": "activity",
                    "heading": "Activity",
                    "content": "Warm up",
                    "position": 20,
                },
                {
                    "item_type": "exercise",
                    "heading": "Exercise",
                    "content": "Page 44",
                    "position": 10,
                },
            ],
        },
    )
    assert append_resp.status_code == 200
    append_rows = append_resp.json()
    assert [row["heading"] for row in append_rows] == ["Intro", "Exercise", "Activity"]
    assert [row["position"] for row in append_rows] == [1, 2, 3]

    detail_resp = client.get(f"/sessions/{session_id}", headers=headers)
    assert detail_resp.status_code == 200
    assert len(detail_resp.json()["progress_items"]) == 3

    audit_resp = client.get(f"/classes/{class_id}/audit-logs", headers=headers)
    assert audit_resp.status_code == 200
    extraction_entries = [row for row in audit_resp.json()["items"] if row["action"] == "extraction.confirm"]
    assert extraction_entries
    assert any((entry.get("details") or {}).get("mode") == "append" for entry in extraction_entries)


def test_get_latest_session_upload_returns_normalized_items(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Latest Upload Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    session_resp = client.post(
        f"/classes/{class_id}/sessions",
        headers=headers,
        json={"session_date": "2026-03-02", "start_time": "08:00:00", "end_time": "09:00:00"},
    )
    assert session_resp.status_code == 201
    session_id = session_resp.json()["id"]

    latest_missing_resp = client.get(f"/sessions/{session_id}/uploads/latest", headers=headers)
    assert latest_missing_resp.status_code == 404

    upload_resp = client.post(
        f"/sessions/{session_id}/uploads",
        headers=headers,
        files={"file": ("board.jpg", b"fake-image", "image/jpeg")},
        data={"raw_text": "1 Algebra\nActivity: Solve quickly\nExercise 1 page 20"},
    )
    assert upload_resp.status_code == 201
    upload_payload = upload_resp.json()
    assert upload_payload["items"]
    upload_hint_ids = [str(row.get("hint_id") or "") for row in upload_payload["items"]]
    assert all(hint.startswith("ex_") for hint in upload_hint_ids)
    assert len(upload_hint_ids) == len(set(upload_hint_ids))

    latest_resp = client.get(f"/sessions/{session_id}/uploads/latest", headers=headers)
    assert latest_resp.status_code == 200
    latest_payload = latest_resp.json()
    assert latest_payload["upload_id"] == upload_payload["upload_id"]
    assert latest_payload["session_id"] == session_id
    assert latest_payload["reviewed"] is False
    assert latest_payload["items"]
    assert latest_payload["items"][0]["item_type"] == "lesson"
    assert latest_payload["items"][0]["position"] == 1
    latest_hint_ids = [str(row.get("hint_id") or "") for row in latest_payload["items"]]
    assert latest_hint_ids == upload_hint_ids

    confirm_resp = client.post(
        f"/sessions/{session_id}/confirm-extraction",
        headers=headers,
        json={"items": latest_payload["items"]},
    )
    assert confirm_resp.status_code == 200
    assert len(confirm_resp.json()) >= 1

    latest_after_confirm_resp = client.get(f"/sessions/{session_id}/uploads/latest", headers=headers)
    assert latest_after_confirm_resp.status_code == 200
    assert latest_after_confirm_resp.json()["reviewed"] is True


def test_quick_submit_rejects_unknown_absent_student(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Quick Submit Validation"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    roster = _build_roster_file([("STD801", "Yasmin")])
    import_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200

    invalid_absent_resp = client.post(
        f"/classes/{class_id}/quick-submit",
        headers=headers,
        files={"file": ("board.jpg", b"fake-image", "image/jpeg")},
        data={
            "absent_student_ids": "[999999]",
            "raw_text": "1 Numbers",
        },
    )
    assert invalid_absent_resp.status_code == 400
    assert "Unknown student ids" in str(invalid_absent_resp.json()["detail"])


def test_class_archive_restore_and_write_lock(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Archive Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    list_before = client.get("/classes", headers=headers)
    assert list_before.status_code == 200
    assert any(item["id"] == class_id for item in list_before.json())

    archive_resp = client.post(f"/classes/{class_id}/archive", headers=headers)
    assert archive_resp.status_code == 200
    assert archive_resp.json()["is_archived"] is True

    list_after_archive = client.get("/classes", headers=headers)
    assert list_after_archive.status_code == 200
    assert all(item["id"] != class_id for item in list_after_archive.json())

    list_with_archived = client.get("/classes?include_archived=true", headers=headers)
    assert list_with_archived.status_code == 200
    archived_row = next(item for item in list_with_archived.json() if item["id"] == class_id)
    assert archived_row["is_archived"] is True

    blocked_session_resp = client.post(
        f"/classes/{class_id}/sessions",
        headers=headers,
        json={"session_date": "2026-03-02"},
    )
    assert blocked_session_resp.status_code == 409

    restore_resp = client.post(f"/classes/{class_id}/restore", headers=headers)
    assert restore_resp.status_code == 200
    assert restore_resp.json()["is_archived"] is False

    create_session_after_restore = client.post(
        f"/classes/{class_id}/sessions",
        headers=headers,
        json={"session_date": "2026-03-02"},
    )
    assert create_session_after_restore.status_code == 201


def test_exam_excel_flow_and_exports(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "2APIC-3"}, headers=headers)
    class_id = class_resp.json()["id"]
    roster = _build_roster_file([("STD010", "Nora"), ("STD011", "Youssef")])
    client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    exam_resp = client.post(
        f"/classes/{class_id}/exams",
        headers=headers,
        json={"title": "CC1", "exam_date": "2026-03-02", "max_score": 20, "weight": 1},
    )
    assert exam_resp.status_code == 201
    exam_id = exam_resp.json()["id"]

    template_resp = client.get(f"/exams/{exam_id}/template", headers=headers)
    assert template_resp.status_code == 200
    workbook = load_workbook(filename=BytesIO(template_resp.content), data_only=True)
    sheet = workbook.active
    assert sheet.max_row == 3

    results_file = _build_exam_file(
        [
            ("STD010", "Nora", 18, "A", "Good"),
            ("STD011", "Youssef", 16.5, "B+", "Can improve speed"),
        ]
    )
    import_resp = client.post(
        f"/exams/{exam_id}/results/import",
        headers=headers,
        files={"file": ("results.xlsx", results_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200
    assert import_resp.json()["imported"] == 2

    results_resp = client.get(f"/exams/{exam_id}/results", headers=headers)
    assert results_resp.status_code == 200
    assert len(results_resp.json()) == 2
    results_csv_resp = client.get(f"/exams/{exam_id}/results.csv", headers=headers)
    assert results_csv_resp.status_code == 200
    assert "text/csv" in results_csv_resp.headers["content-type"]
    assert "student_code,full_name,score,max_score,note,teacher_comment" in results_csv_resp.text
    assert "STD010" in results_csv_resp.text
    results_xlsx_resp = client.get(f"/exams/{exam_id}/results.xlsx", headers=headers)
    assert results_xlsx_resp.status_code == 200
    export_wb = load_workbook(filename=BytesIO(results_xlsx_resp.content), data_only=True)
    export_sheet = export_wb.active
    assert export_sheet.cell(1, 1).value == "student_code"
    assert export_sheet.cell(2, 1).value == "STD010"
    students = client.get(f"/classes/{class_id}/students", headers=headers).json()
    nora = next(student for student in students if student["student_code"] == "STD010")
    nora_profile = client.get(f"/classes/{class_id}/students/{nora['id']}/profile", headers=headers)
    assert nora_profile.status_code == 200
    assert nora_profile.json()["exams"]["count"] == 1
    assert float(nora_profile.json()["exams"]["results"][0]["score"]) == 18.0
    exam_summary_resp = client.get(f"/classes/{class_id}/exam-summary", headers=headers)
    assert exam_summary_resp.status_code == 200
    assert exam_summary_resp.json()["exams"][0]["average_score"] == 17.25
    exam_dashboard = client.get(f"/classes/{class_id}/dashboard", headers=headers)
    assert exam_dashboard.status_code == 200
    assert len(exam_dashboard.json()["exam_trend"]) >= 1

    export_excel_resp = client.get(f"/classes/{class_id}/reports/official-notes.xlsx", headers=headers)
    assert export_excel_resp.status_code == 200
    export_pdf_resp = client.get(f"/classes/{class_id}/reports/full.pdf", headers=headers)
    assert export_pdf_resp.status_code == 200
    assert export_pdf_resp.content.startswith(b"%PDF")


def test_create_linked_exam_workflow_unit_from_exam(client):
    headers = _unique_owner_headers(client)
    class_resp = client.post("/classes", json={"name": f"Exam WF {uuid.uuid4().hex[:6]}"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    exam_resp = client.post(
        f"/classes/{class_id}/exams",
        headers=headers,
        json={"title": "CC2", "exam_date": "2026-06-10", "max_score": 20, "weight": 1},
    )
    assert exam_resp.status_code == 201
    exam_id = int(exam_resp.json()["id"])

    linked_resp = client.post(
        f"/workflow/classes/{class_id}/exams/{exam_id}/linked-unit",
        headers=headers,
        json={"unit_type": "exam"},
    )
    assert linked_resp.status_code == 200
    body = linked_resp.json()
    assert body["created"] is True
    assert body["unit"]["unit_type"] == "exam"
    assert int(body["unit"]["exam_id"]) == exam_id
    assert body["unit"]["exam_title"] == "CC2"
    assert body["unit"]["checklist"]
    assert body["unit"]["checklist"][0]["title"] == "CC2"
    exam_child_titles = [row["title"] for row in body["unit"]["checklist"][0]["children"]]
    assert exam_child_titles == ["Supervision d'examen"]

    linked_again = client.post(
        f"/workflow/classes/{class_id}/exams/{exam_id}/linked-unit",
        headers=headers,
        json={"unit_type": "exam"},
    )
    assert linked_again.status_code == 200
    again_body = linked_again.json()
    assert again_body["created"] is False
    assert int(again_body["unit"]["id"]) == int(body["unit"]["id"])


def test_create_linked_exam_correction_workflow_reuses_exam_structure(client):
    headers = _unique_owner_headers(client)
    class_resp = client.post("/classes", json={"name": f"Exam Correction {uuid.uuid4().hex[:6]}"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    exam_resp = client.post(
        f"/classes/{class_id}/exams",
        headers=headers,
        json={"title": "CC3", "exam_date": "2026-06-17", "max_score": 20, "weight": 1},
    )
    assert exam_resp.status_code == 201
    exam_id = int(exam_resp.json()["id"])

    exam_unit_resp = client.post(
        f"/workflow/classes/{class_id}/exams/{exam_id}/linked-unit",
        headers=headers,
        json={"unit_type": "exam"},
    )
    assert exam_unit_resp.status_code == 200
    exam_unit = exam_unit_resp.json()["unit"]
    exam_unit_id = int(exam_unit["id"])

    close_resp = client.post(f"/workflow/classes/{class_id}/units/{exam_unit_id}/close", headers=headers)
    assert close_resp.status_code == 200

    correction_resp = client.post(
        f"/workflow/classes/{class_id}/exams/{exam_id}/linked-unit",
        headers=headers,
        json={"unit_type": "exam_correction"},
    )
    assert correction_resp.status_code == 200
    correction_body = correction_resp.json()
    assert correction_body["created"] is True
    assert correction_body["unit"]["unit_type"] == "exam_correction"
    assert int(correction_body["unit"]["exam_id"]) == exam_id
    correction_checklist = correction_body["unit"]["checklist"]
    assert correction_checklist
    assert correction_checklist[0]["title"] == "Correction - CC3"
    correction_child_titles = [row["title"] for row in correction_checklist[0]["children"]]
    assert correction_child_titles == ["Correction d'examen"]


def test_exam_list_includes_linked_workflow_status(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": f"Exam Link Status {uuid.uuid4().hex[:6]}"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    exam_resp = client.post(
        f"/classes/{class_id}/exams",
        headers=headers,
        json={"title": "CC4", "exam_date": "2026-06-24", "max_score": 20, "weight": 1},
    )
    assert exam_resp.status_code == 201
    exam_id = int(exam_resp.json()["id"])

    create_exam_workflow_resp = client.post(
        f"/workflow/classes/{class_id}/exams/{exam_id}/linked-unit",
        headers=headers,
        json={"unit_type": "exam"},
    )
    assert create_exam_workflow_resp.status_code == 200

    exams_resp = client.get(f"/classes/{class_id}/exams", headers=headers)
    assert exams_resp.status_code == 200
    row = next(item for item in exams_resp.json() if int(item["id"]) == exam_id)
    assert row["linked_exam_workflow_unit_id"] is not None
    assert row["linked_exam_workflow_status"] == "active"
    assert row["linked_exam_workflow_title"] == "CC4"
    assert row["linked_correction_workflow_unit_id"] is None


def test_create_linked_exam_workflow_reopens_closed_linked_unit(client):
    headers = _unique_owner_headers(client)
    class_resp = client.post("/classes", json={"name": f"Exam Reopen {uuid.uuid4().hex[:6]}"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    exam_resp = client.post(
        f"/classes/{class_id}/exams",
        headers=headers,
        json={"title": "CC5", "exam_date": "2026-07-01", "max_score": 20, "weight": 1},
    )
    assert exam_resp.status_code == 201
    exam_id = int(exam_resp.json()["id"])

    first_link = client.post(
        f"/workflow/classes/{class_id}/exams/{exam_id}/linked-unit",
        headers=headers,
        json={"unit_type": "exam"},
    )
    assert first_link.status_code == 200
    first_body = first_link.json()
    unit_id = int(first_body["unit"]["id"])

    close_resp = client.post(f"/workflow/classes/{class_id}/units/{unit_id}/close", headers=headers)
    assert close_resp.status_code == 200

    reopen_resp = client.post(
        f"/workflow/classes/{class_id}/exams/{exam_id}/linked-unit",
        headers=headers,
        json={"unit_type": "exam"},
    )
    assert reopen_resp.status_code == 200
    reopen_body = reopen_resp.json()
    assert reopen_body["created"] is False
    assert reopen_body["reopened"] is True
    assert int(reopen_body["unit"]["id"]) == unit_id
    assert reopen_body["unit"]["status"] == "active"

    exams_resp = client.get(f"/classes/{class_id}/exams", headers=headers)
    assert exams_resp.status_code == 200
    exam_row = next(item for item in exams_resp.json() if int(item["id"]) == exam_id)
    assert int(exam_row["linked_exam_workflow_unit_id"]) == unit_id
    assert exam_row["linked_exam_workflow_status"] == "active"


def test_linked_exam_workflow_exposes_exam_results_summary(client):
    headers = _unique_owner_headers(client)
    class_resp = client.post("/classes", json={"name": f"Exam Results Link {uuid.uuid4().hex[:6]}"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    roster = _build_roster_file([("STD400", "Rita"), ("STD401", "Amine"), ("STD402", "Salma")])
    roster_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert roster_resp.status_code == 200

    exam_resp = client.post(
        f"/classes/{class_id}/exams",
        headers=headers,
        json={"title": "CC6", "exam_date": "2026-07-08", "max_score": 20, "weight": 1},
    )
    assert exam_resp.status_code == 201
    exam_id = int(exam_resp.json()["id"])

    results_file = _build_exam_file(
        [
            ("STD400", "Rita", 18, "A", ""),
            ("STD401", "Amine", 11, "B", ""),
            ("STD402", "Salma", 7, "C", ""),
        ]
    )
    import_resp = client.post(
        f"/exams/{exam_id}/results/import",
        headers=headers,
        files={"file": ("results.xlsx", results_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200
    assert import_resp.json()["imported"] == 3

    linked_resp = client.post(
        f"/workflow/classes/{class_id}/exams/{exam_id}/linked-unit",
        headers=headers,
        json={"unit_type": "exam"},
    )
    assert linked_resp.status_code == 200
    unit = linked_resp.json()["unit"]
    assert unit["exam_results_count"] == 3
    assert float(unit["exam_results_average_score"]) == 12.0
    assert unit["exam_results_passed_count"] == 2


def test_linked_exam_workflow_uses_exam_paper_outline(client):
    headers = _unique_owner_headers(client)
    class_resp = client.post("/classes", json={"name": f"Exam Outline {uuid.uuid4().hex[:6]}"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    outline = "\n".join([
        "Exercice 1",
        "  a) Calcul direct",
        "  b) Justifier la methode",
        "Exercice 2",
        "Probleme",
    ])
    exam_resp = client.post(
        f"/classes/{class_id}/exams",
        headers=headers,
        json={
            "title": "CC7",
            "exam_date": "2026-07-15",
            "max_score": 20,
            "weight": 1,
            "paper_outline_text": outline,
        },
    )
    assert exam_resp.status_code == 201
    exam_id = int(exam_resp.json()["id"])

    linked_resp = client.post(
        f"/workflow/classes/{class_id}/exams/{exam_id}/linked-unit",
        headers=headers,
        json={"unit_type": "exam"},
    )
    assert linked_resp.status_code == 200
    unit = linked_resp.json()["unit"]
    assert unit["checklist"]
    root = unit["checklist"][0]
    assert root["title"] == "CC7"
    child_titles = [row["title"] for row in root["children"]]
    assert child_titles == ["Supervision d'examen"]


def test_linked_exam_correction_workflow_uses_exam_outline_when_no_exam_workflow_exists(client):
    headers = _unique_owner_headers(client)
    class_resp = client.post("/classes", json={"name": f"Exam Correction Outline {uuid.uuid4().hex[:6]}"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    outline = "\n".join([
        "Exercice 1",
        "Exercice 2",
    ])
    exam_resp = client.post(
        f"/classes/{class_id}/exams",
        headers=headers,
        json={
            "title": "CC8",
            "exam_date": "2026-07-22",
            "max_score": 20,
            "weight": 1,
            "paper_outline_text": outline,
        },
    )
    assert exam_resp.status_code == 201
    exam_id = int(exam_resp.json()["id"])

    correction_resp = client.post(
        f"/workflow/classes/{class_id}/exams/{exam_id}/linked-unit",
        headers=headers,
        json={"unit_type": "exam_correction"},
    )
    assert correction_resp.status_code == 200
    unit = correction_resp.json()["unit"]
    root = unit["checklist"][0]
    child_titles = [row["title"] for row in root["children"]]
    assert child_titles == ["Correction d'examen"]


def test_linked_exam_workflow_checklist_item_accepts_image_attachment(client):
    headers = _unique_owner_headers(client)
    class_resp = client.post("/classes", json={"name": f"Exam Attachment {uuid.uuid4().hex[:6]}"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    exam_resp = client.post(
        f"/classes/{class_id}/exams",
        headers=headers,
        json={
            "title": "CC9",
            "exam_date": "2026-07-29",
            "max_score": 20,
            "weight": 1,
            "paper_outline_text": "Exercice 1\nExercice 2",
        },
    )
    assert exam_resp.status_code == 201
    exam_id = int(exam_resp.json()["id"])

    linked_resp = client.post(
        f"/workflow/classes/{class_id}/exams/{exam_id}/linked-unit",
        headers=headers,
        json={"unit_type": "exam"},
    )
    assert linked_resp.status_code == 200
    unit = linked_resp.json()["unit"]
    supervision_row = next(row for row in unit["checklist"][0]["children"] if row["title"] == "Supervision d'examen")

    upload_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items/{supervision_row['id']}/attachments",
        headers=headers,
        files={"file": ("exercise-1.png", _build_tiny_png(), "image/png")},
    )
    assert upload_resp.status_code == 201
    attachment = upload_resp.json()
    assert attachment["item_id"] == supervision_row["id"]
    assert attachment["file_content_type"] == "image/png"

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_resp.status_code == 200
    active_unit = workspace_resp.json()["active_unit"]
    supervision_row_after = next(
        row for row in active_unit["checklist"][0]["children"] if row["title"] == "Supervision d'examen"
    )
    assert len(supervision_row_after["attachments"]) == 1
    assert supervision_row_after["attachments"][0]["id"] == attachment["id"]

    download_resp = client.get(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items/{supervision_row['id']}/attachments/{attachment['id']}",
        headers=headers,
    )
    assert download_resp.status_code == 200
    assert download_resp.headers["content-type"].startswith("image/png")

    delete_resp = client.delete(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items/{supervision_row['id']}/attachments/{attachment['id']}",
        headers=headers,
    )
    assert delete_resp.status_code == 200
    assert int(delete_resp.json()["deleted_attachment_id"]) == int(attachment["id"])

    workspace_after_delete = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_after_delete.status_code == 200
    active_unit_after_delete = workspace_after_delete.json()["active_unit"]
    supervision_row_after_delete = next(
        row for row in active_unit_after_delete["checklist"][0]["children"] if row["title"] == "Supervision d'examen"
    )
    assert supervision_row_after_delete["attachments"] == []


def test_exercise_series_checklist_item_accepts_image_attachment(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": f"Exercise Attachment {uuid.uuid4().hex[:6]}"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "exercise_series", "title": "Serie d'exercices", "source_text": "Exercice 1"},
    )
    assert unit_resp.status_code == 201
    unit = unit_resp.json()

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_resp.status_code == 200
    active_unit = workspace_resp.json()["active_unit"]
    root_id = int(active_unit["checklist"][0]["id"])

    add_item_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items",
        headers=headers,
        json={"title": "Exercice 1", "item_kind": "exercise", "parent_item_id": root_id},
    )
    assert add_item_resp.status_code == 201
    exercise_1_id = int(add_item_resp.json()["id"])

    upload_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items/{exercise_1_id}/attachments",
        headers=headers,
        files={"file": ("exercise-1.png", _build_tiny_png(), "image/png")},
    )
    assert upload_resp.status_code == 201
    attachment = upload_resp.json()
    assert attachment["item_id"] == exercise_1_id
    assert attachment["file_content_type"] == "image/png"

    workspace_after_upload = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_after_upload.status_code == 200
    exercise_1_after = next(
        row for row in workspace_after_upload.json()["active_unit"]["checklist"][0]["children"] if row["title"] == "Exercice 1"
    )
    assert len(exercise_1_after["attachments"]) == 1
    assert exercise_1_after["attachments"][0]["id"] == attachment["id"]

    download_resp = client.get(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items/{exercise_1_id}/attachments/{attachment['id']}",
        headers=headers,
    )
    assert download_resp.status_code == 200
    assert download_resp.headers["content-type"].startswith("image/png")

    delete_resp = client.delete(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items/{exercise_1_id}/attachments/{attachment['id']}",
        headers=headers,
    )
    assert delete_resp.status_code == 200
    assert int(delete_resp.json()["deleted_attachment_id"]) == int(attachment["id"])

    workspace_after_delete = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_after_delete.status_code == 200
    exercise_1_after_delete = next(
        row for row in workspace_after_delete.json()["active_unit"]["checklist"][0]["children"] if row["title"] == "Exercice 1"
    )
    assert exercise_1_after_delete["attachments"] == []


def test_chapter_checklist_item_accepts_image_attachment(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": f"Chapter Attachment {uuid.uuid4().hex[:6]}"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "chapter", "title": "Chapitre 1", "source_text": "I- Multiplication"},
    )
    assert unit_resp.status_code == 201
    unit = unit_resp.json()

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_resp.status_code == 200
    active_unit = workspace_resp.json()["active_unit"]
    root_id = int(active_unit["checklist"][0]["id"])

    add_item_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items",
        headers=headers,
        json={"title": "I- Multiplication", "item_kind": "section", "parent_item_id": root_id},
    )
    assert add_item_resp.status_code == 201
    section_id = int(add_item_resp.json()["id"])

    upload_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items/{section_id}/attachments",
        headers=headers,
        files={"file": ("section.png", _build_tiny_png(), "image/png")},
    )
    assert upload_resp.status_code == 201
    attachment = upload_resp.json()
    assert attachment["item_id"] == section_id
    assert attachment["file_content_type"] == "image/png"

    workspace_after_upload = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_after_upload.status_code == 200
    section_after = next(
        row for row in workspace_after_upload.json()["active_unit"]["checklist"][0]["children"] if row["title"] == "I- Multiplication"
    )
    assert len(section_after["attachments"]) == 1
    assert section_after["attachments"][0]["id"] == attachment["id"]

    download_resp = client.get(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items/{section_id}/attachments/{attachment['id']}",
        headers=headers,
    )
    assert download_resp.status_code == 200
    assert download_resp.headers["content-type"].startswith("image/png")

    delete_resp = client.delete(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items/{section_id}/attachments/{attachment['id']}",
        headers=headers,
    )
    assert delete_resp.status_code == 200
    assert int(delete_resp.json()["deleted_attachment_id"]) == int(attachment["id"])

    workspace_after_delete = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_after_delete.status_code == 200
    section_after_delete = next(
        row for row in workspace_after_delete.json()["active_unit"]["checklist"][0]["children"] if row["title"] == "I- Multiplication"
    )
    assert section_after_delete["attachments"] == []


def test_checklist_item_teacher_note_persists_and_updates(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": f"Checklist Note {uuid.uuid4().hex[:6]}"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "chapter", "title": "Chapitre note", "source_text": "I- Fraction"},
    )
    assert unit_resp.status_code == 201
    unit = unit_resp.json()

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_resp.status_code == 200
    root_id = int(workspace_resp.json()["active_unit"]["checklist"][0]["id"])

    add_item_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items",
        headers=headers,
        json={
            "title": "I- Fraction",
            "item_kind": "section",
            "parent_item_id": root_id,
            "teacher_note": "Montrer l'idee avec un exemple simple.",
        },
    )
    assert add_item_resp.status_code == 201
    item = add_item_resp.json()
    assert item["teacher_note"] == "Montrer l'idee avec un exemple simple."

    update_resp = client.put(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items/{item['id']}",
        headers=headers,
        json={"teacher_note": "Insister sur la simplification finale."},
    )
    assert update_resp.status_code == 200
    assert update_resp.json()["teacher_note"] == "Insister sur la simplification finale."

    workspace_after = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_after.status_code == 200
    child_row = next(row for row in workspace_after.json()["active_unit"]["checklist"][0]["children"] if row["title"] == "I- Fraction")
    assert child_row["teacher_note"] == "Insister sur la simplification finale."

    clear_resp = client.put(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items/{item['id']}",
        headers=headers,
        json={"teacher_note": ""},
    )
    assert clear_resp.status_code == 200
    assert clear_resp.json()["teacher_note"] is None


def test_unit_assistant_artifact_can_target_checklist_row(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": f"Artifact Row {uuid.uuid4().hex[:6]}"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = int(class_resp.json()["id"])

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "chapter", "title": "Chapitre guidance", "source_text": "I- Produit"},
    )
    assert unit_resp.status_code == 201
    unit = unit_resp.json()

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_resp.status_code == 200
    root_id = int(workspace_resp.json()["active_unit"]["checklist"][0]["id"])

    add_item_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit['id']}/items",
        headers=headers,
        json={"title": "I- Produit", "item_kind": "section", "parent_item_id": root_id},
    )
    assert add_item_resp.status_code == 201
    item_id = int(add_item_resp.json()["id"])

    save_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit['id']}/assistant/artifacts",
        headers=headers,
        json={
            "artifact_kind": "teacher_notes",
            "checklist_item_id": item_id,
            "provider": "notebooklm",
            "section_title": "I- Produit",
            "section_path": ["Chapitre guidance", "I- Produit"],
            "action": "explain_section",
            "title": "Guidance produit",
            "answer_rows": ["Expliquer le sens du produit."],
        },
    )
    assert save_resp.status_code == 200
    artifact = save_resp.json()
    assert int(artifact["checklist_item_id"]) == item_id

    list_resp = client.get(
        f"/workflow/classes/{class_id}/units/{unit['id']}/assistant/artifacts",
        headers=headers,
    )
    assert list_resp.status_code == 200
    rows = list_resp.json()
    saved = next(row for row in rows if int(row["id"]) == int(artifact["id"]))
    assert int(saved["checklist_item_id"]) == item_id


def test_import_students_from_notescc_list_format(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "NotesCC List Class"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    roster = _build_notescc_list_file(
        [
            ("10934196", "A161027646", "Ismail", "23-03-2011"),
            ("10851611", "A166022053", "Adam", "28-08-2010"),
        ]
    )
    import_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("list_notescc.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200
    assert import_resp.json()["created"] == 2

    students = client.get(f"/classes/{class_id}/students", headers=headers).json()
    assert len(students) == 2
    ismail = next(row for row in students if row["student_code"] == "A161027646")
    assert ismail["external_id"] == "10934196"
    assert ismail["birth_date"] == "2011-03-23"


def test_exam_import_from_id_name_birthdate_notes_layout(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Exam Notes Layout"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    roster = _build_notescc_list_file(
        [
            ("2001", "STD2001", "Sara", "14-08-2012"),
            ("2002", "STD2002", "Omar", "01-03-2012"),
        ]
    )
    import_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students_layout.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200

    exam_resp = client.post(
        f"/classes/{class_id}/exams",
        headers=headers,
        json={"title": "CC Layout", "exam_date": "2026-03-09", "max_score": 20, "weight": 1},
    )
    assert exam_resp.status_code == 201
    exam_id = exam_resp.json()["id"]

    results_file = _build_exam_list_file(
        [
            ("STD2001", "Sara", "2012-08-14", 12, 14, 16, 15.5),
            ("STD2002", "Omar", "2012-03-01", 10, 11, 13, 12.0),
        ]
    )
    import_exam_resp = client.post(
        f"/exams/{exam_id}/results/import",
        headers=headers,
        files={"file": ("results_layout.xlsx", results_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_exam_resp.status_code == 200
    assert import_exam_resp.json()["imported"] == 2

    results_resp = client.get(f"/exams/{exam_id}/results", headers=headers)
    assert results_resp.status_code == 200
    result_map = {row["student_code"]: float(row["score"]) for row in results_resp.json()}
    assert result_map["STD2001"] == 15.5
    assert result_map["STD2002"] == 12.0


def test_exam_update_archive_restore(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Exam Archive Class"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    roster = _build_roster_file([("STD300", "Rita"), ("STD301", "Amine")])
    roster_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert roster_resp.status_code == 200

    exam_resp = client.post(
        f"/classes/{class_id}/exams",
        headers=headers,
        json={"title": "CC-Archive", "exam_date": "2026-03-05", "max_score": 20, "weight": 1},
    )
    assert exam_resp.status_code == 201
    exam_id = exam_resp.json()["id"]

    update_resp = client.put(
        f"/exams/{exam_id}",
        headers=headers,
        json={"title": "CC-Archive-Updated", "max_score": 25, "weight": 2},
    )
    assert update_resp.status_code == 200
    assert update_resp.json()["title"] == "CC-Archive-Updated"
    assert float(update_resp.json()["max_score"]) == 25.0
    assert float(update_resp.json()["weight"]) == 2.0

    archive_resp = client.post(f"/exams/{exam_id}/archive", headers=headers)
    assert archive_resp.status_code == 200
    assert archive_resp.json()["is_archived"] is True

    list_default = client.get(f"/classes/{class_id}/exams", headers=headers)
    assert list_default.status_code == 200
    assert all(exam["id"] != exam_id for exam in list_default.json())

    list_all = client.get(f"/classes/{class_id}/exams?include_archived=true", headers=headers)
    assert list_all.status_code == 200
    archived_row = next(exam for exam in list_all.json() if exam["id"] == exam_id)
    assert archived_row["is_archived"] is True

    results_file = _build_exam_file(
        [
            ("STD300", "Rita", 18, "A", ""),
            ("STD301", "Amine", 19, "A+", ""),
        ]
    )
    import_blocked_resp = client.post(
        f"/exams/{exam_id}/results/import",
        headers=headers,
        files={"file": ("results.xlsx", results_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_blocked_resp.status_code == 409

    restore_resp = client.post(f"/exams/{exam_id}/restore", headers=headers)
    assert restore_resp.status_code == 200
    assert restore_resp.json()["is_archived"] is False

    import_ok_resp = client.post(
        f"/exams/{exam_id}/results/import",
        headers=headers,
        files={"file": ("results.xlsx", results_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_ok_resp.status_code == 200
    assert import_ok_resp.json()["imported"] == 2


def test_import_notescc_exam_format_and_template_export(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "2APIC-3", "subject": "Math"}, headers=headers)
    class_id = class_resp.json()["id"]
    roster = _build_roster_file([("STD100", "Amina"), ("STD101", "Salim")])
    client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    exam_resp = client.post(
        f"/classes/{class_id}/exams",
        headers=headers,
        json={"title": "CC1", "exam_date": "2026-03-03", "max_score": 20, "weight": 1},
    )
    exam_id = exam_resp.json()["id"]

    notescc_file = _build_notescc_exam_file([("STD100", "Amina", 14), ("STD101", "Salim", 17.5)])
    import_resp = client.post(
        f"/exams/{exam_id}/results/import",
        headers=headers,
        files={"file": ("notescc.xlsx", notescc_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200
    assert import_resp.json()["imported"] == 2

    template_wb = Workbook()
    template_sheet = template_wb.active
    template_sheet.title = "NotesCC"
    with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        template_wb.save(tmp.name)
        template_path = tmp.name
    os.environ["PRINCIPAL_EXPORT_TEMPLATE"] = template_path
    try:
        notescc_template_resp = client.get(f"/exams/{exam_id}/template?format=notescc", headers=headers)
        assert notescc_template_resp.status_code == 200
        notescc_wb = load_workbook(filename=BytesIO(notescc_template_resp.content), data_only=True)
        assert "NotesCC" in notescc_wb.sheetnames

        export_resp = client.get(f"/classes/{class_id}/reports/official-notes.xlsx", headers=headers)
        assert export_resp.status_code == 200
        out_wb = load_workbook(filename=BytesIO(export_resp.content), data_only=True)
        out_sheet = out_wb["NotesCC"]
        assert out_sheet["I9"].value == "2APIC-3"
        assert out_sheet["O11"].value == "Math"
        # Row ordering is by student full name asc -> Amina first.
        assert out_sheet.cell(18, 3).value == "STD100"
        assert float(out_sheet.cell(18, 7).value) == 14.0
    finally:
        os.environ.pop("PRINCIPAL_EXPORT_TEMPLATE", None)
        try:
            os.remove(template_path)
        except OSError:
            pass


def test_export_history_and_audit_logs(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Audit Export Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    roster = _build_roster_file([("STD200", "Ilyas"), ("STD201", "Nisrine")])
    import_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200

    exam_resp = client.post(
        f"/classes/{class_id}/exams",
        headers=headers,
        json={"title": "CC-Audit", "exam_date": "2026-03-04", "max_score": 20, "weight": 1},
    )
    assert exam_resp.status_code == 201
    exam_id = exam_resp.json()["id"]

    results_file = _build_exam_file(
        [
            ("STD200", "Ilyas", 13.5, "C+", ""),
            ("STD201", "Nisrine", 17.0, "A-", ""),
        ]
    )
    import_exam_resp = client.post(
        f"/exams/{exam_id}/results/import",
        headers=headers,
        files={"file": ("results.xlsx", results_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_exam_resp.status_code == 200

    pdf_export = client.get(f"/classes/{class_id}/reports/full.pdf", headers=headers)
    assert pdf_export.status_code == 200
    assert pdf_export.content.startswith(b"%PDF")
    notes_export = client.get(f"/classes/{class_id}/reports/official-notes.xlsx", headers=headers)
    assert notes_export.status_code == 200

    history_resp = client.get(f"/classes/{class_id}/exports/history", headers=headers)
    assert history_resp.status_code == 200
    history = history_resp.json()["items"]
    assert len(history) >= 2
    first_id = history[0]["id"]
    assert any(item["export_type"] == "class_full_pdf" for item in history)
    assert any(item["export_type"] == "official_notes_xlsx" for item in history)

    download_resp = client.get(f"/exports/{first_id}/download", headers=headers)
    assert download_resp.status_code == 200

    class_audit_resp = client.get(f"/classes/{class_id}/audit-logs", headers=headers)
    assert class_audit_resp.status_code == 200
    actions = [item["action"] for item in class_audit_resp.json()["items"]]
    assert "report.export" in actions
    assert "students.import" in actions

    owner_audit_resp = client.get("/audit/logs?action=report.export", headers=headers)
    assert owner_audit_resp.status_code == 200
    assert owner_audit_resp.json()["count"] >= 1


def test_masked_export_privacy_mode(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Privacy Class", "subject": "Math"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    roster = _build_roster_file([("STD500", "Alice Privacy"), ("STD501", "Bob Privacy")])
    import_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_resp.status_code == 200

    students = client.get(f"/classes/{class_id}/students", headers=headers).json()
    session_resp = client.post(
        f"/classes/{class_id}/sessions",
        headers=headers,
        json={"session_date": "2026-03-08", "note": "privacy test"},
    )
    assert session_resp.status_code == 201
    session_id = session_resp.json()["id"]

    attendance_payload = [
        {"student_id": students[0]["id"], "status": "present", "minutes_late": 0, "comment": "present"},
        {"student_id": students[1]["id"], "status": "absent", "minutes_late": 0, "comment": "absent"},
    ]
    attendance_resp = client.put(f"/sessions/{session_id}/attendance", json=attendance_payload, headers=headers)
    assert attendance_resp.status_code == 200

    masked_csv = client.get(f"/classes/{class_id}/attendance-export.csv?mask_personal_data=true", headers=headers)
    assert masked_csv.status_code == 200
    assert "Alice Privacy" not in masked_csv.text
    assert "Bob Privacy" not in masked_csv.text
    assert "STD500" not in masked_csv.text
    assert "STD501" not in masked_csv.text
    assert "ANON001" in masked_csv.text or "ANON002" in masked_csv.text

    exam_resp = client.post(
        f"/classes/{class_id}/exams",
        headers=headers,
        json={"title": "CC-Privacy", "exam_date": "2026-03-08", "max_score": 20, "weight": 1},
    )
    assert exam_resp.status_code == 201
    exam_id = exam_resp.json()["id"]
    results_file = _build_exam_file(
        [
            ("STD500", "Alice Privacy", 14, "B", ""),
            ("STD501", "Bob Privacy", 16, "A", ""),
        ]
    )
    import_exam_resp = client.post(
        f"/exams/{exam_id}/results/import",
        headers=headers,
        files={"file": ("results.xlsx", results_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert import_exam_resp.status_code == 200

    masked_notes = client.get(f"/classes/{class_id}/reports/official-notes.xlsx?mask_personal_data=true", headers=headers)
    assert masked_notes.status_code == 200
    wb = load_workbook(filename=BytesIO(masked_notes.content), data_only=True)
    if "NotesCC" in wb.sheetnames:
        sheet = wb["NotesCC"]
        assert str(sheet.cell(18, 3).value).startswith("ANON")
        assert str(sheet.cell(18, 4).value).startswith("Student")
    else:
        sheet = wb.active
        assert str(sheet.cell(3, 1).value).startswith("ANON")
        assert str(sheet.cell(3, 2).value).startswith("Student")

    history_resp = client.get(f"/classes/{class_id}/exports/history", headers=headers)
    assert history_resp.status_code == 200
    export_types = {item["export_type"] for item in history_resp.json()["items"]}
    assert "official_notes_xlsx_masked" in export_types


def test_upload_validation_rejects_invalid_file_types(client):
    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Upload Validation"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    invalid_roster_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.txt", b"not-an-excel", "text/plain")},
    )
    assert invalid_roster_resp.status_code == 400

    roster = _build_roster_file([("STD900", "Sara")])
    roster_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert roster_resp.status_code == 200

    session_resp = client.post(
        f"/classes/{class_id}/sessions",
        headers=headers,
        json={"session_date": "2026-03-06"},
    )
    assert session_resp.status_code == 201
    session_id = session_resp.json()["id"]

    invalid_image_resp = client.post(
        f"/sessions/{session_id}/uploads",
        headers=headers,
        files={"file": ("board.txt", b"plain text", "text/plain")},
        data={"raw_text": "1 Algebra"},
    )
    assert invalid_image_resp.status_code == 400

    exam_resp = client.post(
        f"/classes/{class_id}/exams",
        headers=headers,
        json={"title": "CC-Upload", "exam_date": "2026-03-06", "max_score": 20, "weight": 1},
    )
    assert exam_resp.status_code == 201
    exam_id = exam_resp.json()["id"]

    invalid_exam_import_resp = client.post(
        f"/exams/{exam_id}/results/import",
        headers=headers,
        files={"file": ("results.csv", b"code,score", "text/csv")},
    )
    assert invalid_exam_import_resp.status_code == 400


def test_owner_audit_csv_and_teacher_forbidden(client):
    owner_headers = _auth_headers(client)
    _ = client.post("/classes", json={"name": "Owner Audit Class"}, headers=owner_headers)

    owner_csv = client.get("/audit/logs.csv?limit=20", headers=owner_headers)
    assert owner_csv.status_code == 200
    assert "text/csv" in owner_csv.headers["content-type"]
    assert "id,created_at,user_id,action,entity_type,entity_id,class_id,details" in owner_csv.text

    owner_future_json = client.get("/audit/logs?date_from=2100-01-01&limit=20", headers=owner_headers)
    assert owner_future_json.status_code == 200
    assert owner_future_json.json()["count"] == 0

    owner_future_csv = client.get("/audit/logs.csv?date_from=2100-01-01&limit=20", headers=owner_headers)
    assert owner_future_csv.status_code == 200
    csv_lines = [line for line in owner_future_csv.text.strip().splitlines() if line.strip()]
    assert len(csv_lines) == 1
    assert csv_lines[0].startswith("id,created_at,user_id,action,entity_type,entity_id,class_id,details")

    _, teacher_headers = _create_teacher_and_login(client, owner_headers)
    teacher_json_forbidden = client.get("/audit/logs", headers=teacher_headers)
    assert teacher_json_forbidden.status_code == 403
    teacher_csv_forbidden = client.get("/audit/logs.csv", headers=teacher_headers)
    assert teacher_csv_forbidden.status_code == 403


def test_owner_ops_status_and_teacher_forbidden(client):
    owner_headers = _auth_headers(client)
    owner_resp = client.get("/ops/status", headers=owner_headers)
    assert owner_resp.status_code == 200
    payload = owner_resp.json()
    assert payload["status"] == "ok"
    assert "uptime_seconds" in payload
    assert "storage" in payload
    assert "backups" in payload
    assert "alerts" in payload
    assert "enabled" in payload["alerts"]

    _, teacher_headers = _create_teacher_and_login(client, owner_headers)
    teacher_resp = client.get("/ops/status", headers=teacher_headers)
    assert teacher_resp.status_code == 403


def test_rate_limiting_upload_and_export(client):
    from app import config as app_config
    from app.services.rate_limit import reset_rate_limits

    headers = _auth_headers(client)
    class_resp = client.post("/classes", json={"name": "Rate Limit Class"}, headers=headers)
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    roster = _build_roster_file([("STD950", "Nawal")])
    roster_resp = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={"file": ("students.xlsx", roster, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert roster_resp.status_code == 200

    session_resp = client.post(
        f"/classes/{class_id}/sessions",
        headers=headers,
        json={"session_date": "2026-03-07"},
    )
    assert session_resp.status_code == 201
    session_id = session_resp.json()["id"]

    old_upload_count = app_config.UPLOAD_RATE_LIMIT_COUNT
    old_upload_window = app_config.UPLOAD_RATE_LIMIT_WINDOW_SECONDS
    old_export_count = app_config.EXPORT_RATE_LIMIT_COUNT
    old_export_window = app_config.EXPORT_RATE_LIMIT_WINDOW_SECONDS
    try:
        app_config.UPLOAD_RATE_LIMIT_COUNT = 2
        app_config.UPLOAD_RATE_LIMIT_WINDOW_SECONDS = 60
        reset_rate_limits()

        upload1 = client.post(
            f"/sessions/{session_id}/uploads",
            headers=headers,
            files={"file": ("board.jpg", b"fake-image", "image/jpeg")},
            data={"raw_text": "1 Algebra"},
        )
        assert upload1.status_code == 201
        upload2 = client.post(
            f"/sessions/{session_id}/uploads",
            headers=headers,
            files={"file": ("board.jpg", b"fake-image", "image/jpeg")},
            data={"raw_text": "1 Algebra"},
        )
        assert upload2.status_code == 201
        upload3 = client.post(
            f"/sessions/{session_id}/uploads",
            headers=headers,
            files={"file": ("board.jpg", b"fake-image", "image/jpeg")},
            data={"raw_text": "1 Algebra"},
        )
        assert upload3.status_code == 429

        app_config.EXPORT_RATE_LIMIT_COUNT = 1
        app_config.EXPORT_RATE_LIMIT_WINDOW_SECONDS = 60
        reset_rate_limits()
        export1 = client.get(f"/classes/{class_id}/attendance-export.csv", headers=headers)
        assert export1.status_code == 200
        export2 = client.get(f"/classes/{class_id}/attendance-export.csv", headers=headers)
        assert export2.status_code == 429
    finally:
        app_config.UPLOAD_RATE_LIMIT_COUNT = old_upload_count
        app_config.UPLOAD_RATE_LIMIT_WINDOW_SECONDS = old_upload_window
        app_config.EXPORT_RATE_LIMIT_COUNT = old_export_count
        app_config.EXPORT_RATE_LIMIT_WINDOW_SECONDS = old_export_window
        reset_rate_limits()


def test_requires_authentication(client):
    unauthorized = client.get("/classes")
    assert unauthorized.status_code == 401


def test_auth_refresh_rotates_token(client):
    old_headers = _auth_headers(client)
    old_token = old_headers["Authorization"].split(" ", maxsplit=1)[1]
    old_headers = {"Authorization": f"Bearer {old_token}"}

    refresh_resp = client.post("/auth/refresh", headers=old_headers)
    assert refresh_resp.status_code == 200
    new_token = refresh_resp.json()["access_token"]
    assert new_token != old_token

    old_me = client.get("/auth/me", headers=old_headers)
    assert old_me.status_code == 401
    new_me = client.get("/auth/me", headers={"Authorization": f"Bearer {new_token}"})
    assert new_me.status_code == 200


def test_login_lockout_and_owner_unlock(client):
    from app import config as app_config

    owner_headers = _auth_headers(client)
    email = f"locked_{uuid.uuid4().hex[:8]}@app.local"
    password = "TeacherPass123"
    create_resp = client.post(
        "/auth/users",
        headers=owner_headers,
        json={"email": email, "password": password, "full_name": "Lockout Teacher", "role": "teacher"},
    )
    assert create_resp.status_code == 201
    teacher_id = create_resp.json()["id"]

    old_max = app_config.MAX_FAILED_LOGIN_ATTEMPTS
    old_minutes = app_config.LOGIN_LOCKOUT_MINUTES
    try:
        app_config.MAX_FAILED_LOGIN_ATTEMPTS = 2
        app_config.LOGIN_LOCKOUT_MINUTES = 5

        bad1 = client.post("/auth/login", json={"email": email, "password": "wrong-1"})
        assert bad1.status_code == 401
        bad2 = client.post("/auth/login", json={"email": email, "password": "wrong-2"})
        assert bad2.status_code == 401

        locked_login = client.post("/auth/login", json={"email": email, "password": password})
        assert locked_login.status_code == 423

        users_resp = client.get("/auth/users", headers=owner_headers)
        assert users_resp.status_code == 200
        locked_row = next(row for row in users_resp.json() if row["id"] == teacher_id)
        assert locked_row["locked_until"] is not None

        unlock_resp = client.post(f"/auth/users/{teacher_id}/unlock", headers=owner_headers)
        assert unlock_resp.status_code == 200
        assert unlock_resp.json()["locked_until"] is None
        assert int(unlock_resp.json()["failed_login_attempts"]) == 0

        ok_login = client.post("/auth/login", json={"email": email, "password": password})
        assert ok_login.status_code == 200
        assert "access_token" in ok_login.json()
    finally:
        app_config.MAX_FAILED_LOGIN_ATTEMPTS = old_max
        app_config.LOGIN_LOCKOUT_MINUTES = old_minutes


def test_teacher_class_isolation_and_assignment(client):
    owner_headers = _auth_headers(client)
    teacher_id, teacher_headers = _create_teacher_and_login(client, owner_headers)

    owner_class_resp = client.post("/classes", headers=owner_headers, json={"name": "Owner Class"})
    assert owner_class_resp.status_code == 201
    owner_class_id = owner_class_resp.json()["id"]

    teacher_list_before = client.get("/classes", headers=teacher_headers)
    assert teacher_list_before.status_code == 200
    assert all(row["id"] != owner_class_id for row in teacher_list_before.json())

    forbidden_get = client.get(f"/classes/{owner_class_id}", headers=teacher_headers)
    assert forbidden_get.status_code == 403

    assigned_class_resp = client.post(
        "/classes",
        headers=owner_headers,
        json={"name": "Assigned Class", "teacher_user_id": teacher_id},
    )
    assert assigned_class_resp.status_code == 201
    assigned_class_id = assigned_class_resp.json()["id"]

    teacher_list_after = client.get("/classes", headers=teacher_headers)
    assert teacher_list_after.status_code == 200
    ids_after = {row["id"] for row in teacher_list_after.json()}
    assert assigned_class_id in ids_after
    assert owner_class_id not in ids_after

    owner_view_teacher_classes = client.get(f"/classes/by-teacher/{teacher_id}", headers=owner_headers)
    assert owner_view_teacher_classes.status_code == 200
    owner_view_ids = {row["id"] for row in owner_view_teacher_classes.json()}
    assert assigned_class_id in owner_view_ids
    assert owner_class_id not in owner_view_ids

    owner_overview = client.get("/classes/owner-overview", headers=owner_headers)
    assert owner_overview.status_code == 200
    overview_payload = owner_overview.json()
    assert overview_payload["counts"]["teachers"] >= 1
    teacher_overview_rows = [row for row in overview_payload["teachers"] if row["teacher_id"] == teacher_id]
    assert teacher_overview_rows
    assert teacher_overview_rows[0]["assigned_classes"] >= 1

    teacher_forbidden_owner_view = client.get(f"/classes/by-teacher/{teacher_id}", headers=teacher_headers)
    assert teacher_forbidden_owner_view.status_code == 403
    teacher_forbidden_owner_overview = client.get("/classes/owner-overview", headers=teacher_headers)
    assert teacher_forbidden_owner_overview.status_code == 403

    assign_resp = client.post(f"/classes/{owner_class_id}/assign-teacher/{teacher_id}", headers=owner_headers)
    assert assign_resp.status_code == 204

    class_teachers_owner = client.get(f"/classes/{owner_class_id}/teachers", headers=owner_headers)
    assert class_teachers_owner.status_code == 200
    teacher_ids = {row["id"] for row in class_teachers_owner.json()}
    assert teacher_id in teacher_ids

    class_teachers_teacher_forbidden = client.get(f"/classes/{owner_class_id}/teachers", headers=teacher_headers)
    assert class_teachers_teacher_forbidden.status_code == 403

    teacher_get_after_assign = client.get(f"/classes/{owner_class_id}", headers=teacher_headers)
    assert teacher_get_after_assign.status_code == 200

    unassign_resp = client.delete(f"/classes/{owner_class_id}/assign-teacher/{teacher_id}", headers=owner_headers)
    assert unassign_resp.status_code == 204

    owner_view_after_unassign = client.get(f"/classes/by-teacher/{teacher_id}", headers=owner_headers)
    assert owner_view_after_unassign.status_code == 200
    owner_view_after_ids = {row["id"] for row in owner_view_after_unassign.json()}
    assert assigned_class_id in owner_view_after_ids
    assert owner_class_id not in owner_view_after_ids

    teacher_get_after_unassign = client.get(f"/classes/{owner_class_id}", headers=teacher_headers)
    assert teacher_get_after_unassign.status_code == 403


def test_class_assignment_replaces_previous_teacher_and_exposes_teacher_user_id(client):
    owner_headers = _auth_headers(client)
    teacher_one_id, teacher_one_headers = _create_teacher_and_login(client, owner_headers)
    teacher_two_id, teacher_two_headers = _create_teacher_and_login(client, owner_headers)

    create_resp = client.post(
        "/classes",
        headers=owner_headers,
        json={"name": "Replacement Class", "teacher_user_id": teacher_one_id},
    )
    assert create_resp.status_code == 201
    class_payload = create_resp.json()
    class_id = class_payload["id"]
    assert class_payload["teacher_user_id"] == teacher_one_id

    list_resp = client.get("/classes", headers=owner_headers)
    assert list_resp.status_code == 200
    listed_row = next(row for row in list_resp.json() if row["id"] == class_id)
    assert listed_row["teacher_user_id"] == teacher_one_id

    replace_resp = client.post(f"/classes/{class_id}/assign-teacher/{teacher_two_id}", headers=owner_headers)
    assert replace_resp.status_code == 204

    detail_resp = client.get(f"/classes/{class_id}", headers=owner_headers)
    assert detail_resp.status_code == 200
    assert detail_resp.json()["teacher_user_id"] == teacher_two_id

    teachers_resp = client.get(f"/classes/{class_id}/teachers", headers=owner_headers)
    assert teachers_resp.status_code == 200
    teacher_ids = {row["id"] for row in teachers_resp.json()}
    assert teacher_ids == {teacher_two_id}

    stale_access_resp = client.get(f"/classes/{class_id}", headers=teacher_one_headers)
    assert stale_access_resp.status_code == 403

    current_access_resp = client.get(f"/classes/{class_id}", headers=teacher_two_headers)
    assert current_access_resp.status_code == 200


def test_teacher_can_access_multiple_assigned_classes(client):
    owner_headers = _auth_headers(client)
    teacher_id, teacher_headers = _create_teacher_and_login(client, owner_headers)

    first_class_resp = client.post(
        "/classes",
        headers=owner_headers,
        json={"name": "Math 1", "teacher_user_id": teacher_id},
    )
    assert first_class_resp.status_code == 201
    first_class_id = first_class_resp.json()["id"]

    second_class_resp = client.post(
        "/classes",
        headers=owner_headers,
        json={"name": "Math 2", "teacher_user_id": teacher_id},
    )
    assert second_class_resp.status_code == 201
    second_class_id = second_class_resp.json()["id"]

    teacher_classes_resp = client.get("/classes", headers=teacher_headers)
    assert teacher_classes_resp.status_code == 200
    teacher_class_ids = {row["id"] for row in teacher_classes_resp.json()}
    assert first_class_id in teacher_class_ids
    assert second_class_id in teacher_class_ids

    first_access_resp = client.get(f"/classes/{first_class_id}", headers=teacher_headers)
    assert first_access_resp.status_code == 200

    second_access_resp = client.get(f"/classes/{second_class_id}", headers=teacher_headers)
    assert second_access_resp.status_code == 200


def test_owner_can_patch_class_teacher_assignment(client):
    owner_headers = _auth_headers(client)
    teacher_id, teacher_headers = _create_teacher_and_login(client, owner_headers)

    create_resp = client.post("/classes", headers=owner_headers, json={"name": "Patchable Class"})
    assert create_resp.status_code == 201
    class_id = create_resp.json()["id"]

    patch_assign_resp = client.patch(
        f"/classes/{class_id}",
        headers=owner_headers,
        json={"teacher_user_id": teacher_id},
    )
    assert patch_assign_resp.status_code == 200
    assert patch_assign_resp.json()["teacher_user_id"] == teacher_id

    teacher_access_resp = client.get(f"/classes/{class_id}", headers=teacher_headers)
    assert teacher_access_resp.status_code == 200

    patch_clear_resp = client.patch(
        f"/classes/{class_id}",
        headers=owner_headers,
        json={"teacher_user_id": None},
    )
    assert patch_clear_resp.status_code == 200
    assert patch_clear_resp.json()["teacher_user_id"] is None

    teacher_forbidden_resp = client.get(f"/classes/{class_id}", headers=teacher_headers)
    assert teacher_forbidden_resp.status_code == 403


def test_teacher_can_update_exam_result(client):
    headers = _unique_owner_headers(client)

    class_resp = client.post("/classes", headers=headers, json={"name": "Exam Class"})
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    student_import = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={
            "file": (
                "students.xlsx",
                _build_roster_file([("S1", "Student One")]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert student_import.status_code == 200

    exam_resp = client.post(
        f"/classes/{class_id}/exams",
        headers=headers,
        json={"title": "Exam 1", "exam_date": "2026-05-29", "max_score": 20},
    )
    assert exam_resp.status_code == 201
    exam_id = exam_resp.json()["id"]

    import_resp = client.post(
        f"/exams/{exam_id}/results/import",
        headers=headers,
        files={
            "file": (
                "results.xlsx",
                _build_exam_file([("S1", "Student One", 12, "Initial", "Needs review")]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_resp.status_code == 200

    results_resp = client.get(f"/exams/{exam_id}/results", headers=headers)
    assert results_resp.status_code == 200
    result_row = results_resp.json()[0]
    assert result_row["score"] == 12

    update_resp = client.put(
        f"/exams/{exam_id}/results/{result_row['student_id']}",
        headers=headers,
        json={"score": 15.5, "note": "Updated", "teacher_comment": "Better now"},
    )
    assert update_resp.status_code == 200
    updated = update_resp.json()
    assert updated["score"] == 15.5
    assert updated["note"] == "Updated"
    assert updated["teacher_comment"] == "Better now"
    assert updated["id"] is not None


def test_session_detail_exposes_workflow_unit_linkage(client):
    headers = _unique_owner_headers(client)

    class_resp = client.post("/classes", headers=headers, json={"name": "Workflow Session Class"})
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    unit_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "chapter", "title": "Linked Unit", "source_text": "I- Titre"},
    )
    assert unit_resp.status_code == 201
    unit_id = unit_resp.json()["id"]

    student_import = client.post(
        f"/classes/{class_id}/students/import",
        headers=headers,
        files={
            "file": (
                "students.xlsx",
                _build_roster_file([("S1", "Student One")]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert student_import.status_code == 200

    session_resp = client.post(
        f"/workflow/classes/{class_id}/sessions/start",
        headers=headers,
        json={"absent_student_ids": []},
    )
    assert session_resp.status_code == 201
    session_id = session_resp.json()["id"]

    detail_resp = client.get(f"/sessions/{session_id}", headers=headers)
    assert detail_resp.status_code == 200
    payload = detail_resp.json()
    assert payload["unit_id"] == unit_id
    assert payload["session"]["unit_id"] == unit_id
    assert payload["session"]["unit_session_number"] == 1


def test_owner_teacher_status_reset_and_password_change(client):
    owner_headers = _auth_headers(client)
    owner_me = client.get("/auth/me", headers=owner_headers)
    assert owner_me.status_code == 200
    owner_id = owner_me.json()["id"]

    deactivate_owner_resp = client.patch(
        f"/auth/users/{owner_id}/status",
        headers=owner_headers,
        json={"is_active": False},
    )
    assert deactivate_owner_resp.status_code == 400

    email = f"teacher_{uuid.uuid4().hex[:8]}@app.local"
    password = "TeacherPass123"
    create_resp = client.post(
        "/auth/users",
        headers=owner_headers,
        json={"email": email, "password": password, "full_name": "Teacher Status", "role": "teacher"},
    )
    assert create_resp.status_code == 201
    teacher_id = create_resp.json()["id"]

    teacher_headers = _login_headers(client, email, password)
    me_resp = client.get("/auth/me", headers=teacher_headers)
    assert me_resp.status_code == 200

    deactivate_resp = client.patch(
        f"/auth/users/{teacher_id}/status",
        headers=owner_headers,
        json={"is_active": False},
    )
    assert deactivate_resp.status_code == 200
    assert deactivate_resp.json()["is_active"] is False

    stale_token_resp = client.get("/auth/me", headers=teacher_headers)
    assert stale_token_resp.status_code == 401
    inactive_login_resp = client.post("/auth/login", json={"email": email, "password": password})
    assert inactive_login_resp.status_code == 403

    reactivate_resp = client.patch(
        f"/auth/users/{teacher_id}/status",
        headers=owner_headers,
        json={"is_active": True},
    )
    assert reactivate_resp.status_code == 200
    assert reactivate_resp.json()["is_active"] is True
    _ = _login_headers(client, email, password)

    reset_resp = client.post(
        f"/auth/users/{teacher_id}/reset-password",
        headers=owner_headers,
        json={"new_password": "TeacherPass456"},
    )
    assert reset_resp.status_code == 200
    old_login_after_reset = client.post("/auth/login", json={"email": email, "password": password})
    assert old_login_after_reset.status_code == 401

    teacher_headers_new = _login_headers(client, email, "TeacherPass456")
    change_password_resp = client.post(
        "/auth/change-password",
        headers=teacher_headers_new,
        json={"current_password": "TeacherPass456", "new_password": "TeacherPass789"},
    )
    assert change_password_resp.status_code == 200

    stale_after_change = client.get("/auth/me", headers=teacher_headers_new)
    assert stale_after_change.status_code == 401
    final_login = client.post("/auth/login", json={"email": email, "password": "TeacherPass789"})
    assert final_login.status_code == 200


def test_owner_send_invite_requires_smtp_config(client):
    owner_headers = _auth_headers(client)
    create_resp = client.post(
        "/auth/users",
        headers=owner_headers,
        json={
            "email": f"teacher_{uuid.uuid4().hex[:8]}@app.local",
            "password": "TeacherPass123",
            "full_name": "Invite Teacher",
            "role": "teacher",
        },
    )
    assert create_resp.status_code == 201
    teacher_id = create_resp.json()["id"]

    invite_resp = client.post(
        f"/auth/users/{teacher_id}/send-invite",
        headers=owner_headers,
        json={"temporary_password": "TeacherPass123", "app_url": "http://127.0.0.1:8000/app"},
    )
    assert invite_resp.status_code == 400
    assert "SMTP is not configured." in str(invite_resp.json()["detail"])


def test_owner_send_invite_success_with_mocked_mailer(client, monkeypatch):
    owner_headers = _auth_headers(client)
    create_resp = client.post(
        "/auth/users",
        headers=owner_headers,
        json={
            "email": f"teacher_{uuid.uuid4().hex[:8]}@app.local",
            "password": "TeacherPass123",
            "full_name": "Invite Teacher",
            "role": "teacher",
        },
    )
    assert create_resp.status_code == 201
    teacher_id = create_resp.json()["id"]
    teacher_email = create_resp.json()["email"]

    sent_payload: dict[str, str] = {}

    def fake_send_email(*, to_email: str, subject: str, body_text: str) -> None:
        sent_payload["to_email"] = to_email
        sent_payload["subject"] = subject
        sent_payload["body_text"] = body_text

    monkeypatch.setattr("app.routers.auth.smtp_is_configured", lambda: True)
    monkeypatch.setattr("app.routers.auth.send_email", fake_send_email)

    invite_resp = client.post(
        f"/auth/users/{teacher_id}/send-invite",
        headers=owner_headers,
        json={"temporary_password": "TeacherPass123", "app_url": "http://127.0.0.1:8000/app"},
    )
    assert invite_resp.status_code == 200
    payload = invite_resp.json()
    assert payload["sent"] is True
    assert payload["to_email"] == teacher_email
    assert payload["app_url"] == "http://127.0.0.1:8000/app"
    assert payload["included_temporary_password"] is True
    assert sent_payload["to_email"] == teacher_email
    assert "Teacher Platform Login Details" in sent_payload["subject"]
    assert "Temporary password: TeacherPass123" in sent_payload["body_text"]


def _setup_unit_with_leaf(client, headers: dict) -> tuple[int, int, int]:
    """Returns (class_id, unit_id, leaf_item_id) for a unit that has at least one leaf item."""
    class_resp = client.post(
        "/classes",
        json={"name": f"LeafContent Class {uuid.uuid4().hex[:6]}", "subject": "Math"},
        headers=headers,
    )
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    start_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Chapitre Leaf Test",
            "source_text": "Section 1\n1.1 Propriete\n1.2 Exemple\nSection 2\n2.1 Exercice",
        },
    )
    assert start_resp.status_code == 201
    unit = start_resp.json()
    unit_id = unit["id"]

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_resp.status_code == 200
    checklist = workspace_resp.json()["active_unit"]["checklist"]
    leaf = _first_leaf_checklist_item(checklist)
    assert leaf is not None, "Unit must have at least one leaf item"
    return class_id, unit_id, int(leaf["id"])


def test_leaf_content_happy_path(client):
    headers = _auth_headers(client)
    class_id, unit_id, item_id = _setup_unit_with_leaf(client, headers)

    # GET before any record exists returns 404
    get_missing = client.get(
        f"/workflow/classes/{class_id}/units/{unit_id}/leaf-content/{item_id}",
        headers=headers,
    )
    assert get_missing.status_code == 404

    # PUT creates the record
    put_resp = client.put(
        f"/workflow/classes/{class_id}/units/{unit_id}/leaf-content/{item_id}",
        headers=headers,
        json={
            "teaching_goal_md": "Student will understand the property.",
            "explanation_md": "An explanation in **Markdown** with $x^2$.",
            "worked_example_md": "Example: $2x + 3 = 7$",
            "provider": "manual",
            "status": "draft",
            "source_payload": {"requested_by": "test"},
            "raw_provider_response": {"raw": True},
        },
    )
    assert put_resp.status_code == 200
    body = put_resp.json()
    assert body["unit_id"] == unit_id
    assert body["checklist_item_id"] == item_id
    assert body["teaching_goal_md"] == "Student will understand the property."
    assert body["explanation_md"] == "An explanation in **Markdown** with $x^2$."
    assert body["worked_example_md"] == "Example: $2x + 3 = 7$"
    assert body["provider"] == "manual"
    assert body["status"] == "draft"
    assert body["reviewed"] is False
    assert body["source_payload_json"] == {"requested_by": "test"}
    assert body["raw_provider_response_json"] == {"raw": True}
    assert isinstance(body["item_path_json"], list) and body["item_path_json"]
    assert isinstance(body["section_path_json"], list) and body["section_path_json"]

    # GET now returns the same record
    get_resp = client.get(
        f"/workflow/classes/{class_id}/units/{unit_id}/leaf-content/{item_id}",
        headers=headers,
    )
    assert get_resp.status_code == 200
    fetched = get_resp.json()
    assert fetched["id"] == body["id"]
    assert fetched["teaching_goal_md"] == "Student will understand the property."
    assert fetched["explanation_md"] == "An explanation in **Markdown** with $x^2$."
    assert fetched["source_payload_json"] == {"requested_by": "test"}
    assert fetched["raw_provider_response_json"] == {"raw": True}

    # PUT again updates the existing record
    put2_resp = client.put(
        f"/workflow/classes/{class_id}/units/{unit_id}/leaf-content/{item_id}",
        headers=headers,
        json={"practice_md": "Practice: solve $x^2 - 4 = 0$", "status": "ready"},
    )
    assert put2_resp.status_code == 200
    updated = put2_resp.json()
    assert updated["id"] == body["id"]
    assert updated["practice_md"] == "Practice: solve $x^2 - 4 = 0$"
    assert updated["status"] == "ready"
    # Fields not in the second PUT remain unchanged
    assert updated["teaching_goal_md"] == "Student will understand the property."


def test_leaf_content_rejects_non_leaf_item(client):
    headers = _auth_headers(client)
    class_resp = client.post(
        "/classes",
        json={"name": f"NonLeaf Class {uuid.uuid4().hex[:6]}", "subject": "Math"},
        headers=headers,
    )
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    start_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Non-Leaf Unit",
            "source_text": "Section 1\n1.1 Propriete\n1.2 Exemple",
        },
    )
    assert start_resp.status_code == 201
    unit_id = start_resp.json()["id"]

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    checklist = workspace_resp.json()["active_unit"]["checklist"]
    flat = _flatten_checklist(checklist)
    # Find an item that has children (i.e., a parent/non-leaf)
    parent_ids = {
        int(row["parent_item_id"])
        for row in flat
        if isinstance(row, dict) and row.get("parent_item_id") is not None
    }
    non_leaf = next(
        (row for row in flat if isinstance(row, dict) and int(row.get("id") or 0) in parent_ids),
        None,
    )
    if non_leaf is None:
        # All items are leaves (flat checklist) — add a child to make one non-leaf
        parent_item = flat[0]
        add_resp = client.post(
            f"/workflow/classes/{class_id}/units/{unit_id}/items",
            headers=headers,
            json={"title": "Child item", "item_kind": "other", "parent_item_id": parent_item["id"]},
        )
        assert add_resp.status_code == 201
        non_leaf = parent_item

    non_leaf_id = int(non_leaf["id"])

    # GET on non-leaf returns 400
    get_resp = client.get(
        f"/workflow/classes/{class_id}/units/{unit_id}/leaf-content/{non_leaf_id}",
        headers=headers,
    )
    assert get_resp.status_code == 400
    assert "leaf" in get_resp.json()["detail"].lower()

    # PUT on non-leaf also returns 400
    put_resp = client.put(
        f"/workflow/classes/{class_id}/units/{unit_id}/leaf-content/{non_leaf_id}",
        headers=headers,
        json={"explanation_md": "should be rejected"},
    )
    assert put_resp.status_code == 400
    assert "leaf" in put_resp.json()["detail"].lower()


def test_leaf_content_generate_happy_path(client, monkeypatch):
    from app.routers import workflow as workflow_router

    headers = _auth_headers(client)
    class_id, unit_id, item_id = _setup_unit_with_leaf(client, headers)

    captured: dict = {}

    def _fake_generate_leaf_content_package(**kwargs):
        captured.update(kwargs)
        return {
            "provider": "notebooklm",
            "requested_provider": "notebooklm",
            "model": "notebooklm-py",
            "status": "ready",
            "teaching_goal_md": "L'eleve comprend la propriete.",
            "launch_activity_md": "Activite d'amorce: question ouverte.",
            "explanation_md": "Explication en **Markdown** avec $x^2$.",
            "worked_example_md": "Exemple: $2x + 3 = 7 \\Rightarrow x = 2$",
            "practice_md": "Resoudre $x^2 - 4 = 0$.",
            "solution_md": "$x = \\pm 2$",
            "assessment_md": "Mini-evaluation: donner un exemple.",
            "teacher_notes_md": "Note interne enseignant.",
            "source_excerpt_md": "Extrait du document source.",
            "source_payload": {"item_title": "Propriete", "requested_provider": "notebooklm"},
            "raw_provider_response": {"answer": "..."},
            "error_message": None,
        }

    monkeypatch.setattr(workflow_router, "generate_leaf_content_package", _fake_generate_leaf_content_package)

    resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/leaf-content/{item_id}/generate",
        headers=headers,
        json={"provider": "notebooklm", "regenerate": True},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["requested_provider"] == "notebooklm"
    assert body["provider"] == "notebooklm"
    assert body["status"] == "ready"
    lc = body["leaf_content"]
    assert lc["unit_id"] == unit_id
    assert lc["checklist_item_id"] == item_id
    assert lc["teaching_goal_md"] == "L'eleve comprend la propriete."
    assert lc["explanation_md"] == "Explication en **Markdown** avec $x^2$."
    assert lc["worked_example_md"] == "Exemple: $2x + 3 = 7 \\Rightarrow x = 2$"
    assert lc["practice_md"] == "Resoudre $x^2 - 4 = 0$."
    assert lc["solution_md"] == "$x = \\pm 2$"
    assert lc["provider"] == "notebooklm"
    assert lc["status"] == "ready"
    assert isinstance(lc["item_path_json"], list) and lc["item_path_json"]
    assert isinstance(lc["section_path_json"], list) and lc["section_path_json"]
    assert body["leaf_content"]["source_payload_json"]["item_path"] == lc["item_path_json"]
    assert body["leaf_content"]["source_payload_json"]["section_path"] == lc["section_path_json"]
    assert body["leaf_content"]["source_payload_json"]["merge_strategy"] == "fill_missing"

    # GET also returns the persisted record
    get_resp = client.get(
        f"/workflow/classes/{class_id}/units/{unit_id}/leaf-content/{item_id}",
        headers=headers,
    )
    assert get_resp.status_code == 200
    fetched = get_resp.json()
    assert fetched["id"] == lc["id"]
    assert fetched["teaching_goal_md"] == "L'eleve comprend la propriete."

    assert captured["unit_title"]
    assert captured["item_title"]

    second_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/leaf-content/{item_id}/generate",
        headers=headers,
        json={"provider": "notebooklm", "regenerate": False},
    )
    assert second_resp.status_code == 200
    second_body = second_resp.json()
    assert second_body["leaf_content"]["id"] == lc["id"]


def test_leaf_content_generate_fill_missing_preserves_source_fields(client, monkeypatch):
    from app.routers import workflow as workflow_router

    headers = _auth_headers(client)
    class_id, unit_id, item_id = _setup_unit_with_leaf(client, headers)

    put_resp = client.put(
        f"/workflow/classes/{class_id}/units/{unit_id}/leaf-content/{item_id}",
        headers=headers,
        json={
            "provider": "source_extract",
            "status": "ready",
            "teaching_goal_md": "Comprendre la propriete.",
            "explanation_md": "Explication extraite du PDF.",
            "source_excerpt_md": "Extrait source.",
            "source_payload": {"mode": "source_derived"},
            "raw_provider_response": {"mode": "source_derived"},
        },
    )
    assert put_resp.status_code == 200

    def _fake_generate_leaf_content_package(**kwargs):
        return {
            "provider": "notebooklm",
            "requested_provider": "notebooklm",
            "model": "notebooklm-py",
            "status": "ready",
            "teaching_goal_md": "Objectif regenere.",
            "launch_activity_md": None,
            "explanation_md": "Nouvelle explication IA.",
            "worked_example_md": None,
            "practice_md": "Nouvel exercice de pratique.",
            "solution_md": "Correction du nouvel exercice.",
            "assessment_md": None,
            "teacher_notes_md": "Conseil enseignant.",
            "source_excerpt_md": "Nouvel extrait.",
            "source_payload": {"mode": "generated"},
            "raw_provider_response": {"answer": "generated"},
            "error_message": None,
        }

    monkeypatch.setattr(workflow_router, "generate_leaf_content_package", _fake_generate_leaf_content_package)

    resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/leaf-content/{item_id}/generate",
        headers=headers,
        json={"provider": "notebooklm", "regenerate": True, "merge_strategy": "fill_missing"},
    )
    assert resp.status_code == 200
    body = resp.json()
    lc = body["leaf_content"]
    assert lc["provider"] == "notebooklm"
    assert lc["teaching_goal_md"] == "Comprendre la propriete."
    assert lc["explanation_md"] == "Explication extraite du PDF."
    assert lc["source_excerpt_md"] == "Extrait source."
    assert lc["practice_md"] == "Nouvel exercice de pratique."
    assert lc["solution_md"] == "Correction du nouvel exercice."
    assert lc["teacher_notes_md"] == "Conseil enseignant."
    assert lc["source_payload_json"]["mode"] == "hybrid"
    assert "practice_md" in lc["source_payload_json"]["filled_fields"]
    assert "explanation_md" in lc["source_payload_json"]["retained_fields"]


def test_leaf_content_generate_rejects_non_leaf(client, monkeypatch):
    from app.routers import workflow as workflow_router

    headers = _auth_headers(client)
    class_resp = client.post(
        "/classes",
        json={"name": f"GenNonLeaf {uuid.uuid4().hex[:6]}", "subject": "Math"},
        headers=headers,
    )
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    start_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "chapter", "title": "Gen Non-Leaf Unit", "source_text": "Section 1\n1.1 Propriete"},
    )
    assert start_resp.status_code == 201
    unit_id = start_resp.json()["id"]

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    checklist = workspace_resp.json()["active_unit"]["checklist"]
    flat = _flatten_checklist(checklist)
    parent_ids = {
        int(row["parent_item_id"])
        for row in flat
        if isinstance(row, dict) and row.get("parent_item_id") is not None
    }
    non_leaf = next((row for row in flat if int(row.get("id") or 0) in parent_ids), None)
    if non_leaf is None:
        parent_item = flat[0]
        client.post(
            f"/workflow/classes/{class_id}/units/{unit_id}/items",
            headers=headers,
            json={"title": "Child", "item_kind": "other", "parent_item_id": parent_item["id"]},
        )
        non_leaf = parent_item

    monkeypatch.setattr(workflow_router, "generate_leaf_content_package", lambda **kwargs: {})

    resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/leaf-content/{non_leaf['id']}/generate",
        headers=headers,
        json={},
    )
    assert resp.status_code == 400
    assert "leaf" in resp.json()["detail"].lower()


def test_leaf_content_list_by_unit(client):
    headers = _auth_headers(client)
    class_id, unit_id, item_id = _setup_unit_with_leaf(client, headers)

    # Initially empty
    resp = client.get(
        f"/workflow/classes/{class_id}/units/{unit_id}/leaf-content",
        headers=headers,
    )
    assert resp.status_code == 200
    assert resp.json() == []

    # Create a leaf content record via PUT
    client.put(
        f"/workflow/classes/{class_id}/units/{unit_id}/leaf-content/{item_id}",
        headers=headers,
        json={"teaching_goal_md": "Understand the concept.", "status": "ready"},
    )

    # List should now return one summary row with safe fields only
    resp2 = client.get(
        f"/workflow/classes/{class_id}/units/{unit_id}/leaf-content",
        headers=headers,
    )
    assert resp2.status_code == 200
    rows = resp2.json()
    assert len(rows) == 1
    row = rows[0]
    assert row["checklist_item_id"] == item_id
    assert row["status"] == "ready"
    assert "provider" in row
    assert "reviewed" in row
    assert "updated_at" in row
    # Summary should NOT expose content fields
    assert "teaching_goal_md" not in row


def test_leaf_content_source_extract_seeds_on_unit_start(client, monkeypatch):
    from app.routers import workflow as workflow_router

    headers = _auth_headers(client)
    class_resp = client.post(
        "/classes",
        json={"name": f"SeededLeaf {uuid.uuid4().hex[:6]}", "subject": "Math"},
        headers=headers,
    )
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    def _fake_generate_unit_checklist(unit_type, title, source_text, session_count=None, document_path=None):
        return {
            "source": "notebooklm",
            "requested_provider": "notebooklm",
            "model": "notebooklm-py",
            "status": "ready",
            "items": [
                {
                    "title": "Les nombres rationnels",
                    "kind": "chapter",
                    "children": [
                        {
                            "title": "Les denominateurs sont les memes",
                            "kind": "section",
                            "children": [
                                {
                                    "title": "Regle de calcul pour denominateurs communs",
                                    "kind": "property",
                                    "children": [],
                                },
                                {
                                    "title": "Exemples d'application",
                                    "kind": "example",
                                    "children": [],
                                },
                            ],
                        }
                    ],
                }
            ],
            "unit_map": {"unit_title": title, "source": "notebooklm"},
            "content_blocks": [
                {
                    "section_title": "Les denominateurs sont les memes",
                    "section_path": ["Les nombres rationnels", "Les denominateurs sont les memes"],
                    "title": "Regle de calcul pour denominateurs communs",
                    "kind": "property",
                    "teaching_material": "Pour additionner deux rationnels de meme denominateur, on additionne les numerateurs et on garde le denominateur.",
                    "source_excerpt": "Regle: meme denominateur, on garde le denominateur et on additionne les numerateurs.",
                },
                {
                    "section_title": "Les denominateurs sont les memes",
                    "section_path": ["Les nombres rationnels", "Les denominateurs sont les memes"],
                    "title": "Exemples d'application",
                    "kind": "example",
                    "teaching_material": "Exemple: 3/7 + 2/7 = 5/7.",
                    "source_excerpt": "Exemple direct avec meme denominateur.",
                },
            ],
            "raw_provider_response": {"seeded": True},
            "error_message": None,
            "provider_context": {"notebook_id": "test-seeded-leaf"},
        }

    monkeypatch.setattr(workflow_router, "generate_unit_checklist", _fake_generate_unit_checklist)

    start_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Chapitre Seeded Leaf",
            "source_text": "placeholder source text",
        },
    )
    assert start_resp.status_code == 201
    unit_id = start_resp.json()["id"]

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_resp.status_code == 200
    checklist = workspace_resp.json()["active_unit"]["checklist"]
    flat = _flatten_checklist(checklist)
    property_leaf = next(
        row
        for row in flat
        if isinstance(row, dict) and row.get("title") == "Regle de calcul pour denominateurs communs"
    )

    list_resp = client.get(
        f"/workflow/classes/{class_id}/units/{unit_id}/leaf-content",
        headers=headers,
    )
    assert list_resp.status_code == 200
    rows = list_resp.json()
    assert len(rows) == 2
    assert {row["provider"] for row in rows} == {"source_extract"}
    assert {row["status"] for row in rows} == {"ready"}

    get_resp = client.get(
        f"/workflow/classes/{class_id}/units/{unit_id}/leaf-content/{property_leaf['id']}",
        headers=headers,
    )
    assert get_resp.status_code == 200
    body = get_resp.json()
    assert body["provider"] == "source_extract"
    assert body["status"] == "ready"
    assert body["teaching_goal_md"] == "Regle de calcul pour denominateurs communs"
    assert "additionne les numerateurs" in (body["explanation_md"] or "")
    assert body["source_payload_json"]["mode"] == "source_derived"
    assert body["source_payload_json"]["matched_block_count"] >= 1
    extracted_blocks = body["source_payload_json"].get("extracted_blocks") or []
    assert extracted_blocks
    assert extracted_blocks[0]["content_md"]
    assert extracted_blocks[0]["content_source"] in {"source_excerpt", "teaching_material"}
    assert "additionner deux rationnels de meme denominateur" in extracted_blocks[0]["content_md"]


def test_leaf_content_source_extract_assigns_examples_by_sequence(client, monkeypatch):
    from app.routers import workflow as workflow_router

    headers = _auth_headers(client)
    class_resp = client.post(
        "/classes",
        json={"name": f"SeededExampleOrder {uuid.uuid4().hex[:6]}", "subject": "Math"},
        headers=headers,
    )
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    def _fake_generate_unit_checklist(unit_type, title, source_text, session_count=None, document_path=None):
        return {
            "source": "notebooklm",
            "requested_provider": "notebooklm",
            "model": "notebooklm-py",
            "status": "ready",
            "items": [
                {
                    "title": "Fractions",
                    "kind": "chapter",
                    "children": [
                        {
                            "title": "Somme de fractions",
                            "kind": "section",
                            "children": [
                                {"title": "Exemple 1", "kind": "example", "children": []},
                                {"title": "Exemple 2", "kind": "example", "children": []},
                            ],
                        }
                    ],
                }
            ],
            "unit_map": {"unit_title": title, "source": "notebooklm"},
            "content_blocks": [
                {
                    "section_title": "Somme de fractions",
                    "section_path": ["Fractions", "Somme de fractions"],
                    "title": "Addition simple",
                    "kind": "example",
                    "teaching_material": "Exemple: 1/5 + 2/5 = 3/5.",
                    "source_excerpt": "Premier exemple avec meme denominateur.",
                },
                {
                    "section_title": "Somme de fractions",
                    "section_path": ["Fractions", "Somme de fractions"],
                    "title": "Soustraction simple",
                    "kind": "example",
                    "teaching_material": "Exemple: 4/9 - 1/9 = 3/9.",
                    "source_excerpt": "Deuxieme exemple avec meme denominateur.",
                },
            ],
            "raw_provider_response": {"seeded": True},
            "error_message": None,
            "provider_context": {"notebook_id": "test-seeded-examples"},
        }

    monkeypatch.setattr(workflow_router, "generate_unit_checklist", _fake_generate_unit_checklist)

    start_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={
            "unit_type": "chapter",
            "title": "Chapitre Exemples",
            "source_text": "placeholder source text",
        },
    )
    assert start_resp.status_code == 201
    unit_id = start_resp.json()["id"]

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    assert workspace_resp.status_code == 200
    checklist = workspace_resp.json()["active_unit"]["checklist"]
    flat = _flatten_checklist(checklist)
    example1 = next(row for row in flat if isinstance(row, dict) and row.get("title") == "Exemple 1")
    example2 = next(row for row in flat if isinstance(row, dict) and row.get("title") == "Exemple 2")

    first_resp = client.get(
        f"/workflow/classes/{class_id}/units/{unit_id}/leaf-content/{example1['id']}",
        headers=headers,
    )
    second_resp = client.get(
        f"/workflow/classes/{class_id}/units/{unit_id}/leaf-content/{example2['id']}",
        headers=headers,
    )
    assert first_resp.status_code == 200
    assert second_resp.status_code == 200
    first_body = first_resp.json()
    second_body = second_resp.json()
    assert "1/5 + 2/5 = 3/5" in (first_body["worked_example_md"] or "")
    assert "4/9 - 1/9 = 3/9" in (second_body["worked_example_md"] or "")
    assert first_body["worked_example_md"] != second_body["worked_example_md"]


def test_normalize_content_blocks_infers_example_kind_from_generic_block():
    from app.services import workflow_generation

    blocks = workflow_generation._normalize_content_blocks_payload(
        {
            "content_blocks": [
                {
                    "section_title": "Somme et difference",
                    "section_path": ["Rationnels", "Somme et difference"],
                    "title": "Somme de rationnels",
                    "kind": "lesson",
                    "teaching_material": "Exemple: 3/8 + 1/8 = 4/8.",
                    "source_excerpt": "Exemple resolu avec meme denominateur.",
                }
            ]
        },
        unit_map=None,
        fallback_outline=None,
    )
    assert len(blocks) == 1
    assert blocks[0]["kind"] == "example"


def test_normalize_content_blocks_splits_mixed_block_into_ordered_parts():
    from app.services import workflow_generation

    blocks = workflow_generation._normalize_content_blocks_payload(
        {
            "content_blocks": [
                {
                    "section_title": "Somme et difference",
                    "section_path": ["Rationnels", "Somme et difference"],
                    "title": "Somme de rationnels",
                    "kind": "lesson",
                    "teaching_material": (
                        "Definition: Pour additionner deux rationnels de meme denominateur, on garde le denominateur. "
                        "Exemple 1: 3/8 + 1/8 = 4/8. "
                        "Exercice 1: Calculer 5/9 + 2/9."
                    ),
                    "source_excerpt": "Bloc combine du document.",
                }
            ]
        },
        unit_map=None,
        fallback_outline=None,
    )
    assert [row["kind"] for row in blocks] == ["definition", "example", "exercise"]
    assert blocks[0]["title"] == "Definition"
    assert "garde le denominateur" in blocks[0]["teaching_material"]
    assert blocks[1]["title"] == "Exemple 1"
    assert "3/8 + 1/8 = 4/8" in blocks[1]["teaching_material"]
    assert blocks[2]["title"] == "Exercice 1"
    assert "5/9 + 2/9" in blocks[2]["teaching_material"]


def test_source_derived_leaf_content_uses_split_exact_block_for_matching():
    from app.services import workflow_generation
    from app.models import WorkflowChecklistItemKind

    content_blocks = workflow_generation._normalize_content_blocks_payload(
        {
            "content_blocks": [
                {
                    "section_title": "Somme et difference",
                    "section_path": ["Rationnels", "Somme et difference"],
                    "title": "Somme de rationnels",
                    "kind": "lesson",
                    "teaching_material": (
                        "Definition: Pour additionner deux rationnels de meme denominateur, on garde le denominateur. "
                        "Exemple 1: 3/8 + 1/8 = 4/8. "
                        "Exercice 1: Calculer 5/9 + 2/9."
                    ),
                    "source_excerpt": "Bloc combine du document.",
                }
            ]
        },
        unit_map=None,
        fallback_outline=None,
    )

    example_payload = workflow_generation.build_source_derived_leaf_content_package(
        item_title="Exemple 1",
        item_kind=WorkflowChecklistItemKind.EXAMPLE,
        item_path=["Rationnels", "Somme et difference", "Exemple 1"],
        section_path=["Rationnels", "Somme et difference"],
        content_blocks=content_blocks,
    )
    exercise_payload = workflow_generation.build_source_derived_leaf_content_package(
        item_title="Exercice 1",
        item_kind=WorkflowChecklistItemKind.EXERCISE,
        item_path=["Rationnels", "Somme et difference", "Exercice 1"],
        section_path=["Rationnels", "Somme et difference"],
        content_blocks=content_blocks,
    )

    assert "3/8 + 1/8 = 4/8" in (example_payload["worked_example_md"] or "")
    assert "5/9 + 2/9" in (exercise_payload["practice_md"] or "")
    exact_blocks = example_payload["source_payload"]["extracted_blocks"]
    assert exact_blocks
    assert "3/8 + 1/8 = 4/8" in exact_blocks[0]["content_md"]


def test_source_derived_leaf_content_preserves_multiline_exact_source():
    from app.services import workflow_generation
    from app.models import WorkflowChecklistItemKind

    content_blocks = workflow_generation._normalize_content_blocks_payload(
        {
            "content_blocks": [
                {
                    "section_title": "Produit et division",
                    "section_path": ["Rationnels", "Produit et division"],
                    "title": "Serie d'exercices de multiplication",
                    "kind": "exercise",
                    "teaching_material": (
                        "Effectuer les calculs suivants : -11/5 × 2/3 ; 8/9 × -4/9.\n"
                        "Donner le resultat sous forme simplifiee.\n\n"
                        "Surveiller la gestion des signes des produits."
                    ),
                    "source_excerpt": "Bloc d'exercices de multiplication.",
                }
            ]
        },
        unit_map=None,
        fallback_outline=None,
    )

    payload = workflow_generation.build_source_derived_leaf_content_package(
        item_title="Serie d'exercices de multiplication",
        item_kind=WorkflowChecklistItemKind.EXERCISE,
        item_path=["Rationnels", "Produit et division", "Serie d'exercices de multiplication"],
        section_path=["Rationnels", "Produit et division"],
        content_blocks=content_blocks,
    )

    assert "Effectuer les calculs suivants" in (payload["practice_md"] or "")
    assert "Donner le resultat sous forme simplifiee" in (payload["practice_md"] or "")
    assert "\n" in (payload["practice_md"] or "")
    exact_blocks = payload["source_payload"]["extracted_blocks"]
    assert exact_blocks
    assert any("Effectuer les calculs suivants" in row["content_md"] for row in exact_blocks)
    assert any("Donner le resultat sous forme simplifiee" in row["content_md"] for row in exact_blocks)


def test_source_derived_leaf_content_splits_inline_numbered_rows():
    from app.services import workflow_generation
    from app.models import WorkflowChecklistItemKind

    content_blocks = workflow_generation._normalize_content_blocks_payload(
        {
            "content_blocks": [
                {
                    "section_title": "Produit et division",
                    "section_path": ["Rationnels", "Produit et division"],
                    "title": "Serie d'exercices de multiplication",
                    "kind": "exercise",
                    "teaching_material": "1) Calculer 2/3 × 4/5. 2) Simplifier 6/9 × 3/2. 3) Donner le signe du produit.",
                    "source_excerpt": "Bloc d'exercices inline.",
                }
            ]
        },
        unit_map=None,
        fallback_outline=None,
    )

    payload = workflow_generation.build_source_derived_leaf_content_package(
        item_title="Serie d'exercices de multiplication",
        item_kind=WorkflowChecklistItemKind.EXERCISE,
        item_path=["Rationnels", "Produit et division", "Serie d'exercices de multiplication"],
        section_path=["Rationnels", "Produit et division"],
        content_blocks=content_blocks,
    )

    assert "1) Calculer 2/3 × 4/5." in (payload["practice_md"] or "")
    assert "2) Simplifier 6/9 × 3/2." in (payload["practice_md"] or "")
    assert "3) Donner le signe du produit." in (payload["practice_md"] or "")
    assert "\n" in (payload["practice_md"] or "")


def test_source_derived_leaf_content_splits_exact_source_segments_for_numbered_exercises():
    from app.services import workflow_generation
    from app.models import WorkflowChecklistItemKind

    content_blocks = workflow_generation._normalize_content_blocks_payload(
        {
            "content_blocks": [
                {
                    "section_title": "Produit et division",
                    "section_path": ["Rationnels", "Produit et division"],
                    "title": "Serie d'exercices de multiplication",
                    "kind": "exercise",
                    "teaching_material": "1) Calculer 2/3 × 4/5.\n2) Simplifier 6/9 × 3/2.\n3) Donner le signe du produit.",
                    "source_excerpt": "Bloc d'exercices exact.",
                }
            ]
        },
        unit_map=None,
        fallback_outline=None,
    )

    payload = workflow_generation.build_source_derived_leaf_content_package(
        item_title="Serie d'exercices de multiplication",
        item_kind=WorkflowChecklistItemKind.EXERCISE,
        item_path=["Rationnels", "Produit et division", "Serie d'exercices de multiplication"],
        section_path=["Rationnels", "Produit et division"],
        content_blocks=content_blocks,
    )

    exact_blocks = payload["source_payload"]["extracted_blocks"]
    assert len(exact_blocks) == 3
    assert exact_blocks[0]["title"] == "Serie d'exercices de multiplication 1"
    assert "1) Calculer 2/3 × 4/5." in exact_blocks[0]["content_md"]
    assert "2) Simplifier 6/9 × 3/2." in exact_blocks[1]["content_md"]
    assert "3) Donner le signe du produit." in exact_blocks[2]["content_md"]


def test_source_derived_leaf_content_preserves_compact_math_sequences():
    from app.services import workflow_generation
    from app.models import WorkflowChecklistItemKind

    content_blocks = workflow_generation._normalize_content_blocks_payload(
        {
            "content_blocks": [
                {
                    "section_title": "Produit et division",
                    "section_path": ["Rationnels", "Produit et division"],
                    "title": "Exemple guide",
                    "kind": "example",
                    "teaching_material": "A = 2/3 ; B = 4/5 ; A × B = 8/15",
                    "source_excerpt": "Bloc d'exemple compact.",
                }
            ]
        },
        unit_map=None,
        fallback_outline=None,
    )

    payload = workflow_generation.build_source_derived_leaf_content_package(
        item_title="Exemple guide",
        item_kind=WorkflowChecklistItemKind.EXAMPLE,
        item_path=["Rationnels", "Produit et division", "Exemple guide"],
        section_path=["Rationnels", "Produit et division"],
        content_blocks=content_blocks,
    )

    assert "A = 2/3" in (payload["worked_example_md"] or "")
    assert "B = 4/5" in (payload["worked_example_md"] or "")
    assert "A × B = 8/15" in (payload["worked_example_md"] or "")
    assert "\n" in (payload["worked_example_md"] or "")
    exact_blocks = payload["source_payload"]["extracted_blocks"]
    assert len(exact_blocks) == 3
    assert exact_blocks[0]["content_md"] == "A = 2/3"
    assert exact_blocks[1]["content_md"] == "B = 4/5"
    assert exact_blocks[2]["content_md"] == "A × B = 8/15"


def test_source_derived_leaf_content_splits_short_statement_groups():
    from app.services import workflow_generation
    from app.models import WorkflowChecklistItemKind

    content_blocks = workflow_generation._normalize_content_blocks_payload(
        {
            "content_blocks": [
                {
                    "section_title": "Somme et difference",
                    "section_path": ["Rationnels", "Somme et difference"],
                    "title": "Regle generale",
                    "kind": "property",
                    "teaching_material": (
                        "On garde le denominateur.\n"
                        "On additionne les numerateurs.\n"
                        "On simplifie si possible."
                    ),
                    "source_excerpt": "Bloc de regle.",
                }
            ]
        },
        unit_map=None,
        fallback_outline=None,
    )

    payload = workflow_generation.build_source_derived_leaf_content_package(
        item_title="Regle generale",
        item_kind=WorkflowChecklistItemKind.PROPERTY,
        item_path=["Rationnels", "Somme et difference", "Regle generale"],
        section_path=["Rationnels", "Somme et difference"],
        content_blocks=content_blocks,
    )

    assert "On garde le denominateur." in (payload["explanation_md"] or "")
    assert "On additionne les numerateurs." in (payload["explanation_md"] or "")
    exact_blocks = payload["source_payload"]["extracted_blocks"]
    assert len(exact_blocks) == 3
    assert exact_blocks[0]["content_md"] == "On garde le denominateur."
    assert exact_blocks[1]["content_md"] == "On additionne les numerateurs."
    assert exact_blocks[2]["content_md"] == "On simplifie si possible."


def test_source_derived_leaf_content_splits_lettered_rows():
    from app.services import workflow_generation
    from app.models import WorkflowChecklistItemKind

    content_blocks = workflow_generation._normalize_content_blocks_payload(
        {
            "content_blocks": [
                {
                    "section_title": "Somme et difference",
                    "section_path": ["Rationnels", "Somme et difference"],
                    "title": "Exercices rapides",
                    "kind": "exercise",
                    "teaching_material": "a) Calculer 3/7 + 2/7. b) Simplifier 5/10. c) Comparer 4/9 et 5/9.",
                    "source_excerpt": "Bloc d'exercices lettres.",
                }
            ]
        },
        unit_map=None,
        fallback_outline=None,
    )

    payload = workflow_generation.build_source_derived_leaf_content_package(
        item_title="Exercices rapides",
        item_kind=WorkflowChecklistItemKind.EXERCISE,
        item_path=["Rationnels", "Somme et difference", "Exercices rapides"],
        section_path=["Rationnels", "Somme et difference"],
        content_blocks=content_blocks,
    )

    assert "a) Calculer 3/7 + 2/7." in (payload["practice_md"] or "")
    assert "b) Simplifier 5/10." in (payload["practice_md"] or "")
    assert "c) Comparer 4/9 et 5/9." in (payload["practice_md"] or "")
    exact_blocks = payload["source_payload"]["extracted_blocks"]
    assert len(exact_blocks) == 3
    assert "a) Calculer 3/7 + 2/7." in exact_blocks[0]["content_md"]
    assert "b) Simplifier 5/10." in exact_blocks[1]["content_md"]
    assert "c) Comparer 4/9 et 5/9." in exact_blocks[2]["content_md"]


def test_source_derived_leaf_content_splits_short_action_rows():
    from app.services import workflow_generation
    from app.models import WorkflowChecklistItemKind

    content_blocks = workflow_generation._normalize_content_blocks_payload(
        {
            "content_blocks": [
                {
                    "section_title": "Somme et difference",
                    "section_path": ["Rationnels", "Somme et difference"],
                    "title": "Applications rapides",
                    "kind": "exercise",
                    "teaching_material": (
                        "Calculer 3/7 + 2/7.\n"
                        "Simplifier 5/10.\n"
                        "Comparer 4/9 et 5/9."
                    ),
                    "source_excerpt": "Bloc d'applications courtes.",
                }
            ]
        },
        unit_map=None,
        fallback_outline=None,
    )

    payload = workflow_generation.build_source_derived_leaf_content_package(
        item_title="Applications rapides",
        item_kind=WorkflowChecklistItemKind.EXERCISE,
        item_path=["Rationnels", "Somme et difference", "Applications rapides"],
        section_path=["Rationnels", "Somme et difference"],
        content_blocks=content_blocks,
    )

    assert "Calculer 3/7 + 2/7." in (payload["practice_md"] or "")
    assert "Simplifier 5/10." in (payload["practice_md"] or "")
    assert "Comparer 4/9 et 5/9." in (payload["practice_md"] or "")
    exact_blocks = payload["source_payload"]["extracted_blocks"]
    assert len(exact_blocks) == 3
    assert exact_blocks[0]["content_md"] == "Calculer 3/7 + 2/7."
    assert exact_blocks[1]["content_md"] == "Simplifier 5/10."
    assert exact_blocks[2]["content_md"] == "Comparer 4/9 et 5/9."


def test_source_derived_leaf_content_splits_column_style_rows():
    from app.services import workflow_generation
    from app.models import WorkflowChecklistItemKind

    content_blocks = workflow_generation._normalize_content_blocks_payload(
        {
            "content_blocks": [
                {
                    "section_title": "Somme et difference",
                    "section_path": ["Rationnels", "Somme et difference"],
                    "title": "Exercices en colonnes",
                    "kind": "exercise",
                    "teaching_material": "Calculer 3/7 + 2/7.    Simplifier 5/10.    Comparer 4/9 et 5/9.",
                    "source_excerpt": "Bloc d'exercices en colonnes.",
                }
            ]
        },
        unit_map=None,
        fallback_outline=None,
    )

    payload = workflow_generation.build_source_derived_leaf_content_package(
        item_title="Exercices en colonnes",
        item_kind=WorkflowChecklistItemKind.EXERCISE,
        item_path=["Rationnels", "Somme et difference", "Exercices en colonnes"],
        section_path=["Rationnels", "Somme et difference"],
        content_blocks=content_blocks,
    )

    assert "Calculer 3/7 + 2/7." in (payload["practice_md"] or "")
    assert "Simplifier 5/10." in (payload["practice_md"] or "")
    assert "Comparer 4/9 et 5/9." in (payload["practice_md"] or "")
    assert "\n" in (payload["practice_md"] or "")
    exact_blocks = payload["source_payload"]["extracted_blocks"]
    assert len(exact_blocks) == 3
    assert exact_blocks[0]["content_md"] == "Calculer 3/7 + 2/7."
    assert exact_blocks[1]["content_md"] == "Simplifier 5/10."
    assert exact_blocks[2]["content_md"] == "Comparer 4/9 et 5/9."


def test_source_derived_leaf_content_keeps_heading_above_short_math_rows():
    from app.services import workflow_generation
    from app.models import WorkflowChecklistItemKind

    content_blocks = workflow_generation._normalize_content_blocks_payload(
        {
            "content_blocks": [
                {
                    "section_title": "Produit et division",
                    "section_path": ["Rationnels", "Produit et division"],
                    "title": "Exercices guides",
                    "kind": "exercise",
                    "teaching_material": (
                        "Calculer :\n"
                        "3/7 + 2/7 = 5/7\n"
                        "5/10 + 1/10 = 6/10\n"
                        "4/9 + 5/9 = 9/9"
                    ),
                    "source_excerpt": "Bloc d'exercices avec en-tete.",
                }
            ]
        },
        unit_map=None,
        fallback_outline=None,
    )

    payload = workflow_generation.build_source_derived_leaf_content_package(
        item_title="Exercices guides",
        item_kind=WorkflowChecklistItemKind.EXERCISE,
        item_path=["Rationnels", "Produit et division", "Exercices guides"],
        section_path=["Rationnels", "Produit et division"],
        content_blocks=content_blocks,
    )

    exact_blocks = payload["source_payload"]["extracted_blocks"]
    assert len(exact_blocks) == 4
    assert exact_blocks[0]["content_md"] == "Calculer :"
    assert exact_blocks[1]["content_md"] == "3/7 + 2/7 = 5/7"
    assert exact_blocks[2]["content_md"] == "5/10 + 1/10 = 6/10"
    assert exact_blocks[3]["content_md"] == "4/9 + 5/9 = 9/9"


def test_source_derived_leaf_content_splits_pipe_separated_rows():
    from app.services import workflow_generation
    from app.models import WorkflowChecklistItemKind

    content_blocks = workflow_generation._normalize_content_blocks_payload(
        {
            "content_blocks": [
                {
                    "section_title": "Produit et division",
                    "section_path": ["Rationnels", "Produit et division"],
                    "title": "Exercices tableau",
                    "kind": "exercise",
                    "teaching_material": "Calculer 3/7 + 2/7 | Simplifier 5/10 | Comparer 4/9 et 5/9",
                    "source_excerpt": "Bloc d'exercices avec separateur pipe.",
                }
            ]
        },
        unit_map=None,
        fallback_outline=None,
    )

    payload = workflow_generation.build_source_derived_leaf_content_package(
        item_title="Exercices tableau",
        item_kind=WorkflowChecklistItemKind.EXERCISE,
        item_path=["Rationnels", "Produit et division", "Exercices tableau"],
        section_path=["Rationnels", "Produit et division"],
        content_blocks=content_blocks,
    )

    assert "Calculer 3/7 + 2/7" in (payload["practice_md"] or "")
    assert "Simplifier 5/10" in (payload["practice_md"] or "")
    assert "Comparer 4/9 et 5/9" in (payload["practice_md"] or "")
    assert "\n" in (payload["practice_md"] or "")
    exact_blocks = payload["source_payload"]["extracted_blocks"]
    assert len(exact_blocks) == 3
    assert exact_blocks[0]["content_md"] == "Calculer 3/7 + 2/7"
    assert exact_blocks[1]["content_md"] == "Simplifier 5/10"
    assert exact_blocks[2]["content_md"] == "Comparer 4/9 et 5/9"


def test_build_source_section_lesson_package_keeps_exact_section_content():
    from app.services import workflow_generation

    content_blocks = workflow_generation._normalize_content_blocks_payload(
        {
            "content_blocks": [
                {
                    "section_title": "Produit et division",
                    "section_path": ["Rationnels", "Produit et division"],
                    "title": "Exemple de produit",
                    "kind": "example",
                    "teaching_material": "Calculer 2/3 x 4/5 = 8/15.",
                    "source_excerpt": "Bloc exemple produit.",
                },
                {
                    "section_title": "Produit et division",
                    "section_path": ["Rationnels", "Produit et division"],
                    "title": "Exercices de produit",
                    "kind": "exercise",
                    "teaching_material": "a) Calculer 3/7 x 2/7. b) Simplifier 5/10 x 1/2.",
                    "source_excerpt": "Bloc exercices produit.",
                },
                {
                    "section_title": "Somme et difference",
                    "section_path": ["Rationnels", "Somme et difference"],
                    "title": "Exemple de somme",
                    "kind": "example",
                    "teaching_material": "Calculer 3/7 + 2/7 = 5/7.",
                    "source_excerpt": "Bloc exemple somme.",
                },
            ]
        },
        unit_map=None,
        fallback_outline=None,
    )

    payload = workflow_generation.build_source_section_lesson_package(
        section_title="Produit et division",
        section_path=["Rationnels", "Produit et division"],
        item_path=["Rationnels", "Produit et division", "Exemple de produit"],
        item_title="Exemple de produit",
        content_blocks=content_blocks,
    )

    assert payload["section_title"] == "Produit et division"
    assert payload["source_block_count"] >= 2
    combined = "\n".join(block["content_md"] for block in payload["source_blocks"])
    assert "2/3 x 4/5" in combined
    assert "3/7 x 2/7" in combined
    assert "3/7 + 2/7" not in combined


def test_build_raw_section_lesson_package_prefers_exact_notebooklm_section_blocks():
    from app.services import workflow_generation

    content_pack = {
        "unit_title": "Les nombres rationnels",
        "sections": [
            {
                "section_title": "Produit et division",
                "section_path": ["Rationnels", "Produit et division"],
                "order_index": 1,
                "blocks": [
                    {
                        "kind": "example",
                        "title": "Exemple 1",
                        "exact_text": "Calculer 2/3 × 4/5 = 8/15.",
                        "order_index": 1,
                    },
                    {
                        "kind": "exercise",
                        "title": "Exercice 1",
                        "exact_text": "1) Calculer 3/7 × 2/7.\n2) Simplifier 5/10 × 1/2.",
                        "order_index": 2,
                    },
                ],
            }
        ],
    }

    payload = workflow_generation.build_raw_section_lesson_package(
        section_title="Produit et division",
        section_path=["Rationnels", "Produit et division"],
        item_path=["Rationnels", "Produit et division", "Exemple 1"],
        item_title="Exemple 1",
        content_pack=content_pack,
    )

    assert payload is not None
    assert payload["section_title"] == "Produit et division"
    assert payload["source_block_count"] == 2
    assert payload["source_blocks"][0]["content_md"] == "Calculer 2/3 × 4/5 = 8/15."
    assert payload["source_blocks"][1]["content_md"] == "1) Calculer 3/7 × 2/7.\n2) Simplifier 5/10 × 1/2."


def test_build_source_section_index_orders_unique_sections():
    from app.services import workflow_generation

    content_blocks = workflow_generation._normalize_content_blocks_payload(
        {
            "content_blocks": [
                {
                    "section_title": "Produit et division",
                    "section_path": ["Rationnels", "Produit et division"],
                    "title": "Exemple de produit",
                    "kind": "example",
                    "teaching_material": "Calculer 2/3 x 4/5 = 8/15.",
                },
                {
                    "section_title": "Produit et division",
                    "section_path": ["Rationnels", "Produit et division"],
                    "title": "Exercices de produit",
                    "kind": "exercise",
                    "teaching_material": "Calculer 3/7 x 2/7.",
                },
                {
                    "section_title": "Somme et difference",
                    "section_path": ["Rationnels", "Somme et difference"],
                    "title": "Exemple de somme",
                    "kind": "example",
                    "teaching_material": "Calculer 3/7 + 2/7 = 5/7.",
                },
            ]
        },
        unit_map=None,
        fallback_outline=None,
    )

    rows = workflow_generation.build_source_section_index(content_blocks)

    assert [row["section_title"] for row in rows] == ["Produit et division", "Somme et difference"]
    assert rows[0]["section_path_json"] == ["Rationnels", "Produit et division"]
    assert rows[0]["order_index"] == 0
    assert rows[1]["order_index"] == 1
    assert rows[0]["section_key"]


def test_build_raw_section_index_uses_notebooklm_section_order():
    from app.services import workflow_generation

    rows = workflow_generation.build_raw_section_index(
        {
            "sections": [
                {
                    "section_title": "Produit et division",
                    "section_path": ["Rationnels", "Produit et division"],
                    "order_index": 2,
                    "blocks": [{"kind": "example", "title": "Exemple", "exact_text": "Calculer 2/3 × 4/5.", "order_index": 1}],
                },
                {
                    "section_title": "Somme et difference",
                    "section_path": ["Rationnels", "Somme et difference"],
                    "order_index": 1,
                    "blocks": [{"kind": "example", "title": "Exemple", "exact_text": "Calculer 3/7 + 2/7.", "order_index": 1}],
                },
            ]
        }
    )

    assert [row["section_title"] for row in rows] == ["Somme et difference", "Produit et division"]
    assert rows[0]["source_block_count"] == 1
    assert rows[1]["section_path_json"] == ["Rationnels", "Produit et division"]


def test_build_raw_section_lesson_package_accepts_content_bank_sequence_shape():
    from app.services import workflow_generation

    payload = workflow_generation.build_raw_section_lesson_package(
        section_title="1) Les dénominateurs sont les mêmes",
        section_path=["I- Addition et soustraction", "1) Les dénominateurs sont les mêmes"],
        item_path=["I- Addition et soustraction", "1) Les dénominateurs sont les mêmes", "Exemple 1"],
        item_title="Exemple 1",
        content_pack={
            "content_bank": [
                {
                    "content_id": "S01-B01",
                    "content_label": "Activité 1",
                    "content_type": "activity",
                    "source_heading_path": ["Activités"],
                    "pedagogical_section_path": ["I- Addition et soustraction", "1) Les dénominateurs sont les mêmes"],
                    "source_order": 1,
                    "exact_content": "Calculer : 1/3 + 4/3 et 4/7 - 3/7",
                },
                {
                    "content_id": "S01-B02",
                    "content_label": "Règle",
                    "content_type": "property",
                    "source_heading_path": ["Contenu de la leçon", "I- Addition et soustraction", "1) Les dénominateurs sont les mêmes"],
                    "pedagogical_section_path": ["I- Addition et soustraction", "1) Les dénominateurs sont les mêmes"],
                    "source_order": 2,
                    "exact_content": "On garde le dénominateur commun.",
                },
            ],
            "pedagogy_sequence": [
                {
                    "section_title": "1) Les dénominateurs sont les mêmes",
                    "section_path": ["I- Addition et soustraction", "1) Les dénominateurs sont les mêmes"],
                    "sequence_order": 1,
                    "content_ids": ["S01-B01", "S01-B02"],
                }
            ],
        },
    )

    assert payload is not None
    assert payload["source_block_count"] == 2
    assert payload["source_blocks"][0]["title"] == "Activité 1"
    assert payload["source_blocks"][1]["title"] == "Règle"
    assert payload["source_blocks"][1]["content_md"] == "On garde le dénominateur commun."


def test_build_raw_section_index_accepts_content_bank_sequence_shape():
    from app.services import workflow_generation

    rows = workflow_generation.build_raw_section_index(
        {
            "content_bank": [
                {
                    "content_id": "S01-B01",
                    "content_label": "Définition",
                    "content_type": "definition",
                    "source_heading_path": ["Leçon"],
                    "pedagogical_section_path": ["Chapitre", "Section A"],
                    "source_order": 1,
                    "exact_content": "Texte A",
                },
                {
                    "content_id": "S02-B01",
                    "content_label": "Exemple",
                    "content_type": "example",
                    "source_heading_path": ["Leçon"],
                    "pedagogical_section_path": ["Chapitre", "Section B"],
                    "source_order": 2,
                    "exact_content": "Texte B",
                },
            ],
            "pedagogy_sequence": [
                {
                    "section_title": "Section B",
                    "section_path": ["Chapitre", "Section B"],
                    "sequence_order": 2,
                    "content_ids": ["S02-B01"],
                },
                {
                    "section_title": "Section A",
                    "section_path": ["Chapitre", "Section A"],
                    "sequence_order": 1,
                    "content_ids": ["S01-B01"],
                },
            ],
        }
    )

    assert [row["section_title"] for row in rows] == ["Section A", "Section B"]
    assert rows[0]["source_block_count"] == 1
    assert rows[0]["section_path_json"] == ["Chapitre", "Section A"]


def test_section_lesson_endpoint_returns_matching_section_content(client):
    headers = _auth_headers(client)
    class_resp = client.post(
        "/classes",
        json={"name": f"SectionLesson {uuid.uuid4().hex[:6]}", "subject": "Math"},
        headers=headers,
    )
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    start_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "chapter", "title": "Section Lesson Unit", "source_text": "Produit et division\nExemples\nExercices"},
    )
    assert start_resp.status_code == 201
    unit_id = start_resp.json()["id"]

    from app.database import SessionLocal
    from app.models import WorkflowUnitBlueprint

    with SessionLocal() as db_session:
        blueprint = db_session.query(WorkflowUnitBlueprint).filter(WorkflowUnitBlueprint.unit_id == unit_id).one()
        blueprint.content_blocks_json = [
            {
                "section_title": "Produit et division",
                "section_path": ["Rationnels", "Produit et division"],
                "title": "Exemple de produit",
                "kind": "example",
                "teaching_material": "Calculer 2/3 x 4/5 = 8/15.",
                "source_excerpt": "Bloc exemple produit.",
            },
            {
                "section_title": "Produit et division",
                "section_path": ["Rationnels", "Produit et division"],
                "title": "Exercices de produit",
                "kind": "exercise",
                "teaching_material": "a) Calculer 3/7 x 2/7. b) Simplifier 5/10 x 1/2.",
                "source_excerpt": "Bloc exercices produit.",
            },
            {
                "section_title": "Somme et difference",
                "section_path": ["Rationnels", "Somme et difference"],
                "title": "Exemple de somme",
                "kind": "example",
                "teaching_material": "Calculer 3/7 + 2/7 = 5/7.",
                "source_excerpt": "Bloc exemple somme.",
            },
        ]
        db_session.commit()

    index_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/sections/index",
        headers=headers,
    )
    assert index_resp.status_code == 200
    assert len(index_resp.json()) == 2

    prepare_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/sections/prepare",
        headers=headers,
        json={"section_path": ["Rationnels", "Produit et division"]},
    )
    assert prepare_resp.status_code == 200
    assert prepare_resp.json()["status"] == "prepared"

    resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/section-lesson",
        headers=headers,
        json={
            "section_path": ["Rationnels", "Produit et division"],
            "item_path": ["Rationnels", "Produit et division", "Exemple de produit"],
            "item_title": "Exemple de produit",
        },
    )
    assert resp.status_code == 200
    payload = resp.json()
    assert payload["section_title"] == "Produit et division"
    combined = "\n".join(row["content_md"] for row in payload["source_blocks"])
    assert "2/3 x 4/5" in combined
    assert "3/7 x 2/7" in combined
    assert "3/7 + 2/7" not in combined


def test_section_lesson_falls_back_to_blueprint_content_when_unprepared(client):
    headers = _auth_headers(client)
    class_resp = client.post(
        "/classes",
        json={"name": f"SectionPrepGate {uuid.uuid4().hex[:6]}", "subject": "Math"},
        headers=headers,
    )
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    start_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "chapter", "title": "Section Gate Unit", "source_text": "Produit et division\nExemples\nExercices"},
    )
    assert start_resp.status_code == 201
    unit_id = start_resp.json()["id"]

    from app.database import SessionLocal
    from app.models import WorkflowUnitBlueprint

    with SessionLocal() as db_session:
        blueprint = db_session.query(WorkflowUnitBlueprint).filter(WorkflowUnitBlueprint.unit_id == unit_id).one()
        blueprint.content_blocks_json = [
            {
                "section_title": "Produit et division",
                "section_path": ["Rationnels", "Produit et division"],
                "title": "Exemple de produit",
                "kind": "example",
                "teaching_material": "Calculer 2/3 x 4/5 = 8/15.",
                "source_excerpt": "Bloc exemple produit.",
            },
        ]
        db_session.commit()

    resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/section-lesson",
        headers=headers,
        json={"section_path": ["Rationnels", "Produit et division"]},
    )
    assert resp.status_code == 200
    payload = resp.json()
    assert payload["section_title"] == "Produit et division"
    combined = "\n".join(row["content_md"] for row in payload["source_blocks"])
    assert "2/3 x 4/5" in combined


def test_prepare_unit_section_stores_persisted_section_record(client):
    headers = _auth_headers(client)
    class_resp = client.post(
        "/classes",
        json={"name": f"PreparedSection {uuid.uuid4().hex[:6]}", "subject": "Math"},
        headers=headers,
    )
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    start_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "chapter", "title": "Prepared Section Unit", "source_text": "Produit et division\nExemples\nExercices"},
    )
    assert start_resp.status_code == 201
    unit_id = start_resp.json()["id"]

    from app.database import SessionLocal
    from app.models import WorkflowUnitBlueprint

    with SessionLocal() as db_session:
        blueprint = db_session.query(WorkflowUnitBlueprint).filter(WorkflowUnitBlueprint.unit_id == unit_id).one()
        blueprint.content_blocks_json = [
            {
                "section_title": "Produit et division",
                "section_path": ["Rationnels", "Produit et division"],
                "title": "Exemple de produit",
                "kind": "example",
                "teaching_material": "Calculer 2/3 x 4/5 = 8/15.",
                "source_excerpt": "Bloc exemple produit.",
            },
            {
                "section_title": "Produit et division",
                "section_path": ["Rationnels", "Produit et division"],
                "title": "Exercices de produit",
                "kind": "exercise",
                "teaching_material": "Calculer 3/7 x 2/7.",
                "source_excerpt": "Bloc exercices produit.",
            },
        ]
        db_session.commit()

    index_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/sections/index",
        headers=headers,
    )
    assert index_resp.status_code == 200
    rows = index_resp.json()
    assert len(rows) == 1
    section_key = rows[0]["section_key"]

    prepare_resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/sections/prepare",
        headers=headers,
        json={"section_path": ["Rationnels", "Produit et division"]},
    )
    assert prepare_resp.status_code == 200
    prepared = prepare_resp.json()
    assert prepared["status"] == "prepared"
    assert prepared["section_key"] == section_key
    assert prepared["latex_source"]
    assert len(prepared["source_blocks_json"]) >= 2

    get_resp = client.get(
        f"/workflow/classes/{class_id}/units/{unit_id}/sections/{section_key}",
        headers=headers,
    )
    assert get_resp.status_code == 200
    fetched = get_resp.json()
    assert fetched["section_title"] == "Produit et division"
    assert fetched["status"] == "prepared"
    combined = "\n".join(row["content_md"] for row in fetched["source_blocks_json"])
    assert "2/3 x 4/5" in combined
    assert "3/7 x 2/7" in combined


def test_section_lesson_prefers_prepared_section_record_when_available(client):
    headers = _auth_headers(client)
    class_resp = client.post(
        "/classes",
        json={"name": f"PreparedLessonFirst {uuid.uuid4().hex[:6]}", "subject": "Math"},
        headers=headers,
    )
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    start_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "chapter", "title": "Prepared Wins Unit", "source_text": "Produit et division\nExemples\nExercices"},
    )
    assert start_resp.status_code == 201
    unit_id = start_resp.json()["id"]

    from app.database import SessionLocal
    from app.models import WorkflowPreparedSection, WorkflowUnitBlueprint
    from app.services.workflow_generation import build_section_key

    with SessionLocal() as db_session:
        blueprint = db_session.query(WorkflowUnitBlueprint).filter(WorkflowUnitBlueprint.unit_id == unit_id).one()
        blueprint.content_blocks_json = [
            {
                "section_title": "Produit et division",
                "section_path": ["Rationnels", "Produit et division"],
                "title": "Source fallback",
                "kind": "example",
                "teaching_material": "Fallback text from blueprint.",
                "source_excerpt": "Fallback text from blueprint.",
            },
        ]
        db_session.add(
            WorkflowPreparedSection(
                unit_id=unit_id,
                section_key=build_section_key(["Rationnels", "Produit et division"], fallback_title="Produit et division"),
                section_title="Produit et division",
                section_path_json=["Rationnels", "Produit et division"],
                order_index=0,
                source_blocks_json=[
                    {
                        "title": "Prepared block",
                        "kind": "example",
                        "kind_label": "Example",
                        "teaching_phase": "example",
                        "content_md": "Prepared text should win.",
                        "content_source": "source_extract",
                    }
                ],
                source_excerpt_md="Prepared text should win.",
                latex_source="Prepared latex",
                provider="notebooklm",
                status="prepared",
                benchmark_status="pending",
            )
        )
        db_session.commit()

    resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/section-lesson",
        headers=headers,
        json={"section_path": ["Rationnels", "Produit et division"]},
    )
    assert resp.status_code == 200
    payload = resp.json()
    combined = "\n".join(row["content_md"] for row in payload["source_blocks"])
    assert "Prepared text should win." in combined
    assert "Fallback text from blueprint." not in combined


def test_leaf_content_generate_requires_blueprint(client):
    headers = _auth_headers(client)
    class_resp = client.post(
        "/classes",
        json={"name": f"NoBlueprintGen {uuid.uuid4().hex[:6]}", "subject": "Math"},
        headers=headers,
    )
    assert class_resp.status_code == 201
    class_id = class_resp.json()["id"]

    start_resp = client.post(
        f"/workflow/classes/{class_id}/units/start",
        headers=headers,
        data={"unit_type": "chapter", "title": "No Blueprint Unit", "source_text": "Section 1\n1.1 Propriete\n1.2 Exemple"},
    )
    assert start_resp.status_code == 201
    unit_id = start_resp.json()["id"]

    workspace_resp = client.get(f"/workflow/classes/{class_id}", headers=headers)
    checklist = workspace_resp.json()["active_unit"]["checklist"]
    leaf = _first_leaf_checklist_item(checklist)
    assert leaf is not None
    item_id = leaf["id"]

    # Delete the blueprint that was automatically created during unit start
    from app.database import SessionLocal
    from app.models import WorkflowUnitBlueprint
    from sqlalchemy import delete as sa_delete

    with SessionLocal() as db_session:
        db_session.execute(sa_delete(WorkflowUnitBlueprint).where(WorkflowUnitBlueprint.unit_id == unit_id))
        db_session.commit()

    resp = client.post(
        f"/workflow/classes/{class_id}/units/{unit_id}/leaf-content/{item_id}/generate",
        headers=headers,
        json={},
    )
    assert resp.status_code == 409
    assert "blueprint" in resp.json()["detail"].lower()
