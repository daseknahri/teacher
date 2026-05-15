from __future__ import annotations

import csv
from io import BytesIO, StringIO
import re
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook


WEEKDAY_LABELS = {
    1: "Monday",
    2: "Tuesday",
    3: "Wednesday",
    4: "Thursday",
    5: "Friday",
    6: "Saturday",
    7: "Sunday",
}

WEEKDAY_MAP = {
    "1": 1,
    "monday": 1,
    "mon": 1,
    "lundi": 1,
    "2": 2,
    "tuesday": 2,
    "tue": 2,
    "mardi": 2,
    "3": 3,
    "wednesday": 3,
    "wed": 3,
    "mercredi": 3,
    "4": 4,
    "thursday": 4,
    "thu": 4,
    "jeudi": 4,
    "5": 5,
    "friday": 5,
    "fri": 5,
    "vendredi": 5,
    "6": 6,
    "saturday": 6,
    "sat": 6,
    "samedi": 6,
    "7": 7,
    "sunday": 7,
    "sun": 7,
    "dimanche": 7,
}

HEADER_ALIASES = {
    "teacher_key": {"teacherkey", "teacher", "teacheremail", "teacherid", "teachercode"},
    "class_name": {"classname", "class", "classroom", "classgroup", "groupclass"},
    "subject": {"subject", "matiere", "module"},
    "weekday": {"weekday", "day", "jour", "week"},
    "start_time": {"starttime", "start", "timefrom", "from", "begin"},
    "end_time": {"endtime", "end", "timeto", "to", "finish"},
    "room": {"room", "salle"},
    "group": {"group", "section", "grp"},
}

TIME_PATTERN = re.compile(r"^\s*(\d{1,2}):(\d{2})(?::(\d{2}))?\s*$")
ICS_DATETIME_PATTERN = re.compile(r"^\s*(\d{8})(?:T(\d{2})(\d{2})(\d{2})?)?(?:Z)?\s*$")
ICS_RRULE_PART_PATTERN = re.compile(r"^\s*([A-Z0-9_-]+)\s*=\s*(.*?)\s*$")
ICS_WEEKDAY_MAP = {
    "MO": 1,
    "TU": 2,
    "WE": 3,
    "TH": 4,
    "FR": 5,
    "SA": 6,
    "SU": 7,
}


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def _map_headers(raw_headers: list[Any]) -> dict[str, int]:
    mapped: dict[str, int] = {}
    for idx, header in enumerate(raw_headers):
        normalized = _normalize_header(header)
        if not normalized:
            continue
        for canonical, aliases in HEADER_ALIASES.items():
            if canonical in mapped:
                continue
            if normalized in aliases:
                mapped[canonical] = idx
                break
    return mapped


def _safe_text(value: Any) -> str | None:
    text = str(value or "").strip()
    return text or None


def _parse_weekday(value: Any) -> tuple[int | None, str | None]:
    text = str(value or "").strip().lower()
    if not text:
        return None, None
    weekday = WEEKDAY_MAP.get(text)
    if weekday is None:
        return None, None
    return weekday, WEEKDAY_LABELS.get(weekday)


def _parse_time(value: Any) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    match = TIME_PATTERN.match(text)
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2))
    second = int(match.group(3) or 0)
    if hour < 0 or hour > 23 or minute < 0 or minute > 59 or second < 0 or second > 59:
        return None
    return f"{hour:02d}:{minute:02d}:{second:02d}"


def _minutes_from_time(text: str | None) -> int | None:
    if not text:
        return None
    match = TIME_PATTERN.match(text)
    if not match:
        return None
    return (int(match.group(1)) * 60) + int(match.group(2))


def _row_value(row: list[Any], mapped: dict[str, int], key: str) -> Any:
    idx = mapped.get(key)
    if idx is None:
        return None
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _normalize_rows(rows: list[list[Any]], mapped_headers: dict[str, int], header_row_index: int) -> list[dict]:
    output: list[dict] = []
    for offset, row in enumerate(rows, start=1):
        row_index = header_row_index + offset
        if not any(str(cell or "").strip() for cell in row):
            continue

        teacher_key = _safe_text(_row_value(row, mapped_headers, "teacher_key"))
        class_name = _safe_text(_row_value(row, mapped_headers, "class_name"))
        subject = _safe_text(_row_value(row, mapped_headers, "subject"))
        weekday_value = _row_value(row, mapped_headers, "weekday")
        start_value = _row_value(row, mapped_headers, "start_time")
        end_value = _row_value(row, mapped_headers, "end_time")
        room = _safe_text(_row_value(row, mapped_headers, "room"))
        group = _safe_text(_row_value(row, mapped_headers, "group"))

        weekday, weekday_label = _parse_weekday(weekday_value)
        start_time = _parse_time(start_value)
        end_time = _parse_time(end_value)

        issues: list[str] = []
        if class_name is None:
            issues.append("Missing class_name.")
        if weekday is None:
            issues.append("Invalid or missing weekday.")
        if start_time is None:
            issues.append("Invalid or missing start_time.")
        if end_time is None:
            issues.append("Invalid or missing end_time.")
        if start_time is not None and end_time is not None:
            start_minutes = _minutes_from_time(start_time)
            end_minutes = _minutes_from_time(end_time)
            if start_minutes is not None and end_minutes is not None and end_minutes <= start_minutes:
                issues.append("end_time must be after start_time.")

        output.append(
            {
                "row_index": row_index,
                "teacher_key": teacher_key,
                "class_name": class_name,
                "subject": subject,
                "weekday": weekday,
                "weekday_label": weekday_label,
                "start_time": start_time,
                "end_time": end_time,
                "room": room,
                "group": group,
                "is_valid": len(issues) == 0,
                "issues": issues,
            }
        )
    return output


def _decode_ics_text(value: str | None) -> str:
    text = str(value or "")
    # RFC5545 escaping used by common calendar exports.
    text = text.replace("\\N", "\n").replace("\\n", "\n")
    text = text.replace("\\,", ",").replace("\\;", ";").replace("\\\\", "\\")
    return text.strip()


def _unfold_ics_lines(text: str) -> list[str]:
    raw_lines = str(text or "").replace("\r\n", "\n").replace("\r", "\n").split("\n")
    unfolded: list[str] = []
    for raw in raw_lines:
        if (raw.startswith(" ") or raw.startswith("\t")) and unfolded:
            unfolded[-1] = f"{unfolded[-1]}{raw[1:]}"
        else:
            unfolded.append(raw)
    return [line.strip() for line in unfolded if line.strip()]


def _parse_ics_property(line: str) -> tuple[str | None, dict[str, str], str]:
    if ":" not in line:
        return None, {}, ""
    head, value = line.split(":", 1)
    chunks = [chunk.strip() for chunk in head.split(";") if chunk.strip()]
    if not chunks:
        return None, {}, value
    prop_name = chunks[0].upper()
    params: dict[str, str] = {}
    for chunk in chunks[1:]:
        if "=" not in chunk:
            continue
        key, param_value = chunk.split("=", 1)
        params[str(key or "").strip().upper()] = str(param_value or "").strip()
    return prop_name, params, value


def _parse_ics_datetime(value: str | None) -> tuple[int | None, str | None]:
    text = str(value or "").strip()
    if not text:
        return None, None
    match = ICS_DATETIME_PATTERN.match(text)
    if not match:
        return None, None
    date_chunk = str(match.group(1) or "").strip()
    hour_chunk = match.group(2)
    minute_chunk = match.group(3)
    second_chunk = match.group(4)
    if len(date_chunk) != 8:
        return None, None

    try:
        yyyy = int(date_chunk[0:4])
        mm = int(date_chunk[4:6])
        dd = int(date_chunk[6:8])
        # Sakamoto-style weekday, converted to ISO Monday=1..Sunday=7.
        month_table = [0, 3, 2, 5, 0, 3, 5, 1, 4, 6, 2, 4]
        year = yyyy - (1 if mm < 3 else 0)
        weekday = (year + year // 4 - year // 100 + year // 400 + month_table[mm - 1] + dd) % 7
        iso_weekday = 7 if weekday == 0 else weekday
    except Exception:
        iso_weekday = None

    if hour_chunk is None or minute_chunk is None:
        return iso_weekday, None
    try:
        hh = int(hour_chunk)
        minute = int(minute_chunk)
        second = int(second_chunk or 0)
    except Exception:
        return iso_weekday, None
    if hh < 0 or hh > 23 or minute < 0 or minute > 59 or second < 0 or second > 59:
        return iso_weekday, None
    return iso_weekday, f"{hh:02d}:{minute:02d}:{second:02d}"


def _parse_ics_rrule_byday(value: str | None) -> list[int]:
    text = str(value or "").strip()
    if not text:
        return []
    by_day_raw = ""
    for part in text.split(";"):
        match = ICS_RRULE_PART_PATTERN.match(part)
        if not match:
            continue
        key = str(match.group(1) or "").strip().upper()
        parsed_value = str(match.group(2) or "").strip()
        if key == "BYDAY":
            by_day_raw = parsed_value
            break
    if not by_day_raw:
        return []
    days: list[int] = []
    for token in by_day_raw.split(","):
        token_text = str(token or "").strip().upper()
        if not token_text:
            continue
        weekday_key = token_text[-2:]
        day_value = ICS_WEEKDAY_MAP.get(weekday_key)
        if day_value is None:
            continue
        days.append(day_value)
    return sorted(set(days))


def _parse_ics_description_fields(description: str | None) -> dict[str, str]:
    text = _decode_ics_text(description)
    if not text:
        return {}
    output: dict[str, str] = {}
    for line in text.splitlines():
        line_text = str(line or "").strip()
        if not line_text or ":" not in line_text:
            continue
        key, value = line_text.split(":", 1)
        key_norm = str(key or "").strip().lower().replace(" ", "").replace("-", "").replace("_", "")
        parsed = str(value or "").strip()
        if not parsed:
            continue
        if key_norm in {"class", "classname", "classe", "classgroup"}:
            output["class_name"] = parsed
        elif key_norm in {"subject", "matiere", "module"}:
            output["subject"] = parsed
        elif key_norm in {"group", "grp", "section"}:
            output["group"] = parsed
        elif key_norm in {"teacher", "teacherkey", "teacheremail"}:
            output["teacher_key"] = parsed
    return output


def _split_ics_summary(summary: str | None) -> tuple[str | None, str | None]:
    text = _decode_ics_text(summary)
    if not text:
        return None, None
    for separator in ("|", " - ", " / "):
        if separator not in text:
            continue
        left, right = [chunk.strip() for chunk in text.split(separator, 1)]
        if left and right:
            return left, right
    return text, None


def parse_timetable_csv_preview(content: bytes) -> list[dict]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = content.decode("latin-1")
        except UnicodeDecodeError as exc:
            raise HTTPException(status_code=400, detail="Unable to decode CSV file.") from exc

    sample = text[:4096]
    delimiter = ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        delimiter = dialect.delimiter
    except Exception:
        delimiter = "," if text.count(",") >= text.count(";") else ";"

    reader = csv.reader(StringIO(text), delimiter=delimiter)
    all_rows = [list(row) for row in reader]
    if not all_rows:
        raise HTTPException(status_code=400, detail="CSV file is empty.")
    headers = all_rows[0]
    mapped_headers = _map_headers(headers)
    required = {"class_name", "weekday", "start_time", "end_time"}
    missing_required = sorted(required - set(mapped_headers.keys()))
    if missing_required:
        raise HTTPException(status_code=400, detail=f"Missing required timetable columns: {missing_required}")
    return _normalize_rows(all_rows[1:], mapped_headers, header_row_index=1)


def parse_timetable_xlsx_preview(content: bytes) -> list[dict]:
    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Unable to read spreadsheet file.") from exc

    sheet = workbook.active
    matrix = [list(row) for row in sheet.iter_rows(values_only=True)]
    if not matrix:
        raise HTTPException(status_code=400, detail="Spreadsheet is empty.")
    headers = matrix[0]
    mapped_headers = _map_headers(headers)
    required = {"class_name", "weekday", "start_time", "end_time"}
    missing_required = sorted(required - set(mapped_headers.keys()))
    if missing_required:
        raise HTTPException(status_code=400, detail=f"Missing required timetable columns: {missing_required}")
    return _normalize_rows(matrix[1:], mapped_headers, header_row_index=1)


def parse_timetable_ics_preview(content: bytes) -> list[dict]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = content.decode("latin-1")
        except UnicodeDecodeError as exc:
            raise HTTPException(status_code=400, detail="Unable to decode ICS file.") from exc

    lines = _unfold_ics_lines(text)
    events: list[list[str]] = []
    in_event = False
    current_lines: list[str] = []
    for line in lines:
        upper_line = str(line or "").upper()
        if upper_line == "BEGIN:VEVENT":
            in_event = True
            current_lines = []
            continue
        if upper_line == "END:VEVENT":
            if in_event and current_lines:
                events.append(current_lines)
            in_event = False
            current_lines = []
            continue
        if in_event:
            current_lines.append(line)

    if not events:
        raise HTTPException(status_code=400, detail="ICS file has no VEVENT rows.")

    output: list[dict] = []
    row_index = 2
    for event in events:
        props: dict[str, list[str]] = {}
        for line in event:
            prop_name, _, value = _parse_ics_property(line)
            if not prop_name:
                continue
            props.setdefault(prop_name, []).append(value)

        summary = props.get("SUMMARY", [None])[0]
        description = props.get("DESCRIPTION", [None])[0]
        location = props.get("LOCATION", [None])[0]
        categories = props.get("CATEGORIES", [None])[0]
        rrule = props.get("RRULE", [None])[0]
        dtstart = props.get("DTSTART", [None])[0]
        dtend = props.get("DTEND", [None])[0]

        desc_fields = _parse_ics_description_fields(description)
        summary_class_name, summary_subject = _split_ics_summary(summary)
        class_name = (
            desc_fields.get("class_name")
            or summary_class_name
            or _decode_ics_text(summary)
            or None
        )
        subject = (
            desc_fields.get("subject")
            or summary_subject
            or _decode_ics_text(categories)
            or None
        )
        teacher_key = desc_fields.get("teacher_key") or None
        group = desc_fields.get("group") or None
        room = _decode_ics_text(location) or None

        dtstart_weekday, start_time = _parse_ics_datetime(dtstart)
        _, end_time = _parse_ics_datetime(dtend)
        rrule_weekdays = _parse_ics_rrule_byday(rrule)
        weekdays = rrule_weekdays or ([dtstart_weekday] if dtstart_weekday else [])

        base_issues: list[str] = []
        if class_name is None:
            base_issues.append("Missing class_name.")
        if not weekdays:
            base_issues.append("Invalid or missing weekday.")
        if start_time is None:
            base_issues.append("Invalid or missing start_time.")
        if end_time is None:
            base_issues.append("Invalid or missing end_time.")
        if start_time is not None and end_time is not None:
            start_minutes = _minutes_from_time(start_time)
            end_minutes = _minutes_from_time(end_time)
            if start_minutes is not None and end_minutes is not None and end_minutes <= start_minutes:
                base_issues.append("end_time must be after start_time.")

        if not weekdays:
            output.append(
                {
                    "row_index": row_index,
                    "teacher_key": teacher_key,
                    "class_name": class_name,
                    "subject": subject,
                    "weekday": None,
                    "weekday_label": None,
                    "start_time": start_time,
                    "end_time": end_time,
                    "room": room,
                    "group": group,
                    "is_valid": len(base_issues) == 0,
                    "issues": base_issues,
                }
            )
            row_index += 1
            continue

        for weekday in weekdays:
            issues = list(base_issues)
            output.append(
                {
                    "row_index": row_index,
                    "teacher_key": teacher_key,
                    "class_name": class_name,
                    "subject": subject,
                    "weekday": int(weekday),
                    "weekday_label": WEEKDAY_LABELS.get(int(weekday)),
                    "start_time": start_time,
                    "end_time": end_time,
                    "room": room,
                    "group": group,
                    "is_valid": len(issues) == 0,
                    "issues": issues,
                }
            )
            row_index += 1

    return output
