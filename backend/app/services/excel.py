from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
import os
import re

from openpyxl import Workbook, load_workbook

from .holidays import MOROCCO_ACADEMIC_HOLIDAY_RANGES


ROSTER_COLUMNS = ("student_code", "full_name")
EXAM_COLUMNS = ("student_code", "full_name", "score", "note", "teacher_comment")
EXAM_LIST_COLUMNS = ("id", "name", "birth_date", "note_1", "note_2", "note_3", "note")
HOLIDAY_COLUMNS = ("holiday", "start_date", "end_date", "is_blocked", "notes")
HOLIDAY_EXPORT_COLUMNS = ("holiday", "start_date", "end_date", "number_of_days", "is_blocked", "source")


def parse_roster_excel(content: bytes) -> tuple[list[dict], list[str]]:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    notes_sheet = workbook["NotesCC"] if "NotesCC" in workbook.sheetnames else None
    if notes_sheet is not None and _looks_like_notescc_list(notes_sheet):
        return _parse_notescc_roster_template(notes_sheet)

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    return _parse_normalized_roster(rows)


def _parse_normalized_roster(rows: list[tuple]) -> tuple[list[dict], list[str]]:
    if not rows:
        return [], ["Excel file is empty."]

    header = [_normalize_header(v) for v in rows[0]]
    idx_code = _find_index(header, ("student_code", "massar_code", "student_number", "code", "cne", "apogee"))
    idx_fallback_id = _find_index(header, ("id", "student_id"))
    idx_name = _find_index(header, ("full_name", "name", "student_name"))
    idx_external_id = _find_index(header, ("external_id", "id", "student_id"))
    idx_birth_date = _find_index(header, ("birth_date", "date_of_birth", "dob", "birthdate", "birth"))

    code_from_id = False
    if idx_code is None:
        idx_code = idx_fallback_id
        code_from_id = idx_code is not None

    if idx_code is None or idx_name is None:
        return [], [
            "Missing columns. Expected at least student identifier and name "
            "(e.g. student_code/full_name or id/name)."
        ]

    parsed: list[dict] = []
    errors: list[str] = []
    seen: set[str] = set()

    for i, row in enumerate(rows[1:], start=2):
        code = _clean_text(_row_value(row, idx_code))
        name = _clean_text(_row_value(row, idx_name))
        external_id = _clean_text(_row_value(row, idx_external_id)) if idx_external_id is not None else None
        birth_date = _parse_date(_row_value(row, idx_birth_date)) if idx_birth_date is not None else None

        if not code and not name and not external_id:
            continue
        if not code and external_id:
            code = external_id
        if not code or not name:
            errors.append(f"Row {i}: student id/code and name are required.")
            continue
        if code in seen:
            errors.append(f"Row {i}: duplicate student_code '{code}' in file.")
            continue
        seen.add(code)
        if code_from_id and not external_id:
            external_id = code
        parsed.append(
            {
                "student_code": code,
                "external_id": external_id or None,
                "full_name": name,
                "birth_date": birth_date,
            }
        )
    return parsed, errors


def _looks_like_notescc_list(sheet) -> bool:
    marker = _clean_text(sheet.cell(16, 2).value).lower()
    return marker == "id"


def _parse_notescc_roster_template(sheet) -> tuple[list[dict], list[str]]:
    parsed: list[dict] = []
    errors: list[str] = []
    seen: set[str] = set()
    empty_streak = 0

    for row_idx in range(18, sheet.max_row + 1):
        external_id = _clean_text(sheet.cell(row_idx, 2).value)
        code = _clean_text(sheet.cell(row_idx, 3).value)
        name = _clean_text(sheet.cell(row_idx, 4).value)
        birth_date = _parse_date(sheet.cell(row_idx, 6).value, day_first=True)

        if not external_id and not code and not name:
            empty_streak += 1
            if empty_streak >= 5:
                break
            continue
        empty_streak = 0

        code = code or external_id
        if not code or not name:
            errors.append(f"Row {row_idx}: id/code and name are required.")
            continue
        if code in seen:
            errors.append(f"Row {row_idx}: duplicate student_code '{code}' in file.")
            continue
        seen.add(code)

        parsed.append(
            {
                "student_code": code,
                "external_id": external_id or None,
                "full_name": name,
                "birth_date": birth_date,
            }
        )

    if not parsed and not errors:
        errors.append("No student rows found in the provided NotesCC list.")
    return parsed, errors


def build_exam_template(students: list[dict]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "exam_notes"
    sheet.append(EXAM_LIST_COLUMNS)
    for student in students:
        birth_date = student.get("birth_date")
        birth_value = birth_date.isoformat() if isinstance(birth_date, date) else None
        sheet.append(
            [
                student.get("external_id") or student.get("student_code"),
                student.get("full_name"),
                birth_value,
                None,
                None,
                None,
                None,
            ]
        )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_exam_template_notescc(
    students: list[dict], exam_title: str, class_name: str | None = None, subject: str | None = None
) -> bytes:
    template_path = _discover_template_path()
    if not template_path:
        # If NotesCC template is unavailable, return normalized template.
        return build_exam_template(students)

    rows = [
        {
            "student_code": student["student_code"],
            "external_id": student.get("external_id"),
            "full_name": student["full_name"],
            "birth_date": student.get("birth_date"),
            "scores": [None],
            "score": None,
            "note": None,
            "teacher_comment": None,
        }
        for student in students
    ]
    return _build_principal_export_from_template(
        template_path=template_path,
        exam_title=exam_title,
        rows=rows,
        class_name=class_name,
        subject=subject,
    )


def parse_exam_results_excel(content: bytes) -> tuple[list[dict], list[str]]:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    sheet = workbook["NotesCC"] if "NotesCC" in workbook.sheetnames else workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], ["Excel file is empty."]

    header = [_normalize_header(v) for v in rows[0]]
    if all(col in header for col in EXAM_COLUMNS):
        return _parse_normalized_exam(rows)
    if _looks_like_exam_list_header(header):
        return _parse_exam_list(rows)
    return _parse_notescc_exam_template(sheet)


def parse_holiday_excel(content: bytes) -> tuple[list[dict], list[str]]:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], ["Excel file is empty."]

    header = [_normalize_header(v) for v in rows[0]]
    idx_name = _find_index(header, ("holiday", "holiday_name", "name", "title"))
    idx_dates = _find_index(header, ("dates", "date_range", "holiday_dates"))
    idx_start = _find_index(header, ("start_date", "date_from", "from_date", "start"))
    idx_end = _find_index(header, ("end_date", "date_to", "to_date", "end"))
    idx_blocked = _find_index(header, ("is_blocked", "blocked", "block"))

    if idx_name is None or (idx_dates is None and idx_start is None):
        return [], [
            "Missing columns. Expected holiday/name and either dates or start_date/end_date.",
        ]

    parsed: list[dict] = []
    errors: list[str] = []

    for row_index, row in enumerate(rows[1:], start=2):
        name = _clean_text(_row_value(row, idx_name))
        dates_value = _row_value(row, idx_dates)
        start_date = _parse_date(_row_value(row, idx_start), day_first=True) if idx_start is not None else None
        end_date = _parse_date(_row_value(row, idx_end), day_first=True) if idx_end is not None else None
        is_blocked, blocked_error = _parse_bool_value(_row_value(row, idx_blocked), default=True)

        if not name and not _has_value(dates_value) and start_date is None and end_date is None:
            continue
        if not name:
            errors.append(f"Row {row_index}: holiday name is required.")
            continue
        if blocked_error is not None:
            errors.append(f"Row {row_index}: {blocked_error}")
            continue

        if start_date is None and end_date is None and _has_value(dates_value):
            start_date, end_date = _parse_holiday_date_range(dates_value)
        elif start_date is not None and end_date is None:
            end_date = start_date
        elif start_date is None and end_date is not None:
            start_date = end_date

        if start_date is None or end_date is None:
            errors.append(f"Row {row_index}: provide a valid date or date range.")
            continue
        if end_date < start_date:
            errors.append(f"Row {row_index}: end_date must be on or after start_date.")
            continue

        parsed.append(
            {
                "name": name,
                "start_date": start_date,
                "end_date": end_date,
                "is_blocked": bool(is_blocked),
            }
        )

    return parsed, errors


def build_holiday_import_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "holiday_import"
    sheet.append(HOLIDAY_COLUMNS)
    for name, start_date, end_date in [
        (name, start_day, end_day)
        for start_day, end_day, name in MOROCCO_ACADEMIC_HOLIDAY_RANGES
    ]:
        sheet.append(
            (
                name,
                start_date.isoformat(),
                end_date.isoformat(),
                True,
                "",
            )
        )

    instructions = workbook.create_sheet("instructions")
    instructions.append(("How to use",))
    instructions.append(("Keep one holiday row per line.",))
    instructions.append(("Use the holiday_import sheet columns: holiday, start_date, end_date, is_blocked.",))
    instructions.append(("Dates can be real Excel dates, YYYY-MM-DD text, or a single Dates column like 'September 4 - September 5, 2025'.",))
    instructions.append(("The upload replaces imported school-year holiday rows for the years included in the file.",))
    instructions.append(("Fixed Morocco public holidays remain available automatically.",))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_holiday_export_workbook(rows: list[object]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "holiday_export"
    sheet.append(HOLIDAY_EXPORT_COLUMNS)

    for row in _group_holiday_export_rows(rows):
        day_count = int((row["end_date"] - row["start_date"]).days) + 1
        sheet.append(
            (
                row["name"],
                row["start_date"].isoformat(),
                row["end_date"].isoformat(),
                day_count,
                bool(row["is_blocked"]),
                row["source"] or "",
            )
        )

    instructions = workbook.create_sheet("instructions")
    instructions.append(("How to use",))
    instructions.append(("This file contains the current holiday rows for the selected year.",))
    instructions.append(("Add or edit rows, then upload the workbook again from Owner -> Morocco Holidays.",))
    instructions.append(("Uploaded rows merge into the standard calendar for the dates you provide.",))
    instructions.append(("For imports, the required columns remain holiday, start_date, end_date, and optionally is_blocked.",))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _looks_like_exam_list_header(header: list[str]) -> bool:
    has_id = "id" in header or "student_code" in header
    has_name = "name" in header or "full_name" in header
    has_note_column = any(col in header for col in ("note", "score", "note_1", "note_2", "note_3"))
    return has_id and has_name and has_note_column


def _parse_exam_list(rows: list[tuple]) -> tuple[list[dict], list[str]]:
    header = [_normalize_header(v) for v in rows[0]]
    idx_code = _find_index(header, ("student_code", "id", "student_id", "code", "cne", "apogee"))
    idx_name = _find_index(header, ("full_name", "name", "student_name"))
    idx_note_1 = _find_index(header, ("note_1", "note1", "score_1", "score1", "first_note"))
    idx_note_2 = _find_index(header, ("note_2", "note2", "score_2", "score2", "second_note"))
    idx_note_3 = _find_index(header, ("note_3", "note3", "score_3", "score3", "third_note"))
    idx_note = _find_index(header, ("note", "score", "final_note", "finalscore"))

    if idx_code is None:
        return [], ["Missing student identifier column (id/student_code)."]
    if idx_note is None and idx_note_1 is None and idx_note_2 is None and idx_note_3 is None:
        return [], ["Missing score/note columns (note or note_1/note_2/note_3)."]

    parsed: list[dict] = []
    errors: list[str] = []
    seen: set[str] = set()

    for i, row in enumerate(rows[1:], start=2):
        code = _clean_text(_row_value(row, idx_code))
        name = _clean_text(_row_value(row, idx_name)) if idx_name is not None else ""
        score_candidates = [
            _row_value(row, idx_note),
            _row_value(row, idx_note_3),
            _row_value(row, idx_note_2),
            _row_value(row, idx_note_1),
        ]
        score_raw = next((value for value in score_candidates if _has_value(value)), None)

        if not code and not name and score_raw is None:
            continue
        if not code:
            errors.append(f"Row {i}: student id/code is required.")
            continue
        if code in seen:
            errors.append(f"Row {i}: duplicate student_code '{code}' in file.")
            continue
        seen.add(code)
        if score_raw is None:
            errors.append(f"Row {i}: no score/note value found.")
            continue
        try:
            score = float(score_raw)
        except (TypeError, ValueError):
            errors.append(f"Row {i}: note/score must be numeric.")
            continue

        parsed.append(
            {
                "student_code": code,
                "full_name": name,
                "score": score,
                "note": None,
                "teacher_comment": None,
            }
        )

    return parsed, errors


def _parse_normalized_exam(rows: list[tuple]) -> tuple[list[dict], list[str]]:
    header = [_normalize_header(v) for v in rows[0]]

    idx_code = header.index("student_code")
    idx_name = header.index("full_name")
    idx_score = header.index("score")
    idx_note = header.index("note")
    idx_comment = header.index("teacher_comment")

    parsed: list[dict] = []
    errors: list[str] = []
    seen: set[str] = set()

    for i, row in enumerate(rows[1:], start=2):
        code = _clean_text(row[idx_code])
        name = _clean_text(row[idx_name])
        score_raw = row[idx_score]
        note = None if row[idx_note] is None else str(row[idx_note]).strip()
        teacher_comment = None if row[idx_comment] is None else str(row[idx_comment]).strip()

        if not code and not name and score_raw is None:
            continue
        if not code:
            errors.append(f"Row {i}: student_code is required.")
            continue
        if code in seen:
            errors.append(f"Row {i}: duplicate student_code '{code}' in file.")
            continue
        seen.add(code)

        try:
            score = float(score_raw)
        except (TypeError, ValueError):
            errors.append(f"Row {i}: score must be numeric.")
            continue

        parsed.append(
            {
                "student_code": code,
                "full_name": name,
                "score": score,
                "note": note or None,
                "teacher_comment": teacher_comment or None,
            }
        )

    return parsed, errors


def _parse_notescc_exam_template(sheet) -> tuple[list[dict], list[str]]:
    """
    Parse NotesCC-like format:
    - ID in column B (optional)
    - student code in column C (fallback to B)
    - full name in column D
    - score in first numeric among columns M/K/I/G (13/11/9/7)
    """
    parsed: list[dict] = []
    errors: list[str] = []
    seen: set[str] = set()
    empty_streak = 0

    for row_idx in range(18, sheet.max_row + 1):
        external_id = _clean_text(sheet.cell(row_idx, 2).value)
        code = _clean_text(sheet.cell(row_idx, 3).value) or external_id
        name = _clean_text(sheet.cell(row_idx, 4).value)

        if not code and not name:
            empty_streak += 1
            if empty_streak >= 5:
                break
            continue
        empty_streak = 0

        if not code:
            errors.append(f"Row {row_idx}: student_code is required.")
            continue
        if code in seen:
            errors.append(f"Row {row_idx}: duplicate student_code '{code}' in file.")
            continue
        seen.add(code)

        score = None
        text_note = None
        for col_idx in (13, 11, 9, 7):
            raw = sheet.cell(row_idx, col_idx).value
            if not _has_value(raw):
                continue
            try:
                score = float(raw)
                break
            except (TypeError, ValueError):
                if col_idx == 13:
                    text_note = str(raw).strip()
                continue
        if score is None:
            errors.append(f"Row {row_idx}: no numeric score found in columns M/K/I/G.")
            continue

        parsed.append(
            {
                "student_code": code,
                "full_name": name,
                "score": score,
                "note": text_note,
                "teacher_comment": None,
            }
        )

    if not parsed and not errors:
        errors.append("No exam rows found in NotesCC template.")
    return parsed, errors


def build_principal_export(exam_title: str, rows: list[dict], class_name: str | None = None, subject: str | None = None) -> bytes:
    template_path = _discover_template_path()
    if template_path:
        try:
            return _build_principal_export_from_template(template_path, exam_title, rows, class_name, subject)
        except Exception:
            # Fallback to normalized export if template rendering fails.
            pass

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "principal_export"
    sheet.append(("exam_title", exam_title))
    sheet.append(EXAM_LIST_COLUMNS)
    for row in rows:
        scores = row.get("scores") or []
        note_1 = scores[0] if len(scores) > 0 else None
        note_2 = scores[1] if len(scores) > 1 else None
        note_3 = scores[2] if len(scores) > 2 else None
        final_note = row.get("score")
        if final_note is None:
            final_note = next((value for value in (note_3, note_2, note_1) if value is not None), None)
        sheet.append(
            [
                row.get("external_id") or row.get("student_code"),
                row.get("full_name"),
                _format_birth_date(row.get("birth_date")),
                note_1,
                note_2,
                note_3,
                final_note,
            ]
        )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_principal_export_from_template(
    template_path: str, exam_title: str, rows: list[dict], class_name: str | None, subject: str | None
) -> bytes:
    workbook = load_workbook(template_path)
    sheet = workbook["NotesCC"] if "NotesCC" in workbook.sheetnames else workbook.active

    # Common metadata cells in the provided NotesCC sample template.
    if class_name:
        sheet["I9"] = class_name
    if subject:
        sheet["O11"] = subject
    sheet["G16"] = exam_title

    # Reset student area columns B:O from row 18 onward.
    for row_idx in range(18, max(sheet.max_row, 18) + 1):
        for col_idx in range(2, 16):
            sheet.cell(row_idx, col_idx).value = None

    for idx, row in enumerate(rows, start=18):
        sheet.cell(idx, 2).value = row.get("external_id") or row.get("student_code")
        sheet.cell(idx, 3).value = row.get("student_code")
        sheet.cell(idx, 4).value = row.get("full_name")
        sheet.cell(idx, 6).value = _format_birth_date(row.get("birth_date"))
        scores = row.get("scores") or []
        score_cols = (7, 9, 11)  # G, I, K
        for score_idx, score_col in enumerate(score_cols):
            if score_idx < len(scores):
                sheet.cell(idx, score_col).value = scores[score_idx]
        note_value = row.get("note")
        if note_value is None:
            note_value = row.get("score")
        if note_value is not None:
            sheet.cell(idx, 13).value = note_value
        teacher_comment = row.get("teacher_comment")
        if teacher_comment is not None:
            sheet.cell(idx, 14).value = teacher_comment

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _discover_template_path() -> str | None:
    configured = os.getenv("PRINCIPAL_EXPORT_TEMPLATE")
    if configured and Path(configured).exists():
        return configured

    downloads = Path.home() / "Downloads"
    if not downloads.exists():
        return None
    matches = sorted(downloads.glob("export_notesCC*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if matches:
        return str(matches[0])
    return None


def _normalize_header(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower().replace("-", "_").replace(" ", "_")
    while "__" in text:
        text = text.replace("__", "_")
    return text


def _find_index(header: list[str], aliases: tuple[str, ...]) -> int | None:
    for alias in aliases:
        if alias in header:
            return header.index(alias)
    return None


def _clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _row_value(row: tuple, idx: int | None):
    if idx is None:
        return None
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _has_value(value) -> bool:
    return value is not None and str(value).strip() != ""


def _parse_date(value, *, day_first: bool = False) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    formats = ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%B %d, %Y", "%b %d, %Y"]
    if day_first:
        formats = ["%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d", "%B %d, %Y", "%b %d, %Y"]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_bool_value(value, *, default: bool) -> tuple[bool, str | None]:
    if value is None:
        return bool(default), None
    if isinstance(value, bool):
        return value, None
    if isinstance(value, (int, float)):
        return bool(value), None
    text = str(value).strip().lower()
    if not text:
        return bool(default), None
    if text in {"1", "true", "yes", "y", "blocked", "block"}:
        return True, None
    if text in {"0", "false", "no", "n", "open", "unblocked"}:
        return False, None
    return bool(default), "is_blocked must be true/false, yes/no, or 1/0."


def _parse_holiday_date_range(value) -> tuple[date | None, date | None]:
    single_date = _parse_date(value, day_first=True)
    if single_date is not None:
        return single_date, single_date

    text = str(value or "").strip()
    if not text:
        return None, None
    normalized = re.sub(r"\s+", " ", text.replace("–", "-").replace("—", "-")).strip()

    month_range = re.match(
        r"^(?P<month1>[A-Za-z]+)\s+(?P<day1>\d{1,2})\s*-\s*(?P<month2>[A-Za-z]+)\s+(?P<day2>\d{1,2}),\s*(?P<year>\d{4})$",
        normalized,
    )
    if month_range:
        year = int(month_range.group("year"))
        start_value = f"{month_range.group('month1')} {int(month_range.group('day1'))}, {year}"
        end_value = f"{month_range.group('month2')} {int(month_range.group('day2'))}, {year}"
        return _parse_date(start_value), _parse_date(end_value)

    split_match = re.split(r"\s+(?:-|to)\s+", normalized, maxsplit=1)
    if len(split_match) == 2:
        start_value = _parse_date(split_match[0], day_first=True)
        end_value = _parse_date(split_match[1], day_first=True)
        if start_value is not None and end_value is not None:
            return start_value, end_value
    return None, None


def _group_holiday_export_rows(rows: list[object]) -> list[dict]:
    normalized: list[dict] = []
    for row in rows:
        holiday_date = getattr(row, "holiday_date", None)
        name = getattr(row, "name", None)
        if holiday_date is None or not name:
            continue
        normalized.append(
            {
                "holiday_date": holiday_date,
                "name": str(name),
                "is_blocked": bool(getattr(row, "is_blocked", True)),
                "source": str(getattr(row, "source", "") or ""),
            }
        )

    normalized.sort(key=lambda item: item["holiday_date"])
    grouped: list[dict] = []
    for row in normalized:
        if grouped:
            previous = grouped[-1]
            if (
                row["holiday_date"] == previous["end_date"] + timedelta(days=1)
                and row["name"] == previous["name"]
                and row["is_blocked"] == previous["is_blocked"]
                and row["source"] == previous["source"]
            ):
                previous["end_date"] = row["holiday_date"]
                continue
        grouped.append(
            {
                "name": row["name"],
                "start_date": row["holiday_date"],
                "end_date": row["holiday_date"],
                "is_blocked": row["is_blocked"],
                "source": row["source"],
            }
        )
    return grouped


def _format_birth_date(value) -> str | None:
    parsed = _parse_date(value, day_first=True)
    if parsed is None:
        return None
    return parsed.strftime("%d-%m-%Y")
